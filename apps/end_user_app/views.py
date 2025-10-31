# bb_budget_monitoring_system/apps/end_user_app/views.py
from django.conf import settings
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.shortcuts import redirect, get_object_or_404
from apps.admin_panel.models import BudgetAllocation
from .models import PurchaseRequest, PurchaseRequestItems, Budget_Realignment, DepartmentPRE, ActivityDesign, Session, Signatory, CampusApproval, UniversityApproval, PRELineItemBudget, PurchaseRequestAllocation, ActivityDesignAllocations, PREBudgetRealignment, PREDraft, PREDraftSupportingDocument
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from decimal import Decimal
from django.db.models import Sum, Q, F, DecimalField
from django.db.models.functions import Coalesce
from django.contrib.humanize.templatetags.humanize import intcomma
from datetime import datetime
from apps.admin_panel.utils import log_audit_trail
from apps.users.utils import role_required
from .constants import FRIENDLY_LABELS
from docxtpl import DocxTemplate
from openpyxl import load_workbook
import os
import json
from django.db import transaction
from io import BytesIO
import shutil
import xlwings as xw
import tempfile
from apps.budgets.models import ApprovedBudget as NewApprovedBudget, BudgetAllocation as NewBudgetAllocation, DepartmentPRE as NewDepartmentPRE, PurchaseRequest as NewPurchaseRequest, PRELineItem, PurchaseRequestAllocation as NewPurchaseRequestAllocation, PRDraft, PRDraftSupportingDocument, PurchaseRequestSupportingDocument
from .utils.pre_parser import parse_pre_excel
from django.core.files.storage import default_storage
from django.utils import timezone

def generate_pr_number():
    """Generate unique PR number in format PR-YYYY-NNNN"""
    from datetime import datetime
    from apps.budgets.models import PurchaseRequest
    
    year = datetime.now().year
    
    # Get the last PR number for this year
    last_pr = PurchaseRequest.objects.filter(
        pr_number__startswith=f'PR-{year}-'
    ).order_by('-pr_number').first()
    
    if last_pr:
        # Extract the sequence number and increment
        last_sequence = int(last_pr.pr_number.split('-')[-1])
        new_sequence = last_sequence + 1
    else:
        # First PR of the year
        new_sequence = 1
    
    return f'PR-{year}-{new_sequence:04d}'

# Create your views here.
@role_required('end_user', login_url='/')
def user_dashboard(request):
    try:
        # Get the budget allocation of the logged-in user
        budget = BudgetAllocation.objects.filter(department=request.user.department)
        total_budget = sum(a.total_allocated for a in budget)
        total_spent = sum(a.spent for a in budget)
        remaining_balance = total_budget - total_spent
        purchase_requests = PurchaseRequest.objects.filter(requested_by=request.user, pr_status='submitted')
        approved_requests_count = PurchaseRequest.objects.filter(requested_by=request.user, submitted_status='approved').count()
        realignment_requests = PREBudgetRealignment.objects.filter(requested_by=request.user).order_by('-created_at')[:5]
    except BudgetAllocation.DoesNotExist or PurchaseRequest.DoesNotExist or PREBudgetRealignment.DoesNotExist:
        budget = None
        total_budget = None
        remaining_balance = None
        purchase_requests = None
        realignment_requests = None
        approved_requests_count = 0
        
    spent = total_budget - remaining_balance
    if total_budget > 0:
        usage_percentage = (spent / total_budget) * 100
    else:
        usage_percentage = 0
        
    context = {
        "total_budget": total_budget,
        "remaining_balance": remaining_balance,
        'purchase_requests': purchase_requests,
        'approved_requests_count': approved_requests_count,
        "spent": spent,
        "usage_percentage": usage_percentage,
        'realignment_requests': realignment_requests
    }

    return render(request, 'end_user_app/dashboard.html', context)

@role_required('end_user', login_url='/')
def view_budget(request):
    try:
        # Fetch all budget allocations for the logged-in user and include related approved_budget
        budget = BudgetAllocation.objects.select_related('approved_budget').filter(department=request.user.department)

    except BudgetAllocation.DoesNotExist:
        budget = None  # If no budget is assigned to the user
    
    return render(request, 'end_user_app/view_budget.html', {'budget': budget})

@role_required('end_user', login_url='/')
def purchase_request(request):
    """
    Display all Purchase Requests and Activity Designs for the current user
    Uses NEW models from budgets app
    """
    try:
        # ✅ NEW: Get Purchase Requests using NEW model
        purchase_requests = NewPurchaseRequest.objects.filter(
            submitted_by=request.user
        ).select_related(
            'budget_allocation',
            'budget_allocation__approved_budget',
            'source_pre',
            'source_line_item',
            'source_line_item__category'
        ).prefetch_related(
            'supporting_documents',
            'pre_allocations'
        ).order_by('-created_at')
        
        # ✅ Calculate status counts for PRs
        pr_pending_count = purchase_requests.filter(status='Pending').count()
        pr_approved_count = purchase_requests.filter(status='Approved').count()
        pr_rejected_count = purchase_requests.filter(status='Rejected').count()
        
        # ✅ Activity Designs (still using old model - no new version yet)
        activity_designs = ActivityDesign.objects.filter(
            requested_by=request.user
        ).select_related(
            'source_pre', 
            'budget_allocation'
        ).order_by('-created_at')
        
        # ✅ Calculate status counts for ADs
        ad_pending_count = activity_designs.filter(status='Pending').count()
        ad_approved_count = activity_designs.filter(status='Approved').count()
        ad_rejected_count = activity_designs.filter(status='Rejected').count()
        
    except Exception as e:
        # Better error handling
        print(f"Error loading requests: {e}")
        purchase_requests = []
        activity_designs = []
        pr_pending_count = 0
        pr_approved_count = 0
        pr_rejected_count = 0
        ad_pending_count = 0
        ad_approved_count = 0
        ad_rejected_count = 0
    
    # ✅ Pass everything to template
    return render(request, 'end_user_app/purchase_request.html', {
        'purchase_requests': purchase_requests,
        'activity_designs': activity_designs,
        'pr_pending_count': pr_pending_count,
        'pr_approved_count': pr_approved_count,
        'pr_rejected_count': pr_rejected_count,
        'ad_pending_count': ad_pending_count,
        'ad_approved_count': ad_approved_count,
        'ad_rejected_count': ad_rejected_count,
    })

@role_required('end_user', login_url='/')
def user_settings_page(request):
    return render(request, 'end_user_app/settings.html')

@role_required('end_user', login_url='/')
def end_user_logout(request):
    log_audit_trail(
        request=request,
        action='LOGOUT',
        model_name='User',
        detail=f'User {request.user} logged out',
    )
    logout(request)
    return redirect('end_user_login') # redirect to the login page

def build_pre_source_options_v2(user):
    """
    Build available PRE line items as funding source options.
    Directly queries PRELineItem for better performance and reliability.
    
    Args:
        user: User object (request.user)
        
    Returns:
        List of dicts with PRE groups and their available line items
    """
    from collections import defaultdict
    
    # 🔥 DIRECT QUERY: Get all line items from user's approved PREs
    line_items = PRELineItem.objects.filter(
        pre__submitted_by=user,
        pre__status__in=['Approved', 'Partially Approved']  # Include both statuses
    ).select_related(
        'pre',
        'pre__budget_allocation',
        'category',
        'subcategory'
    ).order_by(
        'pre__fiscal_year',
        'category__sort_order',
        'item_name'
    )
    
    # Group line items by PRE
    pre_groups = defaultdict(lambda: {
        'line_items': [],
        'pre_display': '',
        'pre_id': ''
    })
    
    for line_item in line_items:
        # Calculate total budget for this line item
        total = line_item.get_total()
        
        # Skip if no budget allocated
        if total <= 0:
            continue
        
        # Calculate consumed amount
        consumed = calculate_line_item_consumed(line_item)
        available = total - consumed
        
        # Only include line items with available budget
        if available > 0:
            pre_key = str(line_item.pre.id)
            
            # Set PRE info (only once per PRE)
            if not pre_groups[pre_key]['pre_display']:
                pre_groups[pre_key]['pre_display'] = f"{line_item.pre.department} - FY {line_item.pre.fiscal_year}"
                pre_groups[pre_key]['pre_id'] = pre_key
            
            # Build display name
            category_name = line_item.category.name if line_item.category else 'Other'
            if line_item.subcategory:
                display = f"{category_name} - {line_item.subcategory.name} - {line_item.item_name}"
            else:
                display = f"{category_name} - {line_item.item_name}"
            
            # Add line item to group
            pre_groups[pre_key]['line_items'].append({
                'id': line_item.id,
                'display': display,
                'item_name': line_item.item_name,
                'category': category_name,
                'total': total,
                'consumed': consumed,
                'available': available,
                'value': f"{line_item.pre.id}|{line_item.id}",
                'label': f"{display} (Available: ₱{available:,.2f})"
            })
    
    # Convert defaultdict to list
    return list(pre_groups.values())


def calculate_line_item_consumed(line_item):
    """
    Calculate the total consumed/allocated amount for a specific PRE line item.
    Sums up all allocations from approved/pending Purchase Requests.
    
    Args:
        line_item: PRELineItem instance or line_item_id
        
    Returns:
        Decimal: Total consumed amount from all allocations
    """
    if isinstance(line_item, PRELineItem):
        line_item_id = line_item.id
    else:
        line_item_id = line_item
    
    consumed = NewPurchaseRequestAllocation.objects.filter(
        pre_line_item_id=line_item_id,
        purchase_request__status__in=[
            'Pending', 'Partially Approved', 'Approved'
        ]
    ).aggregate(
        total=Coalesce(
            Sum('allocated_amount'),
            Decimal('0.00'),
            output_field=DecimalField(max_digits=15, decimal_places=2)
        )
    )['total']
    
    return consumed


def allocate_funds_to_purchase_request(purchase_request, pre_id, line_item_id, amount):
    """
    Allocate funds from a specific PRE line item to a purchase request.
    Creates a PurchaseRequestAllocation record linking the PR to the PRE line item.
    
    Args:
        purchase_request: PurchaseRequest instance
        pre_id: UUID of the DepartmentPRE
        line_item_id: ID of the PRELineItem
        amount: Decimal amount to allocate
        
    Returns:
        dict: {
            'success': bool,
            'allocation': PurchaseRequestAllocation instance or None,
            'error': str (if success is False)
        }
    """
    try:
        # Get the line item
        line_item = PRELineItem.objects.select_related('pre', 'category', 'subcategory').get(
            id=line_item_id,
            pre_id=pre_id,
            pre__status='Approved'
        )
        
        # Validate availability
        consumed = calculate_line_item_consumed(line_item)
        total = line_item.get_total()
        available = total - consumed
        
        if available < amount:
            return {
                'success': False,
                'allocation': None,
                'error': f'Insufficient funds. Available: ₱{available:,.2f}, Required: ₱{amount:,.2f}'
            }
        
        # Create allocation
        allocation = NewPurchaseRequestAllocation.objects.create(
            purchase_request=purchase_request,
            pre_line_item=line_item,
            allocated_amount=amount
        )
        
        return {
            'success': True,
            'allocation': allocation,
            'error': None
        }
        
    except PRELineItem.DoesNotExist:
        return {
            'success': False,
            'allocation': None,
            'error': 'Selected funding source not found or PRE not approved.'
        }
    except Exception as e:
        return {
            'success': False,
            'allocation': None,
            'error': f'Error allocating funds: {str(e)}'
        }


# Additional helper function for validation
def validate_line_item_availability(line_item_id, required_amount):
    """
    Validate if a PRE line item has sufficient available funds.
    
    Args:
        line_item_id: ID of the PRELineItem
        required_amount: Decimal amount needed
        
    Returns:
        dict: {
            'available': bool,
            'total': Decimal,
            'consumed': Decimal,
            'remaining': Decimal,
            'message': str
        }
    """
    try:
        line_item = PRELineItem.objects.get(id=line_item_id)
        consumed = calculate_line_item_consumed(line_item)
        total = line_item.get_total()
        remaining = total - consumed
        
        return {
            'available': remaining >= required_amount,
            'total': total,
            'consumed': consumed,
            'remaining': remaining,
            'message': f'Available: ₱{remaining:,.2f} of ₱{total:,.2f}'
        }
    except PRELineItem.DoesNotExist:
        return {
            'available': False,
            'total': Decimal('0.00'),
            'consumed': Decimal('0.00'),
            'remaining': Decimal('0.00'),
            'message': 'Line item not found'
        }


# Helper to get line item details for display
def get_line_item_details(pre_id, line_item_id):
    """
    Get detailed information about a PRE line item including availability.
    
    Args:
        pre_id: UUID of the DepartmentPRE
        line_item_id: ID of the PRELineItem
        
    Returns:
        dict with line item details or None if not found
    """
    try:
        line_item = PRELineItem.objects.select_related(
            'pre', 'category', 'subcategory'
        ).get(
            id=line_item_id,
            pre_id=pre_id
        )
        
        consumed = calculate_line_item_consumed(line_item)
        total = line_item.get_total()
        available = total - consumed
        
        category_name = line_item.category.name if line_item.category else 'Other'
        if line_item.subcategory:
            display = f"{category_name} - {line_item.subcategory.name} - {line_item.item_name}"
        else:
            display = f"{category_name} - {line_item.item_name}"
        
        return {
            'id': line_item.id,
            'pre_id': str(line_item.pre.id),
            'item_name': line_item.item_name,
            'display': display,
            'category': category_name,
            'subcategory': line_item.subcategory.name if line_item.subcategory else None,
            'total': total,
            'consumed': consumed,
            'available': available,
            'q1': line_item.q1_amount,
            'q2': line_item.q2_amount,
            'q3': line_item.q3_amount,
            'q4': line_item.q4_amount,
        }
    except PRELineItem.DoesNotExist:
        return None
    
@role_required('end_user', login_url='/')
def get_pre_line_items(request):
    """
    AJAX endpoint to get PRE line items with quarterly breakdown.
    """
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)
    
    allocation_id = request.GET.get('allocation_id')
    
    if not allocation_id:
        return JsonResponse({'success': False, 'error': 'Budget allocation ID is required'}, status=400)
    
    try:
        allocation = NewBudgetAllocation.objects.get(
            id=allocation_id,
            end_user=request.user,
            is_active=True
        )
        
        approved_pres = NewDepartmentPRE.objects.filter(
            budget_allocation=allocation,
            status__in=['Approved', 'Partially Approved']
        ).prefetch_related(
            'line_items__category',
            'line_items__subcategory'
        )
        
        line_items_data = []
        
        for pre in approved_pres:
            for line_item in pre.line_items.all():
                category_name = line_item.category.name if line_item.category else 'Other'
                if line_item.subcategory:
                    display = f"{category_name} - {line_item.subcategory.name} - {line_item.item_name}"
                else:
                    display = f"{category_name} - {line_item.item_name}"
                
                # ✅ Add quarterly breakdown
                quarters = ['Q1', 'Q2', 'Q3', 'Q4']
                for quarter in quarters:
                    # Get quarter amount
                    quarter_amount = getattr(line_item, f'{quarter.lower()}_amount', Decimal('0'))
                    
                    if quarter_amount <= 0:
                        continue
                    
                    # ✅ FIXED: Calculate consumed amount for this quarter
                    consumed = NewPurchaseRequestAllocation.objects.filter(
                        pre_line_item=line_item,
                        quarter=quarter,
                        purchase_request__status__in=['Pending', 'Partially Approved', 'Approved']
                    ).aggregate(
                        total=Coalesce(
                            Sum('allocated_amount'),
                            Decimal('0.00'),
                            output_field=DecimalField(max_digits=15, decimal_places=2)
                        )
                    )['total']
                    
                    available = quarter_amount - consumed
                    
                    if available > 0:
                        line_items_data.append({
                            'id': line_item.id,
                            'pre_id': str(pre.id),
                            'pre_display': f"{pre.department} - FY {pre.fiscal_year}",
                            'display': display,
                            'item_name': line_item.item_name,
                            'category': category_name,
                            'quarter': quarter,
                            'quarter_amount': float(quarter_amount),
                            'consumed': float(consumed),
                            'available': float(available),
                            'value': f"{pre.id}|{line_item.id}|{quarter}",
                            'label': f"{display} - {quarter} (Available: ₱{available:,.2f})"
                        })
        
        return JsonResponse({
            'success': True,
            'line_items': line_items_data,
            'allocation_id': allocation_id,
            'allocation_name': allocation.approved_budget.title
        })
        
    except NewBudgetAllocation.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Budget allocation not found'
        }, status=404)
    except Exception as e:
        # ✅ Add detailed error logging
        import traceback
        print(f"Error in get_pre_line_items: {str(e)}")
        traceback.print_exc()
        
        return JsonResponse({
            'success': False,
            'error': f'Error loading line items: {str(e)}'
        }, status=500)

@role_required('end_user', login_url='/')
def purchase_request_upload(request):
    """Handle PR document upload with supporting files"""
    
    # Get or create draft
    draft, created = PRDraft.objects.get_or_create(
        user=request.user
    )
    
    # Get budget allocations for dropdown
    budget_allocations = NewBudgetAllocation.objects.filter(
        end_user=request.user,
        is_active=True,
        remaining_balance__gt=0
    ).select_related('approved_budget')
    
    # Get approved PRE line items as funding sources
    # source_of_fund_options = build_pre_source_options_v2(request.user)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if action == 'upload_pr':
            # Handle PR document upload
            pr_file = request.FILES.get('pr_document')
            
            if not pr_file:
                if is_ajax:
                    return JsonResponse({'success': False, 'error': 'Please select a PR document to upload.'})
                messages.error(request, "Please select a PR document to upload.")
                return redirect('purchase_request_upload')
            
            # Validate file type
            allowed_extensions = ['pdf', 'doc', 'docx']
            file_ext = pr_file.name.split('.')[-1].lower()
            
            if file_ext not in allowed_extensions:
                error_msg = f"Invalid file type. Allowed: {', '.join(allowed_extensions).upper()}"
                if is_ajax:
                    return JsonResponse({'success': False, 'error': error_msg})
                messages.error(request, error_msg)
                return redirect('purchase_request_upload')
            
            # Validate file size (10MB)
            if pr_file.size > 10 * 1024 * 1024:
                error_msg = "File size must be less than 10MB."
                if is_ajax:
                    return JsonResponse({'success': False, 'error': error_msg})
                messages.error(request, error_msg)
                return redirect('purchase_request_upload')
            
            # Delete old file if exists
            if draft.pr_file:
                draft.pr_file.delete(save=False)
            
            # Save new file
            draft.pr_file = pr_file
            draft.pr_filename = pr_file.name
            draft.save()
            
            success_msg = f"PR document '{pr_file.name}' uploaded successfully."
            if is_ajax:
                return JsonResponse({'success': True, 'message': success_msg})
            messages.success(request, success_msg)
            return redirect('purchase_request_upload')
        
        elif action == 'remove_pr':
            # Remove PR document
            if draft.pr_file:
                draft.pr_file.delete(save=False)
                draft.pr_file = None
                draft.pr_filename = ''
                draft.save()
                
                success_msg = "PR document removed."
                if is_ajax:
                    return JsonResponse({'success': True, 'message': success_msg})
                messages.info(request, success_msg)
            else:
                error_msg = "No PR document to remove."
                if is_ajax:
                    return JsonResponse({'success': False, 'error': error_msg})
                messages.warning(request, error_msg)
                
            return redirect('purchase_request_upload')
        
        elif action == 'upload_supporting':
            # Handle supporting documents upload
            supporting_docs = request.FILES.getlist('supporting_documents')
            
            if not supporting_docs:
                if is_ajax:
                    return JsonResponse({'success': False, 'error': 'Please select at least one supporting document.'})
                messages.error(request, "Please select at least one supporting document.")
                return redirect('purchase_request_upload')
            
            # Validate and save each document
            allowed_extensions = ['pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png']
            count = 0
            errors = []
            
            for doc in supporting_docs:
                # Check file type
                file_ext = doc.name.split('.')[-1].lower()
                if file_ext not in allowed_extensions:
                    errors.append(f"Skipped '{doc.name}': Invalid file type.")
                    continue
                
                # Check file size
                if doc.size > 10 * 1024 * 1024:
                    errors.append(f"Skipped '{doc.name}': File too large (max 10MB).")
                    continue
                
                # Check for duplicates
                if draft.supporting_documents.filter(file_name=doc.name).exists():
                    errors.append(f"Skipped '{doc.name}': Already uploaded.")
                    continue
                
                # Save document
                PRDraftSupportingDocument.objects.create(
                    draft=draft,
                    document=doc,
                    file_name=doc.name,
                    file_size=doc.size
                )
                count += 1
                
            if is_ajax:
                if count > 0:
                    return JsonResponse({
                        'success': True, 
                        'message': f"Uploaded {count} supporting document(s).",
                        'errors': errors
                    })
                else:
                    return JsonResponse({
                        'success': False, 
                        'error': 'No documents were uploaded.',
                        'errors': errors
                    })
            
            if count > 0:
                messages.success(request, f"Uploaded {count} supporting document(s).")
                
            for error in errors:
                messages.warning(request, error)
            return redirect('purchase_request_upload')
        
        elif action == 'remove_supporting':
            # Remove supporting document
            doc_id = request.POST.get('doc_id')
            try:
                doc = draft.supporting_documents.get(id=doc_id)
                doc_name = doc.file_name
                doc.document.delete(save=False)
                doc.delete()
                
                success_msg = f"Removed '{doc_name}'"
                if is_ajax:
                    return JsonResponse({'success': True, 'message': success_msg})
                messages.success(request, success_msg)
            except PRDraftSupportingDocument.DoesNotExist:
                error_msg = "Document not found."
                if is_ajax:
                    return JsonResponse({'success': False, 'error': error_msg})
                messages.error(request, error_msg)
            
            return redirect('purchase_request_upload')
        
        elif action == 'clear_draft':
            # Clear all draft files
            if draft.pr_file:
                draft.pr_file.delete(save=False)
            for doc in draft.supporting_documents.all():
                doc.document.delete(save=False)
                doc.delete()
            draft.delete()
            
            success_msg = "All draft files cleared."
            if is_ajax:
                return JsonResponse({'success': True, 'message': success_msg})
            messages.info(request, success_msg)
            return redirect('purchase_request_upload')
        
        elif action == 'submit':
             # Get form data
            budget_allocation_id = request.POST.get('budget_allocation')
            source_of_fund = request.POST.get('source_of_fund')
            total_amount = request.POST.get('total_amount')
            purpose = request.POST.get('purpose')
            
            # Validate required fields
            if not all([budget_allocation_id, source_of_fund, total_amount, purpose]):
                messages.error(request, "All fields are required.")
                return redirect('purchase_request_upload')
            
            # Validate draft has PR file
            if not draft.pr_file:
                messages.error(request, "Please upload a PR document first.")
                return redirect('purchase_request_upload')
            
            # Parse source_of_fund value
            if '|' in source_of_fund:
                parts = source_of_fund.split('|')
                if len(parts) == 3:  # ✅ Expect 3 parts (pre_id|line_item_id|quarter)
                    pre_id = parts[0]
                    line_item_id = parts[1]
                    quarter = parts[2]  # ✅ Extract quarter
                else:
                    messages.error(request, "Invalid source of fund format.")
                    return redirect('purchase_request_upload')
            else:
                messages.error(request, "Invalid source of fund selection.")
                return redirect('purchase_request_upload')
            
            # ✅ Validate quarter
            if quarter not in ['Q1', 'Q2', 'Q3', 'Q4']:
                messages.error(request, "Invalid quarter selected.")
                return redirect('purchase_request_upload')
            
            # Validate amount
            try:
                total_amount = Decimal(total_amount)
                if total_amount <= 0:
                    messages.error(request, "Amount must be greater than zero.")
                    return redirect('purchase_request_upload')
            except (ValueError, InvalidOperation):
                messages.error(request, "Invalid amount format.")
                return redirect('purchase_request_upload')
            
            # Validate line item availability for the specific quarter
            try:
                line_item = PRELineItem.objects.get(id=line_item_id)
                pre = line_item.pre
                
                # Get quarter amount
                quarter_amount = getattr(line_item, f'{quarter.lower()}_amount', Decimal('0'))
                
                # Calculate consumed for this quarter
                consumed = NewPurchaseRequestAllocation.objects.filter(
                    pre_line_item=line_item,
                    quarter=quarter,  # ✅ Filter by quarter
                    purchase_request__status__in=['Pending', 'Partially Approved', 'Approved']
                ).aggregate(
                    total=Coalesce(
                        Sum('allocated_amount'),
                        Decimal('0.00'),
                        output_field=DecimalField(max_digits=15, decimal_places=2)
                    )
                )['total']
                
                available = quarter_amount - consumed
                
                if total_amount > available:
                    messages.error(request, 
                        f"Amount (₱{total_amount:,.2f}) exceeds available budget for {quarter} "
                        f"(₱{available:,.2f}).")
                    return redirect('purchase_request_upload')
                
            except PRELineItem.DoesNotExist:
                messages.error(request, "Selected PRE line item not found.")
                return redirect('purchase_request_upload')
            
            # Create Purchase Request
            try:
                from django.db import transaction
                from django.core.files.base import ContentFile
                
                with transaction.atomic():
                    # Generate PR number
                    pr_number = generate_pr_number()
                    
                    # Get budget allocation
                    budget_allocation = NewBudgetAllocation.objects.get(id=budget_allocation_id)
                    
                    # ✅ CORRECTED: Create PR with correct field names
                    pr = NewPurchaseRequest.objects.create(
                        pr_number=pr_number,
                        submitted_by=request.user,  # ✅ Correct field name
                        department=budget_allocation.department,  # ✅ Get from allocation
                        budget_allocation=budget_allocation,
                        source_pre=pre,  # ✅ Link to PRE
                        source_line_item=line_item,  # ✅ Link to line item
                        source_of_fund_display=f"{line_item.item_name} - {quarter}",  # ✅ Human-readable
                        purpose=purpose,
                        total_amount=total_amount,
                        status='Pending',
                        uploaded_document=draft.pr_file,  # ✅ Correct field name
                    )
                    
                    # Copy supporting documents from draft
                    for draft_doc in draft.supporting_documents.all():
                        # Read the draft document
                        draft_doc.document.open('rb')
                        file_content = draft_doc.document.read()
                        draft_doc.document.close()
                        
                        # Create new supporting document for PR
                        pr_doc = PurchaseRequestSupportingDocument.objects.create(
                            purchase_request=pr,
                            file_name=draft_doc.file_name,
                            file_size=draft_doc.file_size
                        )
                        
                        # Save the file
                        pr_doc.document.save(
                            draft_doc.file_name,
                            ContentFile(file_content),
                            save=True
                        )
                    
                    # ✅ Create allocation with quarter
                    NewPurchaseRequestAllocation.objects.create(
                        purchase_request=pr,
                        pre_line_item=line_item,
                        quarter=quarter,  # ✅ Save quarter
                        allocated_amount=total_amount,
                        notes=f"Allocated from {line_item.item_name} - {quarter}"
                    )
                    
                    # Clear draft
                    draft.delete()
                    
                    messages.success(request, 
                        f"Purchase Request {pr_number} submitted successfully! "
                        f"Allocated ₱{total_amount:,.2f} from {line_item.item_name} - {quarter}.")
                    
                     # Log audit trail
                    log_audit_trail(
                        request=request,
                        action='CREATE',
                        model_name='PurchaseRequest',
                        record_id=pr.id,
                        detail=f'Submitted Purchase Request {pr_number}'
                    )
                    
                    # ✅ NEW: Redirect to preview instead of upload form
                    return redirect('preview_submitted_pr', pr_id=pr.id)
                    
            except Exception as e:
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error creating purchase request: {str(e)}")
                return redirect('purchase_request_upload')
                    
            except NewBudgetAllocation.DoesNotExist:
                messages.error(request, "Invalid budget allocation selected.")
                return redirect('purchase_request_upload')
            except PRELineItem.DoesNotExist:
                messages.error(request, "Selected funding source not found or PRE not approved.")
                return redirect('purchase_request_upload')
            except Exception as e:
                messages.error(request, f"Error submitting PR: {str(e)}")
                print(f"PR submission error: {e}")
                import traceback
                traceback.print_exc()
                return redirect('purchase_request_upload')
    
    # GET request
    context = {
        'draft': draft,
        'budget_allocations': budget_allocations,
        #'source_of_fund_options': source_of_fund_options,
    }
    
    return render(request, 'end_user_app/purchase_request_upload_form.html', context)

@role_required('end_user', login_url='/')
def preview_submitted_pr(request, pr_id):
    """
    Preview page for submitted Purchase Request
    Shows all details, uploaded documents, and current status
    User can only view their own PRs
    """
    pr = get_object_or_404(
        NewPurchaseRequest.objects.select_related(
            'submitted_by',
            'budget_allocation',
            'budget_allocation__approved_budget',
            'source_pre',
            'source_line_item',
            'source_line_item__category',
            'source_line_item__subcategory'
        ).prefetch_related(
            'supporting_documents',
            'pre_allocations',
            'pre_allocations__pre_line_item',
            'pre_allocations__pre_line_item__category'
        ),
        id=pr_id,
        submitted_by=request.user  # ✅ Security: Only show user's own PRs
    )
    
    # Get allocation details (should only be one for upload-based PRs)
    allocation = pr.pre_allocations.first()
    
    # Check if user can cancel (only pending PRs)
    can_cancel = pr.status == 'Pending'
    
    # Check if user can edit (only draft PRs - shouldn't happen for submitted PRs)
    can_edit = pr.status == 'Draft'
    
    context = {
        'pr': pr,
        'allocation': allocation,
        'can_cancel': can_cancel,
        'can_edit': can_edit,
    }
    
    return render(request, 'end_user_app/preview_submitted_pr.html', context)

@role_required('end_user', login_url='/')
def purchase_request_form(request):
    # Ensure a draft PR exists for the user to accumulate items prior to submission
    purchase_request_obj, _created = PurchaseRequest.objects.get_or_create(
        requested_by=request.user,
        pr_status='draft',
        defaults={
            'entity_name': 'Draft Entity',
            'pr_no': 'TEMP',
        }
    )

    # Budget Allocation options (limit to user's department)
    budget_allocations = BudgetAllocation.objects.select_related('approved_budget').filter(
        department=getattr(request.user, 'department', '')
    )
    
    approved_pres = DepartmentPRE.objects.filter(
            submitted_by=request.user,
            approved_by_approving_officer=True,
            approved_by_admin=True,
        )

    source_of_fund_options = build_pre_source_options(approved_pres)

    if request.method == 'POST':
        # Parse and save header fields
        purchase_request_obj.entity_name = request.POST.get('entity_name') or purchase_request_obj.entity_name
        purchase_request_obj.fund_cluster = request.POST.get('fund_cluster') or purchase_request_obj.fund_cluster
        purchase_request_obj.office_section = request.POST.get('office_section') or purchase_request_obj.office_section
        pr_no_input = request.POST.get('pr_no')
        if pr_no_input and purchase_request_obj.pr_no == 'TEMP':
            purchase_request_obj.pr_no = pr_no_input
        purchase_request_obj.responsibility_center_code = request.POST.get('responsibility_code') or purchase_request_obj.responsibility_center_code
        purchase_request_obj.purpose = request.POST.get('purpose') or purchase_request_obj.purpose

        # Budget Allocation linkage
        ba_id = request.POST.get('budget_allocation')
        if ba_id:
            try:
                purchase_request_obj.budget_allocation = BudgetAllocation.objects.select_related('approved_budget').get(id=ba_id)
            except BudgetAllocation.DoesNotExist:
                print("No such budget allocation for this user.")
                purchase_request_obj.budget_allocation = None
                
        # BUDGET VALIDATION BEFORE SAVING
        sof_encoded = request.POST.get('source_of_fund')
        if sof_encoded:
            try:
                parts = sof_encoded.split('|', 2)
                if len(parts) >= 3:
                    pre_id_str, item_key, quarters_data = parts
                    # Department PRE object instance
                    pre_obj = DepartmentPRE.objects.get(id=int(pre_id_str))
                    
                    # Validate budget availability
                    all_items = PRELineItemBudget.objects.filter(
                        pre=pre_obj,
                        item_key=item_key
                    )
                    
                    available_items = [item for item in all_items if item.remaining_amount > 0]
                    
                    if not available_items:
                        error_message =  f"No budget available for the selected source of fund. All quarters consumed."
                        print(error_message)
                        return JsonResponse({
                            'success': False,
                            'error': error_message,
                            'error_type': 'insufficient_budget'
                        })
                    
                    total_available = sum(item.remaining_amount for item in available_items)
                    
                    if total_available < purchase_request_obj.total_amount:
                        error_message = f"Insufficient budget. Available: ₱{total_available:,.2f}, Required: ₱{purchase_request_obj.total_amount:,.2f}"
                        print(error_message)
                        return JsonResponse({
                            'success': False,
                            'error': error_message,
                            'error_type': 'insufficient_budget'
                        })
                    
                    remaining_to_allocate = purchase_request_obj.total_amount
                    allocations_made = []
                    
                    for item in available_items:
                        if remaining_to_allocate <= 0:
                            break
                        
                        allocation_amount = min(item.remaining_amount, remaining_to_allocate)
                        if allocation_amount > 0:
                            item.consumed_amount += allocation_amount
                            item.save()
                            
                            # Track this allocation for potential reversal
                            allocation_record = PurchaseRequestAllocation.objects.create(
                                purchase_request=purchase_request_obj,
                                pre_line_item=item,
                                allocated_amount=allocation_amount
                            )
                            
                            allocations_made.append(allocation_record)
                            
                            remaining_to_allocate -= allocation_amount
                            
                            print(f"Allocated ₱{allocation_amount} from {item.item_key} {item.quarter}")
                    
                    print(f"Total allocations made: {len(allocations_made)}")
                    
                    # Store source information
                    purchase_request_obj.source_pre = pre_obj
                    purchase_request_obj.source_item_key = item_key
                    
                    # Parse quarters data to get the first quarter for compatibility
                    quarters = quarters_data.split('|')
                    if quarters:
                        first_quarter_data = quarters[0].split(':')
                        if len(first_quarter_data) == 2:
                            quarter, amount_str = first_quarter_data
                            purchase_request_obj.source_quarter = quarter
                            purchase_request_obj.source_amount = Decimal(amount_str)
                            
                    
                    # Calculate total amount across all quarters
                    total_amount = Decimal('0')
                    for quarter_info in quarters:
                        if ':' in quarter_info:
                            q, amount_str = quarter_info.split(':')
                            total_amount += Decimal(amount_str)

                    # Set the display value
                    item_label = FRIENDLY_LABELS.get(item_key, item_key.replace('_', ' ').title())
                    purchase_request_obj.source_of_fund_display = f"{item_label} - ₱{total_amount:,.2f}"
                    
            except Exception as e:
                error_message = f"Error validating budget: {e}"
                print(error_message)
                return JsonResponse({
                    'success': False,
                    'error': error_message,
                    'error_type': 'budget_validation_error'
                })

        # Mark as submitted
        purchase_request_obj.pr_status = 'Submitted'
        purchase_request_obj.submitted_status = 'Pending'
        purchase_request_obj.save()
        
        log_audit_trail(
            request=request,
            action='CREATE',
            model_name='PurchaseRequest',
            record_id=purchase_request_obj.id,
            detail=f'Created a new purchase request {purchase_request_obj.pr_no}',
        )

        from django.urls import reverse
        preview_url = reverse('end_user_preview_purchase_request', args=[purchase_request_obj.id])
        return JsonResponse({'success': True, 'redirect_url': preview_url})

    # GET request
    # Items to render in the table
    items = purchase_request_obj.items.all()

    return render(
        request,
        "end_user_app/purchase_request_form.html",
        {
            'purchase_request': purchase_request_obj,
            'purchase_items': items,
            'budget_allocations': budget_allocations,
            'source_of_fund_options': source_of_fund_options,
        }
    )

@role_required('end_user', login_url='/')
def preview_purchase_request(request, pk: int):
    pr = get_object_or_404(
        PurchaseRequest.objects.select_related('requested_by', 'budget_allocation__approved_budget', 'source_pre'),
        pk=pk,
        requested_by=request.user,
    )
    return render(request, 'end_user_app/preview_purchase_request.html', {'pr': pr})

@role_required('end_user', login_url='/')
def papp_list(request, papp):
    try:
        papp = BudgetAllocation.objects.filter(assigned_user=request.user).values_list('papp', flat=True)
    except BudgetAllocation.DoesNotExist:
        papp = None
    return papp

@role_required('end_user', login_url='/')
def re_alignment(request):
    """
    target_papp = papp_list(request, papp="target_papp")
    source_papp = papp_list(request, papp="source_papp")
    
    
    
    if request.method == 'POST':
        # Get the form data from the request
        get_target_papp = request.POST.get('target_papp')
        get_source_papp = request.POST.get('source_papp')
        amount = request.POST.get('amount')
        reason = request.POST.get('reason')
        
        if not get_target_papp or not get_source_papp:
            messages.error(request, "Target PAPP and Source PAPP is required")
            return redirect('re_alignment')
        
        if get_target_papp == get_source_papp:
            messages.error(request, "Target PAPP and Source PAPP should not equal.")
            return redirect('re_alignment')
        
        # Check if the user has sufficient budget in the source PAPP
        try:
            source_budget = BudgetAllocation.objects.get(department=request.user, papp=get_source_papp)
            if source_budget.remaining_budget < Decimal(amount):
                messages.error(request, "Insufficient budget in the source PAPP.")
                return redirect('re_alignment')
        except BudgetAllocation.DoesNotExist:
            messages.error(request, "Source PAPP not found.")
            return redirect('re_alignment')
        
        re_alignment = Budget_Realignment.objects.create(
                    requested_by=request.user,
                    target_papp=get_target_papp,
                    source_papp=get_source_papp,
                    amount=Decimal(amount),
                    reason=reason
                )
        
        messages.success(request, "Budget for Realignment has been requested successfully")
        return redirect('re_alignment')
        """
            
    return render(request, "end_user_app/re_alignment.html")
    
@require_http_methods(["POST"])
@role_required('end_user', login_url='/')
def add_purchase_request_items(request):
    # Get or create draft purchase request
    purchase_request, created = PurchaseRequest.objects.get_or_create(
        requested_by=request.user,
        pr_status='draft',
        defaults={
            'entity_name': 'Draft Entity',
            'pr_no': 'TEMP',
            'pr_status': 'draft'
        }
    )
    
    print(request.POST)  # Debugging: Check what data is being sent
    
    # Create item
    item = PurchaseRequestItems.objects.create(
        purchase_request=purchase_request,
        stock_property_no=request.POST.get('stock_no'),
        unit=request.POST.get('unit'),
        item_description=request.POST.get('item_desc'),
        quantity=int(request.POST.get('quantity', 0)),
        unit_cost=Decimal(request.POST.get('unit_cost', 0)),
    )
    
    return render(request, "end_user_app/partials/item_with_total_row.html", {"item": item, "purchase_request": purchase_request})
    
    # return JsonResponse({
    #     'success': True,
    #     'item': {
    #         'stock_property_no': item.stock_property_no,
    #         'unit': item.unit,
    #         'item_description': item.item_description,
    #         'quantity': item.quantity,
    #         'unit_cost': float(item.unit_cost),
    #         'total_cost': float(item.total_cost),
    #     }
    # })
    

@role_required('end_user', login_url='/')
@require_http_methods(["DELETE"])
def remove_purchase_item(request, item_id):
    remove_item = get_object_or_404(PurchaseRequestItems, id=item_id)
    purchase_request = remove_item.purchase_request  # ✅ This was missing

    # Delete the item
    remove_item.delete()

    # Update total amount
    total = sum(item.total_cost for item in purchase_request.items.all())
    purchase_request.total_amount = total
    purchase_request.save()

    # Render updated total row with OOB swap
    html = render_to_string("end_user_app/partials/deleted_row_and_total.html", {
        "item_id": item_id,
        "purchase_request": purchase_request
    })

    return HttpResponse(html)

    # return HttpResponse(status=200)  # Empty response with 200 OK
    #return JsonResponse({"success": True})
    
@role_required('end_user', login_url='/')
def department_pre_form(request, pk:int):
    budget_allocation = BudgetAllocation.objects.filter(department=request.user.department, id=pk).first()
    context = {
        'budget': budget_allocation,
        
        'personnel_services': [
            {'label': 'Basic Salary', 'name': 'basic_salary'},
            {'label': 'Honoraria', 'name': 'honoraria'},
            {'label': 'Overtime Pay', 'name': 'overtime_pay'},
        ],
        'MOOE_traveling_expenses': [
            {'label': 'Traveling Expenses - Local', 'name': 'travel_local'},
            {'label': 'Traveling Expenses - Foreign', 'name': 'travel_foreign'},
        ],
        'MOOE_training_and_scholarship_expenses': [
            {'label': 'Training Expenses', 'name': 'training_expenses'},
        ],
        'MOOE_supplies_and_materials_expenses': [
            {'label': 'Office Supplies Expenses', 'name': 'office_supplies_expenses'},
            {'label': 'Accountable Form Expenses', 'name': 'accountable_form_expenses'},
            {'label': 'Agricultural and Marine Supplies Expenses', 'name': 'agri_marine_supplies_expenses'},
            {'label': 'Drugs and Medicines', 'name': 'drugs_medicines'},
            {'label': 'Medical, Dental & Laboratory Supplies Expenses', 'name': 'med_dental_lab_supplies_expenses'},
            {'label': 'Food Supplies Expenses', 'name': 'food_supplies_expenses'},
            {'label': 'Fuel, Oil and Lubricants Expenses', 'name': 'fuel_oil_lubricants_expenses'},
            {'label': 'Textbooks and Instructional Materials Expenses', 'name': 'textbooks_instructional_materials_expenses'},
            {'label': 'Construction Materials Expenses', 'name': 'construction_material_expenses'},
            {'label': 'Other Supplies & Materials Expenses', 'name': 'other_supplies_materials_expenses'},
        ],
        'MOOE_semi_expendable_machinery_equipment_expenses': [
            {'label': 'Machinery', 'name': 'semee_machinery'},
            {'label': 'Office Equipment', 'name': 'semee_office_equipment'},
            {'label': 'Information and Communications Technology Equipment', 'name': 'semee_information_communication'},
            {'label': 'Communications Equipment', 'name': 'semee_communications_equipment'},
            {'label': 'Disaster Response and Rescue Equipment', 'name': 'semee_drr_equipment'},
            {'label': 'Medical Equipment', 'name': 'semee_medical_equipment'},
            {'label': 'Printing Equipment', 'name': 'semee_printing_equipment'},
            {'label': 'Sports Equipment', 'name': 'semee_sports_equipment'},
            {'label': 'Technical and Scientific Equipment', 'name': 'semee_technical_scientific_equipment'},
            {'label': 'ICT Equipment', 'name': 'semee_ict_equipment'},
            {'label': 'Other Machinery and Equipment', 'name': 'semee_other_machinery_equipment'},
        ],
        'MOOE_semi_expendable_furnitures_fixtures_books_expenses': [
            {'label': 'Furniture and Fixtures', 'name': 'furniture_fixtures'},
            {'label': 'Books', 'name': 'books'},
        ],
        'MOOE_utility_expenses': [
            {'label': 'Water Expenses', 'name': 'water_expenses'},
            {'label': 'Electricity Expenses', 'name': 'electricity_expenses'},
        ],
        'MOOE_communication_expenses': [
            {'label': 'Postage and Courier Services', 'name': 'postage_courier_services'},
            {'label': 'Telephone Expenses', 'name': 'telephone_expenses'},
            {'label': 'Telephone Expenses (Landline)', 'name': 'telephone_expenses_landline'},
            {'label': 'Internet Subscription Expenses', 'name': 'internet_subscription_expenses'},
            {'label': 'Cable, Satellite, Telegraph & Radio Expenses', 'name': 'cable_satellite_telegraph_radio_expenses'},
        ],
        'MOOE_awards_rewards_prizes': [
            {'label': 'Awards/Rewards Expenses', 'name': 'awards_rewards_expenses'},
            {'label': 'Prizes', 'name': 'prizes'},
        ],
        'MOOE_survey_research_exploration_development': [
            {'label': 'Survey Expenses', 'name': 'survey_expenses'},
            {'label': 'Survey, Research, Exploration, and Development expenses', 'name': 'survey_research_exploration_development_expenses'},
        ],
        'MOOE_professional_services': [
            {'label': 'Legal Services', 'name': 'legal_services'},
            {'label': 'Auditing Services', 'name': 'auditing_services'},
            {'label': 'Consultancy Services', 'name': 'consultancy_services'},
            {'label': 'Other Professional Services', 'name': 'other_professional_servies'},
        ],
        'MOOE_general_services': [
            {'label': 'Security Services', 'name': 'security_services'},
            {'label': 'Janitorial Services', 'name': 'janitorial_services'},
            {'label': 'Other General Services', 'name': 'other_general_services'},
            {'label': 'Environment/Sanitary Services', 'name': 'environment/sanitary_services'},
        ],
        'repair_land_improvements': [
            {'label': 'Repair & Maintenance - Land Improvements', 'name': 'repair_maintenance_land_improvements'},
        ],
        'repair_buildings_structures': [
            {'label': 'Buildings', 'name': 'buildings'},
            {'label': 'School Buildings', 'name': 'school_buildings'},
            {'label': 'Hostels and Dormitories', 'name': 'hostel_dormitories'},
            {'label': 'Other Structures', 'name': 'other_structures'},
        ],
        'repair_machinery_equipment': [
            {'label': 'Machinery', 'name': 'repair_maintenance_machinery'},
            {'label': 'Office Equipment', 'name': 'repair_maintenance_office_equipment'},
            {'label': 'ICT Equipment', 'name': 'repair_maintenance_ict_equipment'},
            {'label': 'Agricultural and Forestry Equipment', 'name': 'repair_maintenance_agri_forestry_equipment'},
            {'label': 'Marine and Fishery Equipment', 'name': 'repair_maintenance_marine_fishery_equipment'},
            {'label': 'Airport Equipment', 'name': 'repair_maintenance_airport_equipment'},
            {'label': 'Communication Equipment', 'name': 'repair_maintenance_communication_equipment'},
            {'label': 'Disaster, Response and Rescue Equipment', 'name': 'repair_maintenance_drre_equipment'},
            {'label': 'Medical Equipment', 'name': 'repair_maintenance_medical_equipment'},
            {'label': 'Printing Equipment', 'name': 'repair_maintenance_printing_equipment'},
            {'label': 'Sports Equipment', 'name': 'repair_maintenance_sports_equipment'},
            {'label': 'Technical and Scientific Equipment', 'name': 'repair_maintenance_technical_scientific_equipment'},
            {'label': 'Other Machinery and Equipment', 'name': 'repair_maintenance_other_machinery_equipment'},
        ],
        'repair_transportation_equipment': [
            {'label': 'Motor Vehicles', 'name': 'repair_maintenance_motor'},
            {'label': 'Other Transportation Equipment', 'name': 'repair_maintenance_other_transportation_equipment'},
        ],
        'repair_furniture_fixtures': [
            {'label': 'Repairs and Maintenance - Furniture & Fixtures', 'name': 'repair_maintenance_furniture_fixtures'},
        ],
        'repair_semi_expendable_me': [
            {'label': 'Repairs and Maintenance - Semi-Expendable Machinery and Equipment', 'name': 'repair_maintenance_semi_expendable_machinery_equipment'},
        ],
        'repair_other_property_plant_equipment': [
            {'label': 'Repairs and Maintenance - Other Property, Plant and Equipment', 'name': 'repair_maintenance_other_property_plant_equipment'},
        ],
        'taxes_insurance_fees': [
            {'label': 'Taxes, Duties and Licenses', 'name': 'taxes_duties_licenses'},
            {'label': 'Fidelity Bond Premiums', 'name': 'fidelity_bond_premiums'},
            {'label': 'Insurance Expenses', 'name': 'insurance_expenses'},
        ],
        'labor_wages_list': [
            {'label': 'Labor and Wages', 'name': 'labor_wages'},
        ],
        'other_mooe': [
            {'label': 'Advertising Expenses', 'name': 'advertising_expenses'},
            {'label': 'Printing and Publication Expenses', 'name': 'printing_publication_expenses'},
            {'label': 'Representation Expenses', 'name': 'representation_expenses'},
            {'label': 'Transportation and Delivery Expenses', 'name': 'transportation_delivery_expenses'},
            {'label': 'Rent/Lease Expenses', 'name': 'rent/lease_expenses'},
            {'label': 'Membership Dues and contributions to organizations', 'name': 'membership_dues_contribute_to_org'},
            {'label': 'Subscription Expenses', 'name': 'subscription_expenses'},
            {'label': 'Website Maintenance', 'name': 'website_maintenance'},
            {'label': 'Other Maintenance and Operating Expenses', 'name': 'other_maintenance_operating_expenses'},
        ],
        'cap_land': [
            {'label': 'Land', 'name': 'land'},
        ],
        'cap_land_improvements': [
            {'label': 'Land Improvements, Aquaculture Structure', 'name': 'land_improvements_aqua_structure'},
        ],
        'cap_infrastructure_assets': [
            {'label': 'Water Supply Systems', 'name': 'water_supply_systems'},
            {'label': 'Power Supply Systems', 'name': 'power_supply_systems'},
            {'label': 'Other Infrastructure Assets', 'name': 'other_infra_assets'},
        ],
        'cap_bos': [
            {'label': 'Building', 'name': 'bos_building'},
            {'label': 'School Buildings', 'name': 'bos_school_buildings'},
            {'label': 'Hostels and Dormitories', 'name': 'bos_hostels_dorm'},
            {'label': 'Other Structures', 'name': 'other_structures'},
        ],
        'cap_me': [
            {'label': 'Machinery', 'name': 'me_machinery'},
            {'label': 'Office Equipment', 'name': 'me_office_equipment'},
            {'label': 'Information and Communication Technology Equipment', 'name': 'me_ict_equipment'},
            {'label': 'Communication Equipment', 'name': 'me_communication_equipment'},
            {'label': 'Disaster Response and Rescue Equipment', 'name': 'me_drre'},
            {'label': 'Medical Equipment', 'name': 'me_medical_equipment'},
            {'label': 'Printing Equipment', 'name': 'me_printing_equipment'},
            {'label': 'Sports Equipment', 'name': 'me_sports_equipment'},
            {'label': 'Technical and Scientific Equipment', 'name': 'me_technical_scientific_equipment'},
            {'label': 'Other Machinery and Equipment', 'name': 'me_other_machinery_equipment'},
        ],
        'cap_te': [
            {'label': 'Motor Vehicles', 'name': 'te_motor'},
            {'label': 'Other Transportation Equipment', 'name': 'te_other_transpo_equipment'},
        ],
        'cap_ffb': [
            {'label': 'Furniture and Fixtures', 'name': 'ffb_furniture_fixtures'},
            {'label': 'Books', 'name': 'ffb_books'},
        ],
        'cap_cp': [
            {'label': 'Construction in Progress - Land Improvements', 'name': 'cp_land_improvements'},
            {'label': 'Construction in Progress - Infrastructure Assets', 'name': 'cp_infra_assets'},
            {'label': 'Construction in Progress - Buildings and Other Structures', 'name': 'cp_building_other_structures'},
            {'label': 'Construction in Progress - Leased Assets', 'name': 'cp_leased_assets'},
            {'label': 'Construction in Progress - Leased Assets Improvements', 'name': 'cp_leased_assets_improvements'},
        ],
        'cap_ia': [
            {'label': 'Computer Software', 'name': 'ia_computer_software'},
            {'label': 'Websites', 'name': 'ia_websites'},
            {'label': 'Other Tangible Assets', 'name': 'ia_other_tangible_assets'},
        ],
        
    }

    if request.method == 'POST':
        numeric_suffixes = ('_q1', '_q2', '_q3', '_q4', '_total')
        invalid_fields = []
        for key, value in request.POST.items():
            if key == 'csrfmiddlewaretoken':
                continue
            if key in ('prepared_by', 'certified_by', 'approved_by'):
                continue
            if any(key.endswith(suf) for suf in numeric_suffixes):
                if value.strip() == '':
                    continue
                try:
                    num = Decimal(value)
                    if num < 0:
                        invalid_fields.append(key)
                except Exception:
                    invalid_fields.append(key)
        if invalid_fields:
            messages.error(request, f"Please enter valid non-negative numbers for: {', '.join(invalid_fields[:5])}{'…' if len(invalid_fields) > 5 else ''}")
            return render(request, "end_user_app/department_pre_form.html", context)

        payload = {k: v for k, v in request.POST.items() if k not in ['csrfmiddlewaretoken', 'prepared_by', 'certified_by', 'approved_by']}

        # Link to the latest Budget Allocation for this department, if any
        # dept_alloc = BudgetAllocation.objects.filter(department=getattr(request.user, 'department', '')).order_by('-allocated_at').first()
        
        dept_alloc = BudgetAllocation.objects.filter(
            department=getattr(request.user, 'department', ''),
            id=pk
        ).first()

        if not dept_alloc:
            messages.error(request, "No budget allocation found for this department.")
            return render(request, "end_user_app/department_pre_form.html", context)

        print(dept_alloc)

        pre = DepartmentPRE.objects.create(
            submitted_by=request.user,
            department=getattr(request.user, 'department', ''),
            data=payload,
            prepared_by_name=request.user.username,
            certified_by_name= None,
            approved_by_name= None,
            budget_allocation=dept_alloc,
        )
        
        dept_alloc.is_compiled = True
        dept_alloc.save(update_fields=['is_compiled'])

        messages.success(request, "PRE submitted successfully.")
        return render(request, "end_user_app/department_pre_form.html", {'pre_id': pre.id, 'success': True, **context}) 
        # return redirect('preview_pre', pk=pre.id)
    
    return render(request, "end_user_app/department_pre_form.html", context)

# @role_required('end_user', login_url='/')
# def department_pre_page(request):
#     user = request.user
#     user_dept = user.department
#     # has_budget = BudgetAllocation.objects.filter(department=request.user.department).exists()
    
#     is_has_budget = BudgetAllocation.objects.filter(department=user_dept, is_compiled=False).exists()
    
#     if is_has_budget:
#         has_budget = True
#     else:
#         has_budget = False
        
#     budget_allocations = BudgetAllocation.objects.filter(department=request.user.department, is_compiled=False).select_related('approved_budget').order_by('-allocated_at')

#     # Load submitted PREs for this user/department
#     pres = DepartmentPRE.objects.filter(submitted_by=user).order_by('-created_at')

#     return render(request, "end_user_app/department_pre_page.html", {
#         'has_budget': has_budget,
#         'pres': pres,
#         'budget_allocations': budget_allocations,
#     })

@role_required('end_user', login_url='/')
def department_pre_page(request):
    """Main PRE page showing budget allocations and submitted PREs"""
    user = request.user
    
    # Get budget allocations for current user
    budget_allocations = NewBudgetAllocation.objects.filter(
        end_user=user,
        is_active=True
    ).select_related('approved_budget').order_by('-allocated_at')
    
    has_budget = budget_allocations.exists()
    
    # Get submitted PREs
    pres = NewDepartmentPRE.objects.filter(
        submitted_by=user
    ).order_by('-created_at')
    
    # NEW: Count partially approved PREs with PDF
    partially_approved_count = pres.filter(
        status='Partially Approved',
        partially_approved_pdf__isnull=False
    ).count()
    
    context = {
        'has_budget': has_budget,
        'budget_allocations': budget_allocations,
        'pres': pres,
        'partially_approved_count': partially_approved_count,  # NEW
    }
    
    return render(request, 'end_user_app/department_pre_page.html', context)

@role_required('end_user', login_url='/')
def upload_pre(request, allocation_id):
    """Handle PRE Excel upload and supporting documents with draft support"""
    allocation = get_object_or_404(
        NewBudgetAllocation.objects.select_related('approved_budget'),
        id=allocation_id,
        end_user=request.user
    )
    
    # Check if PRE already exists for this allocation
    existing_pre = NewDepartmentPRE.objects.filter(
        budget_allocation=allocation
    ).first()
    
    if existing_pre:
        messages.warning(request, 
            f"PRE already exists for this allocation. "
            f"Status: {existing_pre.status}")
        return redirect('department_pre_page')
    
    # Get or create draft
    draft, created = PREDraft.objects.get_or_create(
        user=request.user,
        budget_allocation=allocation
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'upload_pre':
            # Handle PRE file upload
            pre_file = request.FILES.get('pre_document')
            
            if not pre_file:
                messages.error(request, "Please select a PRE file to upload.")
                return redirect('upload_pre', allocation_id=allocation_id)
            
            if not pre_file.name.endswith('.xlsx'):
                messages.error(request, "Only .xlsx Excel files are accepted for PRE.")
                return redirect('upload_pre', allocation_id=allocation_id)
            
            # Save to draft
            if draft.pre_file:
                # Delete old file if exists
                draft.pre_file.delete(save=False)
            
            draft.pre_file = pre_file
            draft.pre_filename = pre_file.name
            draft.save()
            
            messages.success(request, f"PRE file '{pre_file.name}' saved as draft.")
            return redirect('upload_pre', allocation_id=allocation_id)
        
        elif action == 'upload_supporting':
            # Handle supporting documents upload
            supporting_docs = request.FILES.getlist('supporting_documents')
            
            if not supporting_docs:
                messages.error(request, "Please select at least one supporting document.")
                return redirect('upload_pre', allocation_id=allocation_id)
            
            # Save each supporting document
            count = 0
            for doc in supporting_docs:
                # Check if document with same name already exists
                existing = draft.supporting_documents.filter(file_name=doc.name).first()
                if existing:
                    messages.warning(request, f"File '{doc.name}' already exists. Skipped.")
                    continue
                
                PREDraftSupportingDocument.objects.create(
                    draft=draft,
                    document=doc,
                    file_name=doc.name,
                    file_size=doc.size
                )
                count += 1
            
            if count > 0:
                messages.success(request, f"{count} supporting document(s) saved as draft.")
            return redirect('upload_pre', allocation_id=allocation_id)
        
        elif action == 'remove_supporting':
            # Remove supporting document
            doc_id = request.POST.get('doc_id')
            try:
                doc = draft.supporting_documents.get(id=doc_id)
                doc_name = doc.file_name
                doc.document.delete(save=False)
                doc.delete()
                messages.success(request, f"Removed '{doc_name}'")
            except PREDraftSupportingDocument.DoesNotExist:
                messages.error(request, "Document not found.")
            return redirect('upload_pre', allocation_id=allocation_id)
        
        elif action == 'continue':
            # Validate and continue to preview
            if not draft.pre_file:
                messages.error(request, "Please upload a PRE Excel file first.")
                return redirect('upload_pre', allocation_id=allocation_id)
            
            # Parse Excel file
            try:
                result = parse_pre_excel(draft.pre_file.path)
                
                if not result['success']:
                    messages.error(request, 
                        f"Error parsing PRE file: {', '.join(result['errors'])}")
                    return redirect('upload_pre', allocation_id=allocation_id)
                
                # Validate fiscal year
                if result['fiscal_year']:
                    if result['fiscal_year'] != allocation.approved_budget.fiscal_year:
                        messages.error(request, 
                            f"PRE fiscal year ({result['fiscal_year']}) does not match "
                            f"budget allocation fiscal year ({allocation.approved_budget.fiscal_year}).")
                        return redirect('upload_pre', allocation_id=allocation_id)
                
                # Validate grand total
                if result['grand_total'] > allocation.remaining_balance:
                    messages.error(request, 
                        f"PRE total amount (₱{result['grand_total']:,.2f}) exceeds "
                        f"remaining budget allocation (₱{allocation.remaining_balance:,.2f}).")
                    return redirect('upload_pre', allocation_id=allocation_id)
                
                # Store data in session for preview
                request.session['pre_upload_data'] = {
                    'allocation_id': allocation_id,
                    'draft_id': draft.id,
                    'extracted_data': json.dumps(result['data'], default=str),
                    'grand_total': str(result['grand_total']),
                    'fiscal_year': result['fiscal_year'],
                    'pre_filename': draft.pre_filename,
                    'validation_warnings': result.get('validation_warnings', [])
                }
                
                messages.success(request, "Files validated successfully. Please review the extracted data.")
                return redirect('preview_pre')
                
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                return redirect('upload_pre', allocation_id=allocation_id)
        
        elif action == 'clear_draft':
            # Clear all draft files
            if draft.pre_file:
                draft.pre_file.delete(save=False)
            for doc in draft.supporting_documents.all():
                doc.document.delete(save=False)
                doc.delete()
            draft.delete()
            messages.info(request, "Draft cleared.")
            return redirect('upload_pre', allocation_id=allocation_id)
    
    context = {
        'allocation': allocation,
        'draft': draft,
    }
    
    return render(request, 'end_user_app/upload_pre.html', context)


def create_pre_line_items(pre, extracted_data):
    """
    Create PRELineItem records from extracted data
    
    Args:
        pre: NewDepartmentPRE instance
        extracted_data: Dict with categories (receipts, personnel, mooe, capital)
    
    Returns:
        int: Number of line items created
    """
    from apps.budgets.models import PRELineItem, PRECategory, PRESubCategory
    
    line_items_created = 0
    
    # Category mapping
    category_mapping = {
        'receipts': ('PERSONNEL', 'Receipts'),  # Or create a separate category
        'personnel': ('PERSONNEL', 'Personnel Services'),
        'mooe': ('MOOE', 'Maintenance and Other Operating Expenses'),
        'capital': ('CAPITAL', 'Capital Outlays'),
    }
    
    for section_key, items_list in extracted_data.items():
        if not items_list:
            continue
        
        # Get or create category
        category_type, category_name = category_mapping.get(
            section_key, 
            ('MOOE', section_key.title())
        )
        
        category, _ = PRECategory.objects.get_or_create(
            category_type=category_type,
            defaults={
                'name': category_name,
                'code': category_type[:4].upper(),
                'is_active': True,
                'sort_order': {'receipts': 1, 'personnel': 2, 'mooe': 3, 'capital': 4}.get(section_key, 5)
            }
        )
        
        # Create line items
        for item_data in items_list:
            item_name = item_data.get('item_name', 'Unknown Item')
            
            # Get subcategory from nested data (for MOOE and Capital)
            subcategory = None
            subcategory_name = item_data.get('category')  # MOOE/Capital have subcategories
            
            if subcategory_name:
                subcategory, _ = PRESubCategory.objects.get_or_create(
                    category=category,
                    name=subcategory_name.replace('_', ' ').title(),
                    defaults={
                        'code': subcategory_name[:10].upper(),
                        'is_active': True,
                    }
                )
            
            # Create line item
            PRELineItem.objects.create(
                pre=pre,
                category=category,
                subcategory=subcategory,
                item_name=item_name,
                q1_amount=Decimal(str(item_data.get('q1', 0))),
                q2_amount=Decimal(str(item_data.get('q2', 0))),
                q3_amount=Decimal(str(item_data.get('q3', 0))),
                q4_amount=Decimal(str(item_data.get('q4', 0))),
            )
            
            line_items_created += 1
    
    print(f"✅ Created {line_items_created} PRELineItem records for PRE {pre.id}")
    return line_items_created


@role_required('end_user', login_url='/')
def preview_pre(request):
    """Preview extracted PRE data before final submission"""
    
    # Get data from session
    upload_data = request.session.get('pre_upload_data') 
    supporting_docs_info = request.session.get('supporting_docs_info', [])
    
    if not upload_data:
        messages.error(request, "No upload data found. Please upload PRE again.")
        return redirect('department_pre_page')
    
    # Parse extracted data
    extracted_data = json.loads(upload_data['extracted_data'])
    grand_total = Decimal(upload_data['grand_total'])
    validation_warnings = upload_data.get('validation_warnings', [])
    
    # Calculate section totals
    def calculate_section_totals(items):
        totals = {
            'q1': Decimal('0'),
            'q2': Decimal('0'),
            'q3': Decimal('0'),
            'q4': Decimal('0'),
            'total': Decimal('0')
        }
        for item in items:
            totals['q1'] += Decimal(str(item.get('q1', 0)))
            totals['q2'] += Decimal(str(item.get('q2', 0)))
            totals['q3'] += Decimal(str(item.get('q3', 0)))
            totals['q4'] += Decimal(str(item.get('q4', 0)))
            totals['total'] += Decimal(str(item.get('total', 0)))
        return totals
    
    receipts_total = calculate_section_totals(extracted_data.get('receipts', []))
    personnel_total = calculate_section_totals(extracted_data.get('personnel', []))
    mooe_total = calculate_section_totals(extracted_data.get('mooe', []))
    capital_total = calculate_section_totals(extracted_data.get('capital', []))
    
    # Get allocation
    allocation = get_object_or_404(
        NewBudgetAllocation.objects.select_related('approved_budget'),
        id=upload_data['allocation_id'],
        end_user=request.user
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'cancel':
            # Clean up temporary files
            temp_file_path = upload_data.get('temp_file_path')
            if temp_file_path and default_storage.exists(temp_file_path):
                default_storage.delete(temp_file_path)
            
            temp_doc_paths = request.session.get('temp_doc_paths', [])
            for path in temp_doc_paths:
                if default_storage.exists(path):
                    default_storage.delete(path)
                    
            draft_id = upload_data.get('draft_id')
            try:
                draft = PREDraft.objects.get(id=draft_id, user=request.user)
                if draft.pre_file:
                    draft.pre_file.delete(save=False)
                for doc in draft.supporting_documents.all():
                    doc.document.delete(save=False)
                    doc.delete()
                draft.delete()
            except:
                pass
            
            # Clear session
            request.session.pop('pre_upload_data', None)
            request.session.pop('supporting_docs_info', None)
            request.session.pop('temp_doc_paths', None)
            
            messages.info(request, "Upload cancelled.")
            return redirect('department_pre_page')
        
        elif action == 'submit':
            try:
                from django.db import transaction
                
                # Start transaction - everything succeeds or everything fails
                with transaction.atomic():
                    draft_id = upload_data.get('draft_id')
                    draft = PREDraft.objects.get(id=draft_id, user=request.user)
                    
                    # ✅ Recalculate grand total from extracted data
                    recalculated_total = Decimal('0')
                    for category in ['receipts', 'personnel', 'mooe', 'capital']:
                        for item in extracted_data.get(category, []):
                            item_total = (
                                Decimal(str(item.get('q1', 0))) +
                                Decimal(str(item.get('q2', 0))) +
                                Decimal(str(item.get('q3', 0))) +
                                Decimal(str(item.get('q4', 0)))
                            )
                            recalculated_total += item_total
                    
                    # ✅ Use recalculated total
                    print(f"📊 Session grand_total: ₱{grand_total:,.2f}")
                    print(f"📊 Recalculated total: ₱{recalculated_total:,.2f}")
                    
                    if abs(grand_total - recalculated_total) > Decimal('0.01'):
                        print(f"⚠️ Total mismatch! Using recalculated: ₱{recalculated_total:,.2f}")
                        print(f"   Session: ₱{grand_total:,.2f}")
                        print(f"   Using recalculated value.")
                    
                    final_total = recalculated_total
                    
                    # 1. Create DepartmentPRE record
                    pre = NewDepartmentPRE.objects.create(
                        submitted_by=request.user,
                        department=allocation.department,
                        budget_allocation=allocation,
                        fiscal_year=allocation.approved_budget.fiscal_year,
                        total_amount=final_total,
                        status='Pending',
                        is_valid=True,
                        submitted_at=timezone.now(),
                        prepared_by_name=request.user.get_full_name(),
                    )
                    
                    # 2. Move PRE file from draft
                    if draft.pre_file:
                        from django.core.files.base import ContentFile
                        with draft.pre_file.open('rb') as f:
                            pre.uploaded_excel_file.save(
                                draft.pre_filename, 
                                ContentFile(f.read()), 
                                save=True
                            )
                    
                    # 3. 🔥 CREATE LINE ITEMS FROM EXTRACTED DATA
                    line_items_created = create_pre_line_items(pre, extracted_data)
                    
                    if line_items_created == 0:
                        raise Exception("No line items were created from the PRE data")
                    
                    # 4. Mark draft as submitted
                    draft.is_submitted = True
                    draft.save()
                
                # Move PRE file from temp to permanent location
                # temp_file_path = upload_data['temp_file_path']
                # if default_storage.exists(temp_file_path):
                #     # Read temp file
                #     with default_storage.open(temp_file_path, 'rb') as temp_file:
                #         # Save to PRE's uploaded_excel_file field
                #         from django.core.files.base import ContentFile
                #         pre.uploaded_excel_file.save(
                #             upload_data['pre_filename'],
                #             ContentFile(temp_file.read()),
                #             save=True
                #         )
                #     # Delete temp file
                #     default_storage.delete(temp_file_path)
                
                # Save supporting documents
                temp_doc_paths = request.session.get('temp_doc_paths', [])
                for idx, temp_path in enumerate(temp_doc_paths):
                    if default_storage.exists(temp_path):
                        with default_storage.open(temp_path, 'rb') as temp_doc:
                            # Create SupportingDocument instance
                            # Note: You'll need to create this model or adjust based on your structure
                            from django.core.files.base import ContentFile
                            doc_info = supporting_docs_info[idx]
                            
                            # Save document (adjust this based on your model structure)
                            # For now, we'll assume you want to store it related to the PRE
                            
                        default_storage.delete(temp_path)
                
                # Clear session
                request.session.pop('pre_upload_data', None)
                request.session.pop('supporting_docs_info', None)
                request.session.pop('temp_doc_paths', None)
                request.session.pop('pre_upload_data', None)
                
                messages.success(request, 
                    f"PRE submitted successfully! Total amount: ₱{grand_total:,.2f}. "
                    f"Status: Pending Review")
                
                # Log audit trail
                log_audit_trail(
                    request=request,
                    action='CREATE',
                    model_name='DepartmentPRE',
                    record_id=pre.id,
                    detail=f'Created PRE with {line_items_created} line items, Total: ₱{grand_total:,.2f}'
                )
                
                draft.delete()  # Remove draft after submission
                return redirect('department_pre_page')
                
            except Exception as e:
                messages.error(request, f"Error submitting PRE: {str(e)}")
                print(f"PRE submission error: {e}")
                import traceback
                traceback.print_exc()
                return redirect('preview_pre')
    
    context = {
        'allocation': allocation,
        'extracted_data': extracted_data,
        'grand_total': grand_total,
        'fiscal_year': upload_data['fiscal_year'],
        'pre_filename': upload_data['pre_filename'],
        'supporting_docs': supporting_docs_info,
        'receipts_total': receipts_total,
        'personnel_total': personnel_total,
        'mooe_total': mooe_total,
        'capital_total': capital_total,
        'validation_warnings': validation_warnings,
    }
    
    return render(request, 'end_user_app/preview_pre.html', context)


@role_required('end_user', login_url='/')
def download_pre_template(request):
    """Download blank PRE template"""
    # Path to your PRE template file
    template_path = os.path.join(settings.MEDIA_ROOT, 'templates', 'PRE_Template.xlsx')
    
    if os.path.exists(template_path):
        from django.http import FileResponse
        response = FileResponse(open(template_path, 'rb'))
        response['Content-Disposition'] = 'attachment; filename="PRE_Template.xlsx"'
        return response
    else:
        messages.error(request, "PRE template file not found.")
        return redirect('department_pre_page')
    
@role_required('end_user', login_url='/')
def view_pre_detail(request, pre_id):
    """View PRE details (placeholder for now)"""
    pre = get_object_or_404(
        NewDepartmentPRE.objects.select_related('budget_allocation',
            'budget_allocation__approved_budget'),
        id=pre_id,
        submitted_by=request.user
    )
    
    context = {
        'pre': pre,
    }
    
    return render(request, 'end_user_app/view_pre_detail.html', context)

# @role_required('end_user', login_url='/')
# def preview_pre(request, pk: int):
#     pre = get_object_or_404(DepartmentPRE.objects.select_related('submitted_by'), pk=pk)
#     # Security: ensure user can only view their own PRE (unless you allow admins)
#     if pre.submitted_by != request.user:
#         messages.error(request, "You do not have permission to view this PRE.")
#         return redirect('department_pre_page')

#     # Build grouped sections for Excel-like table
#     def to_decimal(value):
#         try:
#             return Decimal(str(value)) if value not in (None, "") else Decimal('0')
#         except Exception:
#             return Decimal('0')

#     payload = pre.data or {}

#     # Define mapping of sections and items (subset reflecting provided screenshot)
#     sections_spec = [
#         {
#             'title': 'Personnel Services',
#             'color_class': 'bg-yellow-100',
#             'items': [
#                 {'label': 'Basic Salary', 'name': 'basic_salary'},
#                 {'label': 'Honoraria', 'name': 'honoraria'},
#                 {'label': 'Overtime Pay', 'name': 'overtime_pay'},
#             ]
#         },
#         {
#             'title': 'Maintenance and Other Operating Expenses',
#             'color_class': 'bg-blue-100',
#             'items': [
#                 {'label': 'Travelling Expenses', 'is_group': True},
#                 {'label': 'Travelling expenses-local', 'name': 'travel_local', 'indent': True},
#                 {'label': 'Travelling Expenses-foreign', 'name': 'travel_foreign', 'indent': True},
#                 {'label': 'Training and Scholarship expenses', 'is_group': True},
#                 {'label': 'Training Expenses', 'name': 'training_expenses', 'indent': True},
#                 {'label': 'Supplies and materials expenses', 'is_group': True},
#                 {'label': 'Office Supplies Expenses', 'name': 'office_supplies_expenses', 'indent': True},
#                 {'label': 'Accountable Form Expenses', 'name': 'accountable_form_expenses', 'indent': True},
#                 {'label': 'Agricultural and Marine Supplies Expenses', 'name': 'agri_marine_supplies_expenses', 'indent': True},
#                 {'label': 'Drugs and Medicines', 'name': 'drugs_medicines', 'indent': True},
#                 {'label': 'Medical, Dental & Laboratory Supplies Expenses', 'name': 'med_dental_lab_supplies_expenses', 'indent': True},
#                 {'label': 'Food Supplies Expenses', 'name': 'food_supplies_expenses', 'indent': True},
#                 {'label': 'Fuel, Oil and Lubricants Expenses', 'name': 'fuel_oil_lubricants_expenses', 'indent': True},
#                 {'label': 'Textbooks and Instructional Materials Expenses', 'name': 'textbooks_instructional_materials_expenses', 'indent': True},
#                 {'label': 'Construction Materials Expenses', 'name': 'construction_material_expenses', 'indent': True},
#                 {'label': 'Other Supplies & Materials Expenses', 'name': 'other_supplies_materials_expenses', 'indent': True},
#                 {'label': 'Semi-expendable Machinery Equipment', 'is_group': True},
#                 {'label': 'Machinery', 'name': 'semee_machinery', 'indent': True},
#                 {'label': 'Office Equipment', 'name': 'semee_office_equipment', 'indent': True},
#                 {'label': 'Information and Communications Technology Equipment', 'name': 'semee_information_communication', 'indent': True},
#                 {'label': 'Communications Equipment', 'name': 'semee_communications_equipment', 'indent': True},
#                 {'label': 'Disaster Response and Rescue Equipment', 'name': 'semee_drr_equipment', 'indent': True},
#                 {'label': 'Medical Equipment', 'name': 'semee_medical_equipment', 'indent': True},
#                 {'label': 'Printing Equipment', 'name': 'semee_printing_equipment', 'indent': True},
#                 {'label': 'Sports Equipment', 'name': 'semee_sports_equipment', 'indent': True},
#                 {'label': 'Technical and Scientific Equipment', 'name': 'semee_technical_scientific_equipment', 'indent': True},
#                 {'label': 'ICT Equipment', 'name': 'semee_ict_equipment', 'indent': True},
#                 {'label': 'Other Machinery and Equipment', 'name': 'semee_other_machinery_equipment', 'indent': True},
#                 {'label': 'Semi-expendable Furnitures and Fixtures', 'is_group': True},
#                 {'label': 'Furniture and Fixtures', 'name': 'furniture_fixtures', 'indent': True},
#                 {'label': 'Books', 'name': 'books', 'indent': True},
#                 {'label': 'Utility Expenses', 'is_group': True},
#                 {'label': 'Water Expenses', 'name': 'water_expenses', 'indent': True},
#                 {'label': 'Electricity Expenses', 'name': 'electricity_expenses', 'indent': True},
#                 {'label': 'Communication Expenses', 'is_group': True},
#                 {'label': 'Postage and Courier Services', 'name': 'postage_courier_services', 'indent': True},
#                 {'label': 'Telephone Expenses', 'name': 'telephone_expenses', 'indent': True},
#                 {'label': 'Telephone Expenses (Landline)', 'name': 'telephone_expenses_landline', 'indent': True},
#                 {'label': 'Internet Subscription Expenses', 'name': 'internet_subscription_expenses', 'indent': True},
#                 {'label': 'Cable, Satellite, Telegraph & Radio Expenses', 'name': 'cable_satellite_telegraph_radio_expenses', 'indent': True},
#                 {'label': 'Awards/Rewards and Prizes', 'is_group': True},
#                 {'label': 'Awards/Rewards Expenses', 'name': 'awards_rewards_expenses', 'indent': True},
#                 {'label': 'Prizes', 'name': 'prizes', 'indent': True},
#                 {'label': 'Survey, Research, Exploration, and Development Expenses', 'is_group': True},
#                 {'label': 'Survey Expenses', 'name': 'survey_expenses', 'indent': True},
#                 {'label': 'Survey, Research, Exploration, and Development expenses', 'name': 'survey_research_exploration_development_expenses', 'indent': True},
#                 {'label': 'Professional Services', 'is_group': True},
#                 {'label': 'Legal Services', 'name': 'legal_services', 'indent': True},
#                 {'label': 'Auditing Services', 'name': 'auditing_services', 'indent': True},
#                 {'label': 'Consultancy Services', 'name': 'consultancy_services', 'indent': True},
#                 {'label': 'Other Professional Services', 'name': 'other_professional_servies', 'indent': True},
#                 {'label': 'General Services', 'is_group': True},
#                 {'label': 'Security Services', 'name': 'security_services', 'indent': True},
#                 {'label': 'Janitorial Services', 'name': 'janitorial_services', 'indent': True},
#                 {'label': 'Other General Services', 'name': 'other_general_services', 'indent': True},
#                 {'label': 'Environment/Sanitary Services', 'name': 'environment/sanitary_services', 'indent': True},
#                 {'label': 'Repair and Maintenance', 'is_group': True},
#                 {'label': 'Repair & Maintenance - Land Improvements', 'name': 'repair_maintenance_land_improvements', 'indent': True},
#                 {'label': 'Repair & Maintenance - Buildings and Structures', 'is_group': True},
#                 {'label': 'Buildings', 'name': 'buildings', 'indent': True},
#                 {'label': 'School Buildings', 'name': 'school_buildings', 'indent': True},
#                 {'label': 'Hostels and Dormitories', 'name': 'hostel_dormitories', 'indent': True},
#                 {'label': 'Other Structures', 'name': 'other_structures', 'indent': True},
#                 {'label': 'Repairs and Maintenance - Machinery and Equipment', 'is_group': True},
#                 {'label': 'Machinery', 'name': 'repair_maintenance_machinery', 'indent': True},
#                 {'label': 'Office Equipment', 'name': 'repair_maintenance_office_equipment', 'indent': True},
#                 {'label': 'ICT Equipment', 'name': 'repair_maintenance_ict_equipment', 'indent': True},
#                 {'label': 'Agricultural and Forestry Equipment', 'name': 'repair_maintenance_agri_forestry_equipment', 'indent': True},
#                 {'label': 'Marine and Fishery Equipment', 'name': 'repair_maintenance_marine_fishery_equipment', 'indent': True},
#                 {'label': 'Airport Equipment', 'name': 'repair_maintenance_airport_equipment', 'indent': True},
#                 {'label': 'Communication Equipment', 'name': 'repair_maintenance_communication_equipment', 'indent': True},
#                 {'label': 'Disaster, Response and Rescue Equipment', 'name': 'repair_maintenance_drre_equipment', 'indent': True},
#                 {'label': 'Medical Equipment', 'name': 'repair_maintenance_medical_equipment', 'indent': True},
#                 {'label': 'Printing Equipment', 'name': 'repair_maintenance_printing_equipment', 'indent': True},
#                 {'label': 'Sports Equipment', 'name': 'repair_maintenance_sports_equipment', 'indent': True},
#                 {'label': 'Technical and Scientific Equipment', 'name': 'repair_maintenance_technical_scientific_equipment', 'indent': True},
#                 {'label': 'Other Machinery and Equipment', 'name': 'repair_maintenance_other_machinery_equipment', 'indent': True},
#                 {'label': 'Repairs and Maintenance - Transportation Equipment', 'is_group': True},
#                 {'label': 'Motor Vehicles', 'name': 'repair_maintenance_motor', 'indent': True},
#                 {'label': 'Other Transportation Equipment', 'name': 'repair_maintenance_other_transportation_equipment', 'indent': True},
#                 {'label': 'Repairs and Maintenance - Furniture & Fixtures', 'name': 'repair_maintenance_furniture_fixtures'},
#                 {'label': 'Repairs and Maintenance - Semi-Expendable Machinery and Equipment', 'name': 'repair_maintenance_semi_expendable_machinery_equipment'},
#                 {'label': 'Repairs and Maintenance - Other Property, Plant and Equipment', 'name': 'repair_maintenance_other_property_plant_equipment'},
#                 {'label': 'Taxes, Insurance Premiums and Other Fees', 'is_group': True},
#                 {'label': 'Taxes, Duties and Licenses', 'name': 'taxes_duties_licenses', 'indent': True},
#                 {'label': 'Fidelity Bond Premiums', 'name': 'fidelity_bond_premiums', 'indent': True},
#                 {'label': 'Insurance Expenses', 'name': 'insurance_expenses', 'indent': True},
#                 {'label': 'Labor and Wages', 'is_group': True},
#                 {'label': 'Labor and Wages', 'name': 'labor_wages', 'indent': True},
#                 {'label': 'Other Maintenance and Operating Expenses', 'is_group': True},
#                 {'label': 'Advertising Expenses', 'name': 'advertising_expenses', 'indent': True},
#                 {'label': 'Printing and Publication Expenses', 'name': 'printing_publication_expenses', 'indent': True},
#                 {'label': 'Representation Expenses', 'name': 'representation_expenses', 'indent': True},
#                 {'label': 'Transportation and Delivery Expenses', 'name': 'transportation_delivery_expenses', 'indent': True},
#                 {'label': 'Rent/Lease Expenses', 'name': 'rent/lease_expenses', 'indent': True},
#                 {'label': 'Membership Dues and contributions to organizations', 'name': 'membership_dues_contribute_to_org', 'indent': True},
#                 {'label': 'Subscription Expenses', 'name': 'subscription_expenses', 'indent': True},
#                 {'label': 'Website Maintenance', 'name': 'website_maintenance', 'indent': True},
#                 {'label': 'Other Maintenance and Operating Expenses', 'name': 'other_maintenance_operating_expenses', 'indent': True},
#             ]
#         },
#         {
#             'title': 'Capital Outlays',
#             'color_class': 'bg-green-100',
#             'items': [
#                 # Placeholder; extend as needed
#                 {'label': 'Land', 'is_group': True},
#                 {'label': 'Land', 'name': 'land', 'indent': True},
#                 {'label': 'Land Improvements', 'is_group': True},
#                 {'label': 'Land Improvements, Aquaculture Structure', 'name': 'land_improvements_aqua_structure', 'indent': True},
#                 {'label': 'Infrastructure Assets', 'is_group': True},
#                 {'label': 'Water Supply Systems', 'name': 'water_supply_systems', 'indent': True},
#                 {'label': 'Power Supply Systems', 'name': 'power_supply_systems', 'indent': True},
#                 {'label': 'Other Infrastructure Assets', 'name': 'other_infra_assets', 'indent': True},
#                 {'label': 'Building and Other Structures', 'is_group': True},
#                 {'label': 'Building', 'name': 'bos_building', 'indent': True},
#                 {'label': 'School Buildings', 'name': 'bos_school_buildings', 'indent': True},
#                 {'label': 'Hostels and Dormitories', 'name': 'bos_hostels_dorm', 'indent': True},
#                 {'label': 'Other Structures', 'name': 'other_structures', 'indent': True},
#                 {'label': 'Machinery and Equipment', 'is_group': True},
#                 {'label': 'Machinery', 'name': 'me_machinery', 'indent': True},
#                 {'label': 'Office Equipment', 'name': 'me_office_equipment', 'indent': True},
#                 {'label': 'Information and Communication Technology Equipment', 'name': 'me_ict_equipment', 'indent': True},
#                 {'label': 'Communication Equipment', 'name': 'me_communication_equipment', 'indent': True},
#                 {'label': 'Disaster Response and Rescue Equipment', 'name': 'me_drre', 'indent': True},
#                 {'label': 'Medical Equipment', 'name': 'me_medical_equipment', 'indent': True},
#                 {'label': 'Printing Equipment', 'name': 'me_printing_equipment', 'indent': True},
#                 {'label': 'Sports Equipment', 'name': 'me_sports_equipment', 'indent': True},
#                 {'label': 'Technical and Scientific Equipment', 'name': 'me_technical_scientific_equipment', 'indent': True},
#                 {'label': 'Other Machinery and Equipment', 'name': 'me_other_machinery_equipment', 'indent': True},
#                 {'label': 'Transportation Equipment', 'is_group': True},
#                 {'label': 'Motor Vehicles', 'name': 'te_motor', 'indent': True},
#                 {'label': 'Other Transportation Equipment', 'name': 'te_other_transpo_equipment', 'indent': True},
#                 {'label': 'Furniture, Fixtures and Books', 'is_group': True},
#                 {'label': 'Furniture and Fixtures', 'name': 'ffb_furniture_fixtures', 'indent': True},
#                 {'label': 'Books', 'name': 'ffb_books', 'indent': True},
#                 {'label': 'Construction in Progress', 'is_group': True},
#                 {'label': 'Construction in Progress - Land Improvements', 'name': 'cp_land_improvements', 'indent': True},
#                 {'label': 'Construction in Progress - Infrastructure Assets', 'name': 'cp_infra_assets', 'indent': True},
#                 {'label': 'Construction in Progress - Buildings and Other Structures', 'name': 'cp_building_other_structures', 'indent': True},
#                 {'label': 'Construction in Progress - Leased Assets', 'name': 'cp_leased_assets', 'indent': True},
#                 {'label': 'Construction in Progress - Leased Assets Improvements', 'name': 'cp_leased_assets_improvements', 'indent': True},
#                 {'label': 'Intangible Assets', 'is_group': True},
#                 {'label': 'Computer Software', 'name': 'ia_computer_software', 'indent': True},
#                 {'label': 'Websites', 'name': 'ia_websites', 'indent': True},
#                 {'label': 'Other Tangible Assets', 'name': 'ia_other_tangible_assets', 'indent': True},
                
#             ]
#         },
#     ]

#     sections = []
#     for spec in sections_spec:
#         items = []
#         for row in spec['items']:
#             if row.get('is_group'):
#                 items.append({'is_group': True, 'label': row['label']})
#             else:
#                 name = row['name']
#                 q1 = to_decimal(payload.get(f"{name}_q1"))
#                 q2 = to_decimal(payload.get(f"{name}_q2"))
#                 q3 = to_decimal(payload.get(f"{name}_q3"))
#                 q4 = to_decimal(payload.get(f"{name}_q4"))
#                 total = q1 + q2 + q3 + q4
#                 items.append({
#                     'label': row['label'],
#                     'indent': row.get('indent', False),
#                     'q1': q1,
#                     'q2': q2,
#                     'q3': q3,
#                     'q4': q4,
#                     'total': total,
#                 })

#         # Compute section totals excluding group rows
#         total_q1 = sum((it['q1'] for it in items if not it.get('is_group')), Decimal('0'))
#         total_q2 = sum((it['q2'] for it in items if not it.get('is_group')), Decimal('0'))
#         total_q3 = sum((it['q3'] for it in items if not it.get('is_group')), Decimal('0'))
#         total_q4 = sum((it['q4'] for it in items if not it.get('is_group')), Decimal('0'))
#         total_overall = sum((it['total'] for it in items if not it.get('is_group')), Decimal('0'))

#         sections.append({
#             'title': spec['title'],
#             'color_class': spec['color_class'],
#             'items': items,
#             'total_q1': total_q1,
#             'total_q2': total_q2,
#             'total_q3': total_q3,
#             'total_q4': total_q4,
#             'total_overall': total_overall,
#         })

#     # Compute grand totals across all sections
#     grand_total_q1 = sum((sec['total_q1'] for sec in sections), Decimal('0'))
#     grand_total_q2 = sum((sec['total_q2'] for sec in sections), Decimal('0'))
#     grand_total_q3 = sum((sec['total_q3'] for sec in sections), Decimal('0'))
#     grand_total_q4 = sum((sec['total_q4'] for sec in sections), Decimal('0'))
#     grand_total_overall = sum((sec['total_overall'] for sec in sections), Decimal('0'))

#     # Pass the grouped table, raw data and signatories
#     context = {
#         'pre': pre,
#         'data': payload,
#         'sections': sections,
#         'grand_total_q1': grand_total_q1,
#         'grand_total_q2': grand_total_q2,
#         'grand_total_q3': grand_total_q3,
#         'grand_total_q4': grand_total_q4,
#         'grand_total_overall': grand_total_overall,
#         'prepared_by': pre.prepared_by_name,
#         'certified_by': pre.certified_by_name,
#         'approved_by': pre.approved_by_name,
#     }
#     return render(request, "end_user_app/preview_pre.html", context)
    
# Activity Design Form Logic
@role_required('end_user', login_url='/')
def activity_design_form(request):
    budget_allocations = BudgetAllocation.objects.select_related('approved_budget').filter(department=getattr(request.user, 'department', ''))
    
    approved_pres = DepartmentPRE.objects.filter(
            submitted_by=request.user,
            approved_by_approving_officer=True,
            approved_by_admin=True,
        )
    
    source_of_fund_options = build_pre_source_options(approved_pres)
    
    if request.method == 'POST':    
        # Initialize variable to avoid scope issues
        item_key = None
        quarter = None
        source_amount_decimal = None
        source_of_fund_display = None
        source_pre_instance = None
        total_amount = Decimal(request.POST.get('total_amount', '0'))
        budget_allocation_instance = None
        allocations_to_make = []
        
        ba_id = request.POST.get('budget_allocation')
        if ba_id:
            try:
                budget_allocation_instance = BudgetAllocation.objects.select_related('approved_budget').get(id=ba_id)
            except BudgetAllocation.DoesNotExist:
                messages.error(request, "No such budget allocation for this user.")
                budget_allocation_instance = None
        
        # Extracting the source of fund from the encoded string
        sof_encoded = request.POST.get('source_of_fund')
        if sof_encoded:
            try:
                parts = sof_encoded.split('|', 2)
                if len(parts) >= 3:
                    pre_id_str, item_key, quarters_data = parts
                    # DepartmentPRE object instance
                    source_pre_instance = DepartmentPRE.objects.get(id=int(pre_id_str))
                    
                    # Validate budget availability
                    all_items = PRELineItemBudget.objects.filter(
                        pre=source_pre_instance,
                        item_key=item_key
                    )
                    
                    available_items = [item for item in all_items if item.remaining_amount > 0]
                    
                    if not available_items:
                        messages.error(request, "No budget available for the selected source of fund. All quarters consumed.")
                        return redirect('activity_design_form')
                    
                    total_available = sum(item.remaining_amount for item in available_items)
                    
                    if total_available < total_amount:
                        messages.error(f"Insufficient budget. Available: ₱{total_available:,.2f}, Required: ₱{total_amount:,.2f}")
                        return redirect('activity_design_form')
                    
                    remaining_to_allocate = total_amount
                    
                    for item in available_items:
                        if remaining_to_allocate <= 0:
                            break
                        
                        allocation_amount = min(item.remaining_amount, remaining_to_allocate)
                        if allocation_amount > 0:
                            allocations_to_make.append({
                                'item': item,
                                'amount': allocation_amount
                            })
                            remaining_to_allocate -= allocation_amount
                            
                    # Prepare source information
                    quarters = quarters_data.split('|')
                    if quarters:
                        first_quarter_data = quarters[0].split(':')
                        if len(first_quarter_data) == 2:
                            quarter, amount_str = first_quarter_data
                            source_amount_decimal = Decimal(amount_str)
                            
                    # Calculate display value
                    total_amount_display = sum(Decimal(q.split(':')[1]) for q in quarters if ':' in q)
                    item_label = FRIENDLY_LABELS.get(item_key, item_key.replace('_', ' ').title())
                    source_of_fund_display = f"{item_label} - ₱{total_amount_display:,.2f}"

            except (Exception, ValueError, DepartmentPRE.DoesNotExist):
                source_pre_instance = None
                messages.error(request, f"Invalid source of fund format or {Exception.message}.")
                return redirect('activity_design_form')
                
        
        activity = ActivityDesign.objects.create(
            title_of_activity=request.POST.get('title_of_activity'),
            schedule_date=request.POST.get('schedule_date'),
            venue=request.POST.get('venue'),
            rationale=request.POST.get('rationale'),
            objectives=request.POST.get('objectives'),
            methodology=request.POST.get('methodology'),
            participants=request.POST.get('participants'),
            resource_persons=request.POST.get('resource_persons'),
            materials_needed=request.POST.get('materials_needed'),
            budget_allocation=budget_allocation_instance,
            evaluation_plan=request.POST.get('evaluation_outcomes'),
            source_pre=source_pre_instance,
            source_item_key=item_key,
            source_quarter=quarter,
            source_amount=source_amount_decimal,
            total_amount=total_amount,
            source_of_fund_display=source_of_fund_display,
            requested_by=request.user,
        )
        
        # Execute budget allocations (now we have the activity object)
        for allocation_data in allocations_to_make:
            item = allocation_data['item']
            amount = allocation_data['amount']
            
            # Update consumed amount
            item.consumed_amount += amount
            item.save()
            
            # CReate tracking record
            ActivityDesignAllocations.objects.create(
                activity_design=activity,
                pre_line_item=item,
                allocated_amount=amount
            )
        
        # Save sessions (many-to-one)
        session_contents = request.POST.getlist('sessions[]')
        for order, content in enumerate(session_contents, start=1):
            if content.strip():
                Session.objects.create(
                    activity=activity,
                    content=content,
                    order=order
                )
             
        # Temporary commented this code for form enhancing   
        # Save signatories (many-to-many)
        # names = request.POST.getlist('signatory_name[]')
        # positions = request.POST.getlist('signatory_position[]')
        # dates = request.POST.getlist('signatory_date[]')
        
        # for name, position, date_str in zip(names, positions, dates):
        #     if name.strip():
        #         date_obj = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else None
        #         Signatory.objects.create(
        #             activity=activity,
        #             name=name,
        #             position=position,
        #             date=date_obj
        #         )
                
        # # Save campus approval (one-to-one)
        # campus_approval = CampusApproval.objects.create(
        #     activity=activity,
        #     name=request.POST.get('campus_name'),
        #     position=request.POST.get('campus_position'),
        #     date=datetime.strptime(request.POST.get('campus_date'), "%Y-%m-%d").date() if request.POST.get('campus_date') else None,
        #     remarks=request.POST.get('campus_remarks')
        # )
        
        # # Save University Approval (one-to-one)
        # university_approval = UniversityApproval.objects.create(
        #     activity=activity,
        #     name=request.POST.get('univ_name'),
        #     position=request.POST.get('univ_position'),
        #     date=datetime.strptime(request.POST.get('univ_date'), "%Y-%m-%d").date() if request.POST.get('univ_date') else None,
        #     remarks=request.POST.get('univ_remarks')
        # )
        
        
        # Log audit trail
        log_audit_trail(
            request=request,
            action='CREATE',
            model_name='ActivityDesign',
            record_id=activity.id,
            detail=f'Created activity design: {activity.title_of_activity}',
        )
        
        # Show message success 
        messages.success(request, "Activity Design submitted successfully.")
        
        return render(request, "end_user_app/activity_design_form.html", {'ad_id': activity.id, 'success': True})  # Redirect to the same form or a success page
        
    return render(request, "end_user_app/activity_design_form.html", {
        'budget_allocations': budget_allocations,
    })

@role_required('end_user', login_url='/')
def preview_activity_design(request, pk):
    activity = get_object_or_404(ActivityDesign.objects.prefetch_related('sessions', 'signatories').select_related('campus_approval', 'university_approval'), pk=pk)
    context = {
        "activity": activity,
    }
    return render(request, "end_user_app/preview_activity_design.html", context)

def build_pre_source_options(pres_queryset):
        options = []
        # approved_pres = DepartmentPRE.objects.filter(
        #     submitted_by=request,
        #     approved_by_approving_officer=True,
        #     approved_by_admin=True,
        # )

        # Friendly labels for known PRE item keys (fallback to humanized key)
        friendly_labels = FRIENDLY_LABELS
        
        line_item_groups = {}
        
        for pre in pres_queryset:
            # Get line item budgets instead of parsing JSON
            line_items = PRELineItemBudget.objects.filter(pre=pre)
            
            for line_item in line_items:
                if line_item.remaining_amount > 0:
                    key = f"{pre.id}|{line_item.item_key}"
                    
                    if key not in line_item_groups:
                        line_item_groups[key] = {
                            'pre_id': pre.id,
                            'item_key': line_item.item_key,
                            'label': friendly_labels.get(line_item.item_key, line_item.item_key.replace('_', ' ').title()),
                            'total_remaining': Decimal('0'),
                            'total_allocated': Decimal('0'),
                            'quarters': []
                        }
                        
                    line_item_groups[key]['total_remaining'] += line_item.remaining_amount
                    line_item_groups[key]['total_allocated'] += line_item.allocated_amount
                    line_item_groups[key]['quarters'].append({
                        'quarter': line_item.quarter,
                        'remaining': line_item.remaining_amount
                    })   
                                     
        # Create options from grouped data
        for group_key, group_data in line_item_groups.items():
            display = f"{group_data['label']} - ₱{intcomma(group_data['total_remaining'])}"
            
            # Encode all quarters for this line item
            quarters_data = "|".join([f"{q['quarter']}:{q['remaining']}" for q in group_data['quarters']])
            encoded = f"{group_data['pre_id']}|{group_data['item_key']}|{quarters_data}"
            
            options.append({'value': encoded, 'label': display})
        return options
    
@role_required('end_user', login_url='/')
def load_source_of_fund(request):
    ba_id = request.GET.get("budget_allocation")
    source_of_fund_options = []

    if ba_id:
        print("budget allocation ID:", ba_id)
        try:
            allocation = BudgetAllocation.objects.get(
                id=ba_id,
                department=getattr(request.user, 'department', '')
            )

            # ✅ filter PREs linked to that allocation
            approved_pres = DepartmentPRE.objects.filter(
                budget_allocation=allocation,
                submitted_by=request.user.id,
                approved_by_approving_officer=True,
                approved_by_admin=True,
            )

            # ✅ reuse your builder
            source_of_fund_options = build_pre_source_options(approved_pres)

        except BudgetAllocation.DoesNotExist:
            print("No such budget allocation for this user.")
            source_of_fund_options = []
        
    else:
        print("No budget allocation ID provided.")

    return render(
        request,
        "end_user_app/partials/source_of_fund_options.html",
        {"source_of_fund_options": source_of_fund_options}
    )

@role_required('end_user', login_url='/')
def budget_details(request, budget_id):
    budget_allocation = get_object_or_404(BudgetAllocation.objects.select_related('approved_budget'), id=budget_id, department=getattr(request.user, 'department', ''))
    
    try:
        # Get all PRE line items for this budget allocation
        related_pres = DepartmentPRE.objects.filter(
            budget_allocation=budget_allocation,
            submitted_by=request.user
        ).order_by('-created_at')
    except DepartmentPRE.DoesNotExist:
        messages.error(request, "No PREs found for this budget allocation.")
        return redirect('user_view_budget')
    
    # Get all PRE line items with friendly names
    pre_line_items = []
    total_pre_allocated = Decimal('0')
    total_pre_consumed = Decimal('0')
    
    # Friendly labels (you can import this from constants.py if you created it)
    friendly_labels = FRIENDLY_LABELS
    
    for pre in related_pres:
        line_items = PRELineItemBudget.objects.filter(pre=pre).order_by('item_key', 'quarter')
        
        for item in line_items:
            # Add friendly name to the item
            item.friendly_name = friendly_labels.get(
                item.item_key, 
                item.item_key.replace('_', ' ').title()
            )
            
            pre_line_items.append(item)
            total_pre_allocated += item.allocated_amount
            total_pre_consumed += item.consumed_amount
    
    total_pre_remaining = total_pre_allocated - total_pre_consumed
    
    # Calculate utilization percentage
    utilization_percentage = 0
    if budget_allocation.total_allocated > 0:
        utilization_percentage = (budget_allocation.spent / budget_allocation.total_allocated) * 100
    
    context = {
        'budget_allocation': budget_allocation,
        'pre_line_items': pre_line_items,
        'related_pres': related_pres,
        'total_pre_allocated': total_pre_allocated,
        'total_pre_consumed': total_pre_consumed,
        'total_pre_remaining': total_pre_remaining,
        'utilization_percentage': utilization_percentage,
    }
    
    return render(request, 'end_user_app/budget_details.html', context)
     
@role_required('end_user', login_url='/')
def pre_budget_realignment(request):
    """PRE-based budget realignment form"""
    
    # Get user's approved PREs
    approved_pres = DepartmentPRE.objects.filter(
        submitted_by=request.user,
        approved_by_approving_officer=True,
        approved_by_admin=True,
    ).order_by('-created_at')
    
    # Build available budget categories
    available_categories = []
    
    for pre in approved_pres:
        line_items = PRELineItemBudget.objects.filter(pre=pre)
        
        # Group by item_key and calculate totals
        category_totals = {}
        for item in line_items:
            if item.item_key not in category_totals:
                category_totals[item.item_key] = {
                    'pre_id': pre.id,
                    'item_key': item.item_key,
                    'label': FRIENDLY_LABELS.get(item.item_key, item.item_key.replace('_', ' ').title()),
                    'total_allocated': Decimal('0'),
                    'total_consumed': Decimal('0'),
                    'total_remaining': Decimal('0'),
                    'quarters': []
                }
            
            category_totals[item.item_key]['total_allocated'] += item.allocated_amount
            category_totals[item.item_key]['total_consumed'] += item.consumed_amount
            category_totals[item.item_key]['total_remaining'] += item.remaining_amount
            category_totals[item.item_key]['quarters'].append({
                'quarter': item.quarter,
                'allocated': item.allocated_amount,
                'consumed': item.consumed_amount,
                'remaining': item.remaining_amount
            })
            
        # Add categories with remaining budget
        for item_key, data in category_totals.items():
            if data['total_remaining'] > 0:
                available_categories.append({
                    'value': f"{data['pre_id']}|{item_key}",
                    'label': f"{data['label']} - ₱{data['total_remaining']:,.2f} available",
                    'remaining': data['total_remaining'],
                    'pre_id': data['pre_id'],
                    'item_key': item_key
                })
                
    if request.method == 'POST':
        source_encoded = request.POST.get('source_category')
        target_encoded = request.POST.get('target_category')
        amount = Decimal(request.POST.get('amount', '0'))
        reason = request.POST.get('reason', '').strip()
        
        # Validation
        if not source_encoded or not target_encoded:
            messages.error(request, "Please select both source and target categories.")
            return redirect('pre_budget_realignment')
        
        if source_encoded == target_encoded:
            messages.error(request, "Source and target categories cannot be the same.")
            return redirect('pre_budget_realignment')
        
        if amount <= 0:
            messages.error(request, "Amount must be greater than zero.")
            return redirect('pre_budget_realignment')
        
        if not reason:
            messages.error(request, "Please provide a reason for the realignment.")
            return redirect('pre_budget_realignment')
        
        try:
            # Parse source and target
            source_pre_id, source_item_key = source_encoded.split('|')
            target_pre_id, target_item_key = target_encoded.split('|')
            
            source_pre = DepartmentPRE.objects.get(id=source_pre_id, submitted_by=request.user)
            target_pre = DepartmentPRE.objects.get(id=target_pre_id, submitted_by=request.user)
            
            # Validate source has sufficient funds
            source_items = PRELineItemBudget.objects.filter(
                pre=source_pre,
                item_key=source_item_key
            )
            
            total_available = sum(item.remaining_amount for item in source_items)
            
            if total_available < amount:
                messages.error(request, f"Insufficient funds. Available: ₱{total_available:,.2f}, Requested: ₱{amount:,.2f}")
                return redirect('pre_budget_realignment')
            
            # Validate target category exists
            target_items = PRELineItemBudget.objects.filter(
                pre=target_pre,
                item_key=target_item_key
            )
            
            if not target_items.exists():
                messages.error(request, "Target category not found in your PRE.")
                return redirect('pre_budget_realignment')
            
            # Create realignment request
            realignment = PREBudgetRealignment.objects.create(
                requested_by=request.user,
                source_pre=source_pre,
                source_item_key=source_item_key,
                target_pre=target_pre,
                target_item_key=target_item_key,
                amount=amount,
                reason=reason,
                source_item_display=FRIENDLY_LABELS.get(source_item_key, source_item_key.replace('_', ' ').title()),
                target_item_display=FRIENDLY_LABELS.get(target_item_key, target_item_key.replace('_', ' ').title()),
            )
            
            # Log audit trail
            log_audit_trail(
                request=request,
                action='CREATE',
                model_name='PREBudgetRealignment',
                record_id=realignment.id,
                detail=f'Requested budget realignment: {realignment.source_item_display} → {realignment.target_item_display} (₱{amount:,.2f})',
            )
            
            messages.success(request, f"Budget realignment request submitted successfully. Request ID: {realignment.id}")
            return redirect('pre_budget_realignment')
            
        except (ValueError, DepartmentPRE.DoesNotExist) as e:
            messages.error(request, f"Invalid selection: {str(e)}")
            return redirect('pre_budget_realignment')
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('pre_budget_realignment')
    
    # GET request - render form
    context = {
        'available_categories': available_categories,
        'approved_pres': approved_pres,
    }
    
    return render(request, "end_user_app/pre_budget_realignment.html", context)

@role_required('end_user', login_url='/')
def realignment_history(request):
    """View for displaying user's budget realignment history"""
    
    # Get all realignment requests for the current user
    realignment_requests = PREBudgetRealignment.objects.filter(
        requested_by=request.user
    ).select_related('source_pre', 'target_pre', 'approved_by').order_by('-created_at')
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(realignment_requests, 10)  # Show 10 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        realignment_requests = realignment_requests.filter(status=status_filter)
        paginator = Paginator(realignment_requests, 10)
        page_obj = paginator.get_page(page_number)
    
    # Calculate summary statistics
    total_requests = realignment_requests.count()
    pending_count = realignment_requests.filter(status='Pending').count()
    approved_count = realignment_requests.filter(status='Approved').count()
    rejected_count = realignment_requests.filter(status='Rejected').count()
    
    context = {
        'page_obj': page_obj,
        'realignment_requests': page_obj,
        'status_filter': status_filter,
        'status_choices': PREBudgetRealignment.STATUS_CHOICES,
        'total_requests': total_requests,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
    }
    
    return render(request, "end_user_app/realignment_history.html", context)

@role_required('end_user', login_url='/')
def download_activity_design_word(request, pk):
    """Generate and download Activity Design as Word document"""
    
    # Get the activity design
    activity = get_object_or_404(
        ActivityDesign.objects.prefetch_related('sessions').select_related('requested_by', 'source_pre'),
        pk=pk,
        requested_by=request.user  # Ensure user can only download their own
    )
    
    try:
        # Path to your Word template
        template_path = r"C:\Users\John Romel Lucot\OneDrive\Desktop\Capstone project\bb_budget_monitoring_system\document_templates\activity_design_template.docx"
        
        # Load the template
        doc = DocxTemplate(template_path)
        
        # Prepare context data
        context = {
            'title_of_activity': activity.title_of_activity or '',
            'schedule_date': activity.schedule_date.strftime('%B %d, %Y') if activity.schedule_date else '',
            'venue': activity.venue or '',
            'rationale': activity.rationale or '',
            'objectives': activity.objectives or '',
            'methodology': activity.methodology or '',
            'participants': activity.participants or '',
            'resource_persons': activity.resource_persons or '',
            'materials_needed': activity.materials_needed or '',
            'evaluation_plan': activity.evaluation_plan or '',
            'source_of_fund_display': activity.source_of_fund_display or 'Not specified',
            'total_amount': f"{activity.total_amount:,.2f}" if activity.total_amount else '0.00',
            'requested_by_name': activity.requested_by.get_full_name() if activity.requested_by else 'Unknown',
            'date_prepared': datetime.now().strftime('%B %d, %Y'),
            'sessions': [
                {
                    'order': session.order,
                    'content': session.content
                }
                for session in activity.sessions.all().order_by('order')
            ]
        }
        
        # Render the template with context
        doc.render(context, autoescape=True)
        
        # Generate filename
        safe_title = "".join(c for c in activity.title_of_activity if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_title = safe_title.replace(' ', '_')[:50]  # Limit length
        filename = f"Activity_Design_{safe_title}_{activity.id}.docx"
        
        # Create HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
        # Save document to response
        doc.save(response)
        
        # Log the download
        log_audit_trail(
            request=request,
            action='DOWNLOAD',
            model_name='ActivityDesign',
            record_id=activity.id,
            detail=f'Downloaded Word document for activity design: {activity.title_of_activity}',
        )
        
        return response
        
    except FileNotFoundError:
        messages.error(request, "Word template not found. Please contact administrator.")
        return redirect('preview_activity_design', pk=pk)
    except Exception as e:
        messages.error(request, f"Error generating Word document: {str(e)}")
        return redirect('preview_activity_design', pk=pk)
    
# This views is for inspecting the excel template and its structure
def inspect_excel_template(request):
    """Inspect the Excel template structure"""
    
    try:
        template_path = r"c:\Users\John Romel Lucot\OneDrive\Desktop\Capstone project\bb_budget_monitoring_system\excel_templates\Departmental-PRE.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active
        
        inspection_data = []
        inspection_data.append(f"Sheet name: {ws.title}")
        inspection_data.append(f"Max row: {ws.max_row}")
        inspection_data.append(f"Max column: {ws.max_column}")
        inspection_data.append("")
        
        # Inspect first 30 rows and 10 columns
        for row in range(1, 31):
            row_data = []
            for col in range(1, 11):  # A to J columns
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    col_letter = chr(64 + col)  # Convert to letter (A, B, C...)
                    row_data.append(f"{col_letter}{row}: '{cell.value}'")
            
            if row_data:
                inspection_data.append(f"Row {row}: {' | '.join(row_data)}")
        
        # Return as plain text response for easy reading
        return HttpResponse(
            "\n".join(inspection_data),
            content_type="text/plain"
        )
        
    except Exception as e:
        return HttpResponse(f"Error inspecting template: {e}", content_type="text/plain")
    

# This views is for downloading the PRE as an Excel document but uses openpyxl
@role_required('end_user', login_url='/')
def download_pre_excel(request, pk):
    """Generate and download PRE as Excel - CORRECT SHEET VERSION"""
    
    pre = get_object_or_404(
        DepartmentPRE.objects.select_related('submitted_by'),
        pk=pk,
        submitted_by=request.user
    )
    
    try:
        template_path = r"c:\Users\John Romel Lucot\OneDrive\Desktop\Capstone project\bb_budget_monitoring_system\excel_templates\Departmental-PRE.xlsx"
        
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Excel template not found")
        
        # Create temporary copy
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            shutil.copy2(template_path, temp_file.name)
            temp_path = temp_file.name
        
        def to_decimal(value):
            try:
                return Decimal(str(value)) if value not in (None, "") else Decimal('0')
            except Exception:
                return Decimal('0')
        
        payload = pre.data or {}
        
        # ✅ Open Excel with xlwings
        app = xw.App(visible=False)
        wb = app.books.open(temp_path)
        
        # ✅ CRITICAL FIX: Select the correct sheet
        print(f"=== SHEET SELECTION ===")
        print(f"Available sheets: {[sheet.name for sheet in wb.sheets]}")
        
        try:
            ws = wb.sheets["Departmental PRE (2)"]  # ✅ Correct sheet
            print(f"Selected sheet: {ws.name}")
        except Exception as e:
            print(f"Could not find 'Departmental PRE (2)' sheet: {e}")
            # Fallback to second sheet by index
            ws = wb.sheets[1] if len(wb.sheets) > 1 else wb.sheets[0]
            print(f"Using fallback sheet: {ws.name}")
        
        # ✅ Update department name
        print(f"=== UPDATING DEPARTMENT ===")
        try:
            old_dept = ws.range('C5').value
            ws.range('C5').value = pre.department
            new_dept = ws.range('C5').value
            print(f"Department: '{old_dept}' -> '{new_dept}'")
        except Exception as e:
            print(f"Department update error: {e}")
        
        # ✅ Budget data mapping
        budget_items_mapping = {
            # Personnel Services Section
            'basic_salary': 15,
            'honoraria': 16,
            'overtime_pay': 17,
            
            # MOOE Section
            'travel_local': 23,
            'travel_foreign': 24,
            'training_expenses': 26,
            'office_supplies_expenses': 28,
            'accountable_form_expenses': 29,
            'agri_marine_supplies_expenses': 30,
            'drugs_medicines': 31,
            'med_dental_lab_supplies_expenses': 32,
            'food_supplies_expenses': 33,
            'fuel_oil_lubricants_expenses': 39,
            'textbooks_instructional_materials_expenses': 40,
            'construction_material_expenses': 41,
            'other_supplies_materials_expenses': 42,
            'semee_machinery': 44,
            'semee_office_equipment': 45,
            'semee_information_communication': 46,
            'semee_communications_equipment': 47,
            'semee_drr_equipment': 48,
            'semee_medical_equipment': 49,
            'semee_printing_equipment': 50,
            'semee_sports_equipment': 51,
            'semee_technical_scientific_equipment': 52,
            'semee_ict_equipment': 53,
            'semee_other_machinery_equipment': 54,
            'furniture_fixtures': 56,
            'books': 57,
            'water_expenses': 59,
            'electricity_expenses': 60,
            'postage_courier_services': 62,
            'telephone_expenses': 63,
            'telephone_expenses_landline': 64,
            'internet_subscription_expenses': 65,
            'cable_satellite_telegraph_radio_expenses': 66,
            'awards_rewards_expenses': 72,
            'prizes': 73,
            'survey_expenses': 75,
            'survey_research_exploration_development_expenses': 76,
            'legal_services': 78,
            'auditing_services': 79,
            'consultancy_services': 80,
            'other_professional_servies': 81,
            'security_services': 83,
            'janitorial_services': 84,
            'other_general_services': 85,
            'environment/sanitary_services': 86,
            'repair_maintenance_land_improvements': 88,
            'buildings': 90,
            'school_buildings': 91,
            'hostel_dormitories': 92,
            'other_structures': 93,
            'repair_maintenance_machinery': 95,
            'repair_maintenance_office_equipment': 96,
            'repair_maintenance_ict_equipment': 97,
            'repair_maintenance_agri_forestry_equipment': 98,
            'repair_maintenance_marine_fishery_equipment': 99,
            'repair_maintenance_airport_equipment': 100,
            'repair_maintenance_communication_equipment': 101,
            'repair_maintenance_drre_equipment': 102,
            'repair_maintenance_medical_equipment': 103,
            'repair_maintenance_printing_equipment': 104,
            'repair_maintenance_sports_equipment': 105,
            'repair_maintenance_technical_scientific_equipment': 106,
            'repair_maintenance_other_machinery_equipment': 107,
            'repair_maintenance_motor': 109,
            'repair_maintenance_other_transportation_equipment': 110,
            'repair_maintenance_furniture_fixtures': 111,
            'repair_maintenance_semi_expendable_machinery_equipment': 112,
            'repair_maintenance_other_property_plant_equipment': 113,
            'taxes_duties_licenses': 115,
            'fidelity_bond_premiums': 116,
            'insurance_expenses': 117,
            'labor_wages': 119,
            'advertising_expenses': 121,
            'printing_publication_expenses': 122,
            'representation_expenses': 123,
            'transportation_delivery_expenses': 124,
            'rent/lease_expenses': 125,
            'membership_dues_contribute_to_org': 126,
            'subscription_expenses': 127,
            'website_maintenance': 129,
            'other_maintenance_operating_expenses': 130,
            
            # Capital Outlays Section
            'land': 136,
            'land_improvements_aqua_structure': 138,
            'water_supply_systems': 140,
            'power_supply_systems': 141,
            'other_infra_assets': 142,
            'bos_building': 144,
            'bos_school_buildings': 145,
            'bos_hostels_dorm': 146,
            'other_structures': 147,
            'me_machinery': 149,
            'me_office_equipment': 150,
            'me_ict_equipment': 151,
            'me_communication_equipment': 152,
            'me_drre': 153,
            'me_medical_equipment': 154,
            'me_printing_equipment': 155,
            'me_sports_equipment': 156,
            'me_technical_scientific_equipment': 157,
            'me_other_machinery_equipment': 158,
            'te_motor': 160,
            'te_other_transpo_equipment': 161,
            'ffb_furniture_fixtures': 163,
            'ffb_books': 164,
            'cp_land_improvements': 166,
            'cp_infra_assets': 167,
            'cp_building_other_structures': 168,
            'cp_leased_assets': 169,
            'cp_leased_assets_improvements': 170,
            'ia_computer_software': 172,
            'ia_websites': 173,
            'ia_other_tangible_assets': 174,
        }
        
        # ✅ Fill budget data
        print(f"=== UPDATING BUDGET DATA ===")
        updates_made = 0
        
        for item_key, row_num in budget_items_mapping.items():
            try:
                # Get quarterly values
                q1 = to_decimal(payload.get(f"{item_key}_q1"))
                q2 = to_decimal(payload.get(f"{item_key}_q2"))
                q3 = to_decimal(payload.get(f"{item_key}_q3"))
                q4 = to_decimal(payload.get(f"{item_key}_q4"))
                total = q1 + q2 + q3 + q4
                
                if total > 0:
                    print(f"Updating {item_key} (Row {row_num}): Q1={q1}, Q2={q2}, Q3={q3}, Q4={q4}")
                    
                    # Update quarterly data
                    if q1 > 0:
                        ws.range(f'E{row_num}').value = float(q1)
                    if q2 > 0:
                        ws.range(f'F{row_num}').value = float(q2)
                    if q3 > 0:
                        ws.range(f'G{row_num}').value = float(q3)
                    if q4 > 0:
                        ws.range(f'H{row_num}').value = float(q4)
                    if total > 0:
                        ws.range(f'I{row_num}').value = float(total)
                    
                    updates_made += 1
                    
            except Exception as e:
                print(f"Error updating {item_key}: {e}")
                continue
        
        print(f"Total updates made: {updates_made}")
        
        # ✅ Fill certification section
        try:
            # Search for certification fields
            for row in range(130, 200):
                try:
                    cell_value = ws.range(f'A{row}').value
                    if cell_value:
                        cell_value_lower = str(cell_value).lower()
                        
                        if 'prepared' in cell_value_lower and 'by' in cell_value_lower:
                            if pre.prepared_by_name:
                                ws.range(f'C{row}').value = pre.prepared_by_name
                                print(f"Updated prepared by at row {row}")
                        elif 'certified' in cell_value_lower and 'by' in cell_value_lower:
                            if pre.certified_by_name:
                                ws.range(f'C{row}').value = pre.certified_by_name
                                print(f"Updated certified by at row {row}")
                        elif 'approved' in cell_value_lower and 'by' in cell_value_lower:
                            if pre.approved_by_name:
                                ws.range(f'C{row}').value = pre.approved_by_name
                                print(f"Updated approved by at row {row}")
                except Exception:
                    continue
        except Exception as e:
            print(f"Could not update certification: {e}")
        
        # ✅ Save and close
        print("Saving workbook...")
        wb.save()
        wb.close()
        app.quit()
        
        # Read the file
        with open(temp_path, 'rb') as f:
            file_content = f.read()
        
        # Clean up
        try:
            os.unlink(temp_path)
        except:
            pass
        
        # Generate filename
        safe_dept = "".join(c for c in pre.department if c.isalnum() or c in (' ', '-', '_')).replace(' ', '_')
        filename = f"PRE_{safe_dept}_{pre.id}_{pre.created_at.strftime('%Y%m%d')}.xlsx"
        
        # Create response
        response = HttpResponse(
            file_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
        # Log the download
        log_audit_trail(
            request=request,
            action='DOWNLOAD',
            model_name='DepartmentPRE',
            record_id=pre.id,
            detail=f'Downloaded Excel document for PRE: {pre.department}',
        )
        
        return response
        
    except Exception as e:
        messages.error(request, f"Error generating Excel document: {str(e)}")
        print(f"DEBUG: Excel generation error: {e}")
        import traceback
        traceback.print_exc()
        return redirect('preview_pre', pk=pk)


# This views is for downloading the PRE into excel format but this one uses xlwings
# @role_required('end_user', login_url='/')
# def download_pre_excel(request, pk):
#     """Debug version - Generate and download PRE as Excel with detailed logging"""
    
#     pre = get_object_or_404(
#         DepartmentPRE.objects.select_related('submitted_by'),
#         pk=pk,
#         submitted_by=request.user
#     )
    
#     try:
#         template_path = r"c:\Users\John Romel Lucot\OneDrive\Desktop\Capstone project\bb_budget_monitoring_system\excel_templates\Departmental-PRE.xlsx"
        
#         if not os.path.exists(template_path):
#             raise FileNotFoundError(f"Excel template not found at {template_path}")
        
#         # Create temporary copy
#         with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
#             shutil.copy2(template_path, temp_file.name)
#             temp_path = temp_file.name
        
#         def to_decimal(value):
#             try:
#                 return Decimal(str(value)) if value not in (None, "") else Decimal('0')
#             except Exception:
#                 return Decimal('0')
        
#         payload = pre.data or {}
        
#         # ✅ DEBUG: Print PRE data to understand what we have
#         print(f"=== PRE DATA DEBUG ===")
#         print(f"PRE ID: {pre.id}")
#         print(f"Department: {pre.department}")
#         print(f"Payload keys: {list(payload.keys()) if payload else 'No payload'}")
        
#         # Print some sample data
#         if payload:
#             for key, value in list(payload.items())[:10]:  # First 10 items
#                 print(f"  {key}: {value}")
        
#         # Open with xlwings
#         app = xw.App(visible=False)
#         wb = app.books.open(temp_path)
#         ws = wb.sheets[0]
        
#         # ✅ DEBUG: Test basic cell update first
#         print(f"=== TESTING BASIC UPDATES ===")
#         try:
#             # Test department update
#             old_dept = ws.range('C5').value
#             ws.range('C5').value = pre.department
#             new_dept = ws.range('C5').value
#             print(f"Department update: '{old_dept}' -> '{new_dept}'")
#         except Exception as e:
#             print(f"Error updating department: {e}")
        
#         # ✅ DEBUG: Test a few budget items with detailed logging
#         test_items = {
#             'basic_salary': 15,
#             'travel_local': 23,
#             'office_supplies_expenses': 28,
#             'food_supplies_expenses': 33,
#         }
        
#         print(f"=== TESTING BUDGET DATA UPDATES ===")
#         for item_key, row_num in test_items.items():
#             try:
#                 # Get quarterly values
#                 q1 = to_decimal(payload.get(f"{item_key}_q1"))
#                 q2 = to_decimal(payload.get(f"{item_key}_q2"))
#                 q3 = to_decimal(payload.get(f"{item_key}_q3"))
#                 q4 = to_decimal(payload.get(f"{item_key}_q4"))
#                 total = q1 + q2 + q3 + q4
                
#                 print(f"Item: {item_key} (Row {row_num})")
#                 print(f"  Q1: {q1}, Q2: {q2}, Q3: {q3}, Q4: {q4}, Total: {total}")
                
#                 # Check current cell values
#                 current_e = ws.range(f'E{row_num}').value
#                 current_f = ws.range(f'F{row_num}').value
#                 current_g = ws.range(f'G{row_num}').value
#                 current_h = ws.range(f'H{row_num}').value
#                 current_i = ws.range(f'I{row_num}').value
                
#                 print(f"  Current cells: E={current_e}, F={current_f}, G={current_g}, H={current_h}, I={current_i}")
                
#                 # Update cells if we have data
#                 if q1 > 0:
#                     ws.range(f'E{row_num}').value = float(q1)
#                     print(f"  Updated E{row_num} = {float(q1)}")
#                 if q2 > 0:
#                     ws.range(f'F{row_num}').value = float(q2)
#                     print(f"  Updated F{row_num} = {float(q2)}")
#                 if q3 > 0:
#                     ws.range(f'G{row_num}').value = float(q3)
#                     print(f"  Updated G{row_num} = {float(q3)}")
#                 if q4 > 0:
#                     ws.range(f'H{row_num}').value = float(q4)
#                     print(f"  Updated H{row_num} = {float(q4)}")
#                 if total > 0:
#                     ws.range(f'I{row_num}').value = float(total)
#                     print(f"  Updated I{row_num} = {float(total)}")
                
#                 # Verify the updates
#                 new_e = ws.range(f'E{row_num}').value
#                 new_f = ws.range(f'F{row_num}').value
#                 new_g = ws.range(f'G{row_num}').value
#                 new_h = ws.range(f'H{row_num}').value
#                 new_i = ws.range(f'I{row_num}').value
                
#                 print(f"  After update: E={new_e}, F={new_f}, G={new_g}, H={new_h}, I={new_i}")
#                 print("  ---")
                
#             except Exception as e:
#                 print(f"Error updating {item_key}: {e}")
#                 continue
        
#         # ✅ Save and close
#         print(f"=== SAVING FILE ===")
#         wb.save()
#         wb.close()
#         app.quit()
        
#         # Read the saved file
#         with open(temp_path, 'rb') as f:
#             file_content = f.read()
        
#         # Clean up
#         try:
#             os.unlink(temp_path)
#         except:
#             pass
        
#         # Generate filename
#         safe_dept = "".join(c for c in pre.department if c.isalnum() or c in (' ', '-', '_')).replace(' ', '_')
#         filename = f"PRE_{safe_dept}_{pre.id}_{pre.created_at.strftime('%Y%m%d')}.xlsx"
        
#         # Create response
#         response = HttpResponse(
#             file_content,
#             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#         response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
#         # Log the download
#         log_audit_trail(
#             request=request,
#             action='DOWNLOAD',
#             model_name='DepartmentPRE',
#             record_id=pre.id,
#             detail=f'Downloaded Excel document for PRE: {pre.department}',
#         )
        
#         return response
        
#     except Exception as e:
#         messages.error(request, f"Error generating Excel document: {str(e)}")
#         print(f"DEBUG: xlwings error: {e}")
#         import traceback
#         traceback.print_exc()
#         return redirect('preview_pre', pk=pk)