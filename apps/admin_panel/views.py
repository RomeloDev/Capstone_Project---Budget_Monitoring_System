# bb_budget_monitoring_system/apps/admin_panel/views.py
from datetime import datetime, timedelta
import decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from apps.users.models import User
from .models import BudgetAllocation, Budget, ApprovedBudget, AuditTrail
from apps.budgets.models import ApprovedBudget as NewApprovedBudget, SupportingDocument, BudgetAllocation as NewBudgetAllocation, DepartmentPRE as NewDepartmentPRE, RequestApproval, SystemNotification
from apps.users.models import User
from django.contrib import messages
from decimal import Decimal
from apps.end_user_app.models import PurchaseRequest, Budget_Realignment, DepartmentPRE, ActivityDesign, PRELineItemBudget, PurchaseRequestAllocation, ActivityDesignAllocations, PREBudgetRealignment
from apps.users.models import User
from django.db import transaction
from django.db.models import Sum
from django.template.defaultfilters import floatformat
from django.contrib.humanize.templatetags.humanize import intcomma
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .utils import log_audit_trail, generate_pre_number
from django.db.models import Count
from datetime import timedelta
from apps.users.utils import role_required
import mimetypes
from django.http import FileResponse, Http404, JsonResponse
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from .forms import ApprovedDocumentUploadForm

# Create your views here.
@role_required('admin', login_url='/admin/')
def admin_dashboard(request):
    try:
        # Total Users (Active users who logged in within last 30 days)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        end_users_total = User.objects.filter(
            is_staff=False, 
            is_approving_officer=True,
            last_login__gte=thirty_days_ago
        ).count()
        
        # Alternative: All registered users if you prefer total count
        # end_users_total = User.objects.filter(is_staff=False, is_approving_officer=True).count()
        
        # Total Budget from ApprovedBudget (since Budget model seems different)
        total_budget = ApprovedBudget.objects.aggregate(
            Sum('amount')
        )['amount__sum'] or 0
        
        # Pending Requests Count
        total_pending_realignment_request = Budget_Realignment.objects.filter(
            status='pending'
        ).count()
        
        # Approved Requests Count  
        total_approved_realignment_request = Budget_Realignment.objects.filter(
            status='approved'
        ).count()
        
        # Budget allocations for the table
        budget_allocated = BudgetAllocation.objects.select_related('approved_budget').all()
        
        # Calculate percentage changes for trends (you can customize this based on your needs)
        # This is a simple example - you might want to compare with previous period
        user_trend = "up"  # You can calculate actual trend
        budget_trend = "up"
        pending_trend = "down" if total_pending_realignment_request < 10 else "up"
        approved_trend = "up"
        
        # Chart data (group by department)
        dept_agg = (
            BudgetAllocation.objects
            .values('department')
            .annotate(
                total_allocated=Sum('total_allocated'),
                total_spent=Sum('spent')
            )
            .order_by('department')
        )
        
        dept_labels = [row['department'] or 'Unknown' for row in dept_agg]
        dept_allocated = [float(row['total_allocated'] or 0) for row in dept_agg]
        dept_spent = [float(row['total_spent'] or 0) for row in dept_agg]
        dept_remaining = [max(0.0, a - s) for a, s in zip(dept_allocated, dept_spent)]
        
        # Recent Activity (latest 8 for better display)
        recent_activities = (
            AuditTrail.objects
            .select_related('user')
            .order_by('-timestamp')[:8]
        )
        
        # Additional metrics for enhanced dashboard
        # Total departments with active budgets
        active_departments = BudgetAllocation.objects.values('department').distinct().count()
        
        # Average budget utilization
        total_allocated_sum = sum(dept_allocated) or 1  # Avoid division by zero
        total_spent_sum = sum(dept_spent)
        avg_utilization = (total_spent_sum / total_allocated_sum * 100) if total_allocated_sum > 0 else 0
        
        # Low budget departments (less than 20% remaining)
        low_budget_depts = BudgetAllocation.objects.filter(
            total_allocated__gt=0
        ).extra(
            where=["(total_allocated - spent) / total_allocated < 0.2"]
        ).count()
        
    except Exception as e:
        # Fallback values in case of any errors
        print(f"Dashboard error: {e}")  # For debugging
        end_users_total = 0
        total_budget = 0
        total_pending_realignment_request = 0
        total_approved_realignment_request = 0
        budget_allocated = BudgetAllocation.objects.none()
        dept_labels, dept_allocated, dept_spent, dept_remaining = [], [], [], []
        recent_activities = AuditTrail.objects.none()
        user_trend = budget_trend = pending_trend = approved_trend = "neutral"
        active_departments = 0
        avg_utilization = 0
        low_budget_depts = 0

    context = {
        # Main metrics
        'end_users_total': end_users_total,
        'total_budget': total_budget,
        'total_pending_realignment_request': total_pending_realignment_request,
        'total_approved_realignment_request': total_approved_realignment_request,
        
        # Trends (you can implement actual trend calculation)
        'user_trend': user_trend,
        'budget_trend': budget_trend,
        'pending_trend': pending_trend,
        'approved_trend': approved_trend,
        
        # Data for table and charts
        'budget_allocated': budget_allocated,
        'dept_labels': dept_labels,
        'dept_allocated': dept_allocated,
        'dept_spent': dept_spent,
        'dept_remaining': dept_remaining,
        
        # Recent activity
        'recent_activities': recent_activities,
        
        # Additional metrics
        'active_departments': active_departments,
        'avg_utilization': round(avg_utilization, 1),
        'low_budget_depts': low_budget_depts,
        
        # Template helper
        'current_time': timezone.now(),
    }
    
    return render(request, 'admin_panel/dashboard.html', context)
    
@role_required('admin', login_url='/admin/')
def client_accounts(request):
    try:
        end_users = User.objects.filter(is_staff=False)
    except User.DoesNotExist:
        end_users = None
    return render(request, 'admin_panel/client_accounts.html', {'end_users': end_users})

@role_required('admin', login_url='/admin/')
def register_account(request):
    
    if request.method == "POST":
         # Retrieval of value in input fields
        username = request.POST.get('username')
        fullname = request.POST.get('fullname')
        email = request.POST.get('email')
        mfo = request.POST.get('mfo')  # Main Department
        department = request.POST.get('department')  # Sub-department
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm-password')
        position = request.POST.get('position')
        
        # Validate password confirmation
        if password != confirm_password:
            return render(request, 'admin_panel/client_accounts.html', {'error': 'Passwords do not match'})
        
        # Check if the username or email already exists
        if User.objects.filter(username=username).exists():
            return render(request, 'admin_panel/client_accounts.html', {'error': 'Username already taken.'})
        if User.objects.filter(email=email).exists():
            return render(request, 'admin_panel/client_accounts.html', {'error': f'Email {email} already registered.'})
        
        # Check if the department already exists (specific sub-department)
        if User.objects.filter(department=department).exists():
            return render(request, 'admin_panel/client_accounts.html', {'error': f'Department {department} already registered.'})
        
        if department == 'Approving-Officer':
            approving_officer = User.objects.create_approving_officer(
                username=username, 
                fullname=fullname, 
                email=email, 
                password=password,
                mfo=mfo
            )
            approving_officer.position = position
            approving_officer.save()
            return render(request, "admin_panel/client_accounts.html", {'success': "Approving Officer Account registered successfully!"})
        else:
            # Create and save the user with both mfo and department
            user = User.objects.create_user(
                username=username, 
                fullname=fullname, 
                email=email, 
                password=password, 
                department=department,
                mfo=mfo,
                position=position
            )
            
            try:
                end_users = User.objects.filter(is_staff=False)
            except User.DoesNotExist:
                end_users = None
            return render(request, "admin_panel/client_accounts.html", {'success': "Account registered successfully!", 'end_users': end_users})

@role_required('admin', login_url='/admin/')
def departments_pr_request(request):
    """
    Admin view to manage Purchase Requests from departments
    Uses NEW models from budgets app
    """
    STATUS_CHOICES = (
        ('Pending', 'Pending'),
        ('Partially Approved', 'Partially Approved'),
        ('Rejected', 'Rejected'),
        ('Approved', 'Approved')
    )
    
    # Get distinct departments from users who submitted PRs
    departments = User.objects.filter(
        is_staff=False, 
        is_approving_officer=False,
        purchase_requests__isnull=False  # Has submitted PRs
    ).values_list('department', flat=True).distinct()
    
    # ✅ NEW: Use NewPurchaseRequest from budgets app
    from apps.budgets.models import PurchaseRequest as NewPurchaseRequest
    
    try:
        users_purchase_requests = NewPurchaseRequest.objects.select_related(
            'submitted_by',
            'budget_allocation__approved_budget',
            'source_pre',
            'source_line_item'
        ).prefetch_related(
            'supporting_documents',
            'pre_allocations'
        ).order_by('-created_at')
        
    except Exception as e:
        print(f"Error loading PRs: {e}")
        users_purchase_requests = NewPurchaseRequest.objects.none()
    
    # ✅ Calculate status counts
    status_counts = {
        'total': users_purchase_requests.count(),
        'pending': users_purchase_requests.filter(status='Pending').count(),
        'approved': users_purchase_requests.filter(status='Approved').count(),
        'partially_approved': users_purchase_requests.filter(status='Partially Approved').count(),
        'rejected': users_purchase_requests.filter(status='Rejected').count(),
    }
    
    # Apply filters
    filter_department = request.GET.get('department')
    if filter_department:
        users_purchase_requests = users_purchase_requests.filter(
            submitted_by__department=filter_department
        )
        
    status_filter = request.GET.get('status')
    if status_filter:
        users_purchase_requests = users_purchase_requests.filter(status=status_filter)
    
    context = {
        'users_purchase_requests': users_purchase_requests,
        'departments': departments,
        'status_choices': STATUS_CHOICES,
        'status_counts': status_counts
    }
    
    return render(request, 'admin_panel/departments_pr_request.html', context)

@role_required('admin', login_url='/admin/')
def handle_departments_request(request, request_id):
    """
    Admin handler for PRs with document conversion
    """
    from apps.budgets.models import PurchaseRequest as NewPurchaseRequest
    from apps.budgets.models import PurchaseRequestAllocation as NewPurchaseRequestAllocation
    from .document_converter import convert_pr_documents_to_pdf  # ← NEW IMPORT
    
    purchase_request = get_object_or_404(
        NewPurchaseRequest.objects.select_related(
            'submitted_by',
            'budget_allocation',
            'source_pre',
            'source_line_item'
        ).prefetch_related('pre_allocations', 'supporting_documents'),
        id=request_id
    )

    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Get allocation info for detailed feedback
        allocation = purchase_request.pre_allocations.first()
        line_item_name = allocation.pre_line_item.item_name if allocation else "Unknown"
        quarter = allocation.quarter if allocation else "Unknown"

        if action == 'approve':
            if purchase_request.status != 'Pending':
                messages.warning(request, 
                    f"PR {purchase_request.pr_number} cannot be approved. "
                    f"Current status: {purchase_request.status}")
                return redirect('department_pr_request')
            
            try:
                # 1. Update status to Partially Approved
                purchase_request.status = 'Partially Approved'
                purchase_request.partially_approved_at = timezone.now()
                purchase_request.save(update_fields=['status', 'partially_approved_at', 'updated_at'])
                
                # 2. 🔥 AUTO-CONVERT DOCUMENTS TO PDF
                print(f"🔄 Starting document conversion for PR {purchase_request.pr_number}")
                conversion_results = convert_pr_documents_to_pdf(purchase_request)

                # 3. Check conversion results and show appropriate messages
                if conversion_results['main_document']:
                    if conversion_results['main_document_format'] == 'PDF':
                        messages.success(request, 
                            f'✅ PR {purchase_request.pr_number} partially approved!')
                        messages.info(request,
                            f'📄 Main document converted to PDF successfully.')
                    else:
                        messages.success(request, 
                            f'✅ PR {purchase_request.pr_number} partially approved!')
                        messages.info(request,
                            f'📄 Main document already in PDF format.')
                else:
                    # Conversion completely failed
                    messages.success(request, 
                        f'✅ PR {purchase_request.pr_number} partially approved!')
                    messages.warning(request,
                        f'⚠️ Document conversion failed. Original files are available for download. Manual PDF conversion may be needed.')

                # Show specific warnings (only if there are issues)
                if conversion_results['warnings']:
                    for warning in conversion_results['warnings']:
                        messages.warning(request, warning)

                # Show errors (if any critical errors)
                if conversion_results['errors']:
                    for error in conversion_results['errors']:
                        messages.error(request, f"❌ {error}")
                
                # 4. Log audit trail
                log_audit_trail(
                    request=request,
                    action='APPROVE',
                    model_name='PurchaseRequest',
                    record_id=purchase_request.id,
                    detail=f'Purchase Request {purchase_request.pr_number} partially approved by Admin. '
                           f'Amount: ₱{purchase_request.total_amount:,.2f} from {line_item_name} - {quarter}. '
                           f'Documents converted: Main={conversion_results["main_document"]}, '
                           f'Supporting={len(conversion_results["supporting_docs"])}'
                )
                
                # 5. Create notification for end user
                from apps.budgets.models import SystemNotification
                SystemNotification.objects.create(
                    recipient=purchase_request.submitted_by,
                    title='PR Partially Approved - Ready to Print',
                    message=f'Your PR {purchase_request.pr_number} has been partially approved. '
                            f'Download the PDF documents, print them, and get them signed by the Approving Officer.',
                    content_type='pr',
                    object_id=purchase_request.id
                )
                
            except Exception as e:
                print(f"❌ Error in approval process: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error processing approval: {str(e)}")
                return redirect('department_pr_request')
            
        elif action == 'reject':
            if purchase_request.status != 'Pending':
                messages.warning(request, 
                    f"PR {purchase_request.pr_number} cannot be rejected. "
                    f"Current status: {purchase_request.status}")
                return redirect('department_pr_request')
            
            try:
                # Get allocations
                allocations = NewPurchaseRequestAllocation.objects.filter(
                    purchase_request=purchase_request
                ).select_related('pre_line_item')
                
                if not allocations.exists():
                    messages.warning(request, "No allocations found to reverse.")
                    return redirect('department_pr_request')
                
                total_released = Decimal('0')
                released_details = []
                
                # Reverse allocations
                for pr_allocation in allocations:
                    line_item = pr_allocation.pre_line_item
                    allocated_amount = pr_allocation.allocated_amount
                    quarter = pr_allocation.quarter
                    
                    # ✅ NO NEED to update line_item - consumption is calculated dynamically!
                    # Just track for logging
                    total_released += allocated_amount
                    released_details.append(
                        f"{line_item.item_name} {quarter}: ₱{allocated_amount:,.2f}"
                    )
                    
                    print(f"✅ Will release ₱{allocated_amount} from {line_item.item_name} {quarter}")

                # ✅ Delete allocations - this automatically "releases" the budget
                allocations.delete()

                print(f"✅ Deleted {len(released_details)} allocation(s), budget automatically released")
                
                # Update PR status
                purchase_request.status = 'Rejected'
                purchase_request.save(update_fields=['status', 'updated_at'])
                
                # Log audit trail
                log_audit_trail(
                    request=request,
                    action='REJECT',
                    model_name='PurchaseRequest',
                    record_id=purchase_request.id,
                    detail=f'Purchase Request {purchase_request.pr_number} rejected by Admin. '
                           f'Released ₱{total_released:,.2f} back to budget. '
                           f'Details: {", ".join(released_details)}'
                )
                
                # Create notification
                from apps.budgets.models import SystemNotification
                SystemNotification.objects.create(
                    recipient=purchase_request.submitted_by,
                    title='PR Rejected',
                    message=f'Your PR {purchase_request.pr_number} has been rejected. '
                            f'₱{total_released:,.2f} has been returned to your budget.',
                    content_type='pr',
                    object_id=purchase_request.id
                )
                
                messages.success(request, 
                    f'❌ Purchase Request {purchase_request.pr_number} has been rejected.')
                messages.info(request, 
                    f'💰 Released ₱{total_released:,.2f} back to budget')
                
                # Show detailed breakdown if multiple allocations
                if len(released_details) > 1:
                    for detail in released_details:
                        messages.info(request, f"  • {detail}")
                
            except Exception as e:
                print(f"❌ ERROR: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error processing rejection: {str(e)}")
                return redirect('department_pr_request')
        
        else:
            messages.error(request, "Invalid action specified.")

    return redirect('department_pr_request')

@role_required('admin', login_url='/admin/')
def admin_manual_pdf_upload(request, pr_id):
    """Admin manually uploads converted PDF if auto-conversion failed"""
    from apps.budgets.models import PurchaseRequest as NewPurchaseRequest
    
    pr = get_object_or_404(NewPurchaseRequest, id=pr_id)
    
    if request.method == 'POST' and request.FILES.get('manual_pdf'):
        pdf_file = request.FILES['manual_pdf']
        
        # Validate it's a PDF
        if not pdf_file.name.lower().endswith('.pdf'):
            messages.error(request, "Please upload a PDF file")
            return redirect('admin_preview_pr', pr_id=pr_id)
        
        # Save it
        pr.partially_approved_pdf = pdf_file
        pr.save()
        
        messages.success(request, "PDF uploaded successfully!")
    
    return redirect('admin_preview_pr', pr_id=pr_id)

@role_required('admin', login_url='/admin/')
def admin_upload_pr_signed_copy(request, pr_id):
    """
    Admin uploads scanned signed PR copy + signed supporting documents
    Changes status: Partially Approved → Approved
    """
    from apps.budgets.models import PurchaseRequest as NewPurchaseRequest
    from apps.budgets.models import PurchaseRequestSupportingDocument  # ✅ CORRECT MODEL
    
    pr = get_object_or_404(
        NewPurchaseRequest.objects.select_related('submitted_by', 'budget_allocation'),
        id=pr_id
    )
    
    # Check if PR is in correct status
    if pr.status != 'Partially Approved':
        messages.error(request, 
            f'Cannot upload signed copy. PR status is "{pr.status}". '
            f'Only "Partially Approved" PRs can receive signed documents.')
        return redirect('admin_preview_pr', pr_id=pr_id)
    
    if request.method == 'POST':
        main_signed_file = request.FILES.get('signed_copy')
        supporting_signed_files = request.FILES.getlist('supporting_signed_copies')  # Multiple files
        admin_notes = request.POST.get('notes', '').strip()
        
        if not main_signed_file:
            messages.error(request, "Main PR signed copy is required.")
            return redirect('admin_upload_pr_signed_copy', pr_id=pr_id)
        
        # Validate main file
        allowed_extensions = ['pdf', 'jpg', 'jpeg', 'png']
        file_ext = main_signed_file.name.split('.')[-1].lower()
        
        if file_ext not in allowed_extensions:
            messages.error(request, 
                f"Invalid file type for main document. Allowed: {', '.join(allowed_extensions).upper()}")
            return redirect('admin_upload_pr_signed_copy', pr_id=pr_id)
        
        # Validate file size (max 10MB)
        max_size = 10 * 1024 * 1024
        if main_signed_file.size > max_size:
            messages.error(request, "Main document size must be less than 10MB.")
            return redirect('admin_upload_pr_signed_copy', pr_id=pr_id)
        
        try:
            # Save the main signed copy
            pr.approved_documents = main_signed_file
            pr.status = 'Approved'
            pr.final_approved_at = timezone.now()
            pr.admin_notes = admin_notes
            pr.save()
            
            uploaded_count = 1  # Main document
            skipped_files = []
            
            # Process supporting signed documents
            if supporting_signed_files:
                for supporting_file in supporting_signed_files:
                    # Validate each supporting file
                    file_ext = supporting_file.name.split('.')[-1].lower()
                    
                    if file_ext not in allowed_extensions:
                        skipped_files.append(f"{supporting_file.name} (invalid format)")
                        continue
                    
                    if supporting_file.size > max_size:
                        skipped_files.append(f"{supporting_file.name} (too large)")
                        continue
                    
                    # ✅ Create supporting document record for signed copy
                    PurchaseRequestSupportingDocument.objects.create(
                        purchase_request=pr,
                        document=supporting_file,
                        file_name=supporting_file.name,
                        file_size=supporting_file.size,
                        uploaded_by=request.user,  # Track admin who uploaded
                        is_signed_copy=True  # Flag as signed copy
                    )
                    uploaded_count += 1
            
            # 🔒 Budget is now officially consumed (already deducted on submission)
            
            # Log audit trail
            log_audit_trail(
                request=request,
                action='APPROVE',
                model_name='PurchaseRequest',
                record_id=pr.id,
                detail=f'PR {pr.pr_number} fully approved. '
                       f'Uploaded {uploaded_count} signed document(s). '
                       f'Budget ₱{pr.total_amount:,.2f} officially consumed.'
            )
            
            # Create notification
            from apps.budgets.models import SystemNotification
            SystemNotification.objects.create(
                recipient=pr.submitted_by,
                title='PR Fully Approved',
                message=f'Your PR {pr.pr_number} has been fully approved. '
                        f'{uploaded_count} signed document(s) uploaded.',
                content_type='pr',
                object_id=pr.id
            )
            
            # Success message
            messages.success(request, 
                f'✅ PR {pr.pr_number} fully approved! '
                f'Uploaded {uploaded_count} signed document(s). '
                f'Budget ₱{pr.total_amount:,.2f} is officially consumed.')
            
            # Show warnings for skipped files
            if skipped_files:
                messages.warning(request, 
                    f"Skipped {len(skipped_files)} file(s): {', '.join(skipped_files)}")
            
        except Exception as e:
            print(f"❌ Error uploading signed copy: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error uploading file: {str(e)}')
        
        return redirect('admin_preview_pr', pr_id=pr_id)
    
    # GET request - show upload form
    original_docs = pr.supporting_documents.filter(is_signed_copy=False)

    context = {
        'pr': pr,
        'supporting_docs_count': original_docs.count(),
        'original_supporting_docs': original_docs,  # Pass the actual documents too
    }
    
    return render(request, 'admin_panel/upload_pr_signed_copy.html', context)

@role_required('admin', login_url='/admin/')
def admin_preview_pr(request, pr_id):
    """
    Admin preview of Purchase Request details
    """
    from apps.budgets.models import PurchaseRequest as NewPurchaseRequest
    
    pr = get_object_or_404(
        NewPurchaseRequest.objects.select_related(
            'submitted_by',
            'budget_allocation',
            'budget_allocation__approved_budget',
            'source_pre',
            'source_line_item',
            'source_line_item__category'
        ).prefetch_related(
            'supporting_documents',
            'pre_allocations',
            'pre_allocations__pre_line_item'
        ),
        id=pr_id
    )
    
    # Get allocation details
    allocation = pr.pre_allocations.first()
    
    context = {
        'pr': pr,
        'allocation': allocation,
    }
    
    return render(request, 'admin_panel/preview_pr.html', context)

@role_required('admin', login_url='/admin/')
def admin_preview_ad(request, ad_id):
    """
    Admin preview of Activity Design details with multiple line items
    """
    from apps.budgets.models import ActivityDesign, ActivityDesignAllocation
    from decimal import Decimal

    ad = get_object_or_404(
        ActivityDesign.objects.select_related(
            'submitted_by',
            'budget_allocation',
            'budget_allocation__approved_budget',
            'admin_approved_by'
        ).prefetch_related(
            'supporting_documents',
            'pre_allocations',
            'pre_allocations__pre_line_item',
            'pre_allocations__pre_line_item__category'
        ),
        id=ad_id
    )

    # Get all allocations with quarter remaining calculations
    allocations_data = []
    for allocation in ad.pre_allocations.all():
        line_item = allocation.pre_line_item
        quarter = allocation.quarter

        # Get quarter-specific available amount AFTER this allocation
        # This shows how much budget remains in this quarter after using this AD
        quarter_total = line_item.get_quarter_amount(quarter)
        quarter_consumed = line_item.get_quarter_consumed(quarter)
        quarter_remaining = quarter_total - quarter_consumed

        allocations_data.append({
            'allocation': allocation,
            'quarter_remaining': quarter_remaining
        })

    # Budget summary
    budget_summary = {
        'line_items_count': ad.pre_allocations.count(),
    }

    context = {
        'ad': ad,
        'allocations': allocations_data,
        'budget_summary': budget_summary,
    }

    return render(request, 'admin_panel/admin_preview_ad.html', context)

@role_required('admin', login_url='/admin/')
def admin_upload_ad_signed_copy(request, ad_id):
    """
    Admin uploads scanned signed AD copy + signed supporting documents
    Changes status: Partially Approved → Approved
    """
    from apps.budgets.models import ActivityDesign, ActivityDesignSupportingDocument

    ad = get_object_or_404(
        ActivityDesign.objects.select_related('submitted_by', 'budget_allocation'),
        id=ad_id
    )

    # Check if AD is in correct status
    if ad.status != 'Partially Approved':
        messages.error(request,
            f'Cannot upload signed copy. AD status is "{ad.status}". '
            f'Only "Partially Approved" ADs can receive signed documents.')
        return redirect('admin_preview_ad', ad_id=ad_id)

    if request.method == 'POST':
        main_signed_file = request.FILES.get('signed_copy')
        supporting_signed_files = request.FILES.getlist('supporting_signed_copies')  # Multiple files
        admin_notes = request.POST.get('notes', '').strip()

        if not main_signed_file:
            messages.error(request, "Main AD signed copy is required.")
            return redirect('admin_upload_ad_signed_copy', ad_id=ad_id)

        # Validate main file
        allowed_extensions = ['pdf', 'jpg', 'jpeg', 'png']
        file_ext = main_signed_file.name.split('.')[-1].lower()

        if file_ext not in allowed_extensions:
            messages.error(request,
                f"Invalid file type for main document. Allowed: {', '.join(allowed_extensions).upper()}")
            return redirect('admin_upload_ad_signed_copy', ad_id=ad_id)

        # Validate file size (max 10MB)
        max_size = 10 * 1024 * 1024
        if main_signed_file.size > max_size:
            messages.error(request, "Main document size must be less than 10MB.")
            return redirect('admin_upload_ad_signed_copy', ad_id=ad_id)

        try:
            # Save the main signed copy
            ad.approved_document = main_signed_file
            ad.status = 'Approved'
            ad.final_approved_at = timezone.now()
            ad.admin_approved_by = request.user
            ad.save()

            uploaded_count = 1  # Main document
            skipped_files = []

            # Process supporting signed documents
            if supporting_signed_files:
                for supporting_file in supporting_signed_files:
                    # Validate each supporting file
                    file_ext = supporting_file.name.split('.')[-1].lower()

                    if file_ext not in allowed_extensions:
                        skipped_files.append(f"{supporting_file.name} (invalid format)")
                        continue

                    if supporting_file.size > max_size:
                        skipped_files.append(f"{supporting_file.name} (too large)")
                        continue

                    # Create supporting document record for signed copy
                    ActivityDesignSupportingDocument.objects.create(
                        activity_design=ad,
                        document=supporting_file,
                        file_name=supporting_file.name,
                        file_size=supporting_file.size,
                        uploaded_by=request.user,  # Track admin who uploaded
                        is_signed_copy=True  # Flag as signed copy
                    )
                    uploaded_count += 1

            # Budget is now officially consumed (already deducted on submission)

            # Log audit trail
            log_audit_trail(
                request=request,
                action='APPROVE',
                model_name='ActivityDesign',
                record_id=ad.id,
                detail=f'AD {ad.ad_number} fully approved. '
                       f'Uploaded {uploaded_count} signed document(s). '
                       f'Budget ₱{ad.total_amount:,.2f} officially consumed.'
            )

            # Create notification
            from apps.budgets.models import SystemNotification
            SystemNotification.objects.create(
                recipient=ad.submitted_by,
                title='AD Fully Approved',
                message=f'Your AD {ad.ad_number} has been fully approved. '
                        f'{uploaded_count} signed document(s) uploaded.',
                content_type='ad',
                object_id=ad.id
            )

            # Success message
            messages.success(request,
                f'✅ AD {ad.ad_number} fully approved! '
                f'Uploaded {uploaded_count} signed document(s). '
                f'Budget ₱{ad.total_amount:,.2f} is officially consumed.')

            # Show warnings for skipped files
            if skipped_files:
                messages.warning(request,
                    f"Skipped {len(skipped_files)} file(s): {', '.join(skipped_files)}")

        except Exception as e:
            print(f"❌ Error uploading signed copy: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error uploading file: {str(e)}')

        return redirect('admin_preview_ad', ad_id=ad_id)

    # GET request - show upload form
    context = {
        'ad': ad,
    }
    return render(request, 'admin_panel/upload_ad_signed_copy.html', context)

@role_required('admin', login_url='/admin/')
def budget_allocation(request):
    """Handle budget allocation to departments and end users"""
    
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        
        if action == 'edit':
            # Handle Edit
            allocation_id = request.POST.get('allocation_id')
            new_amount = request.POST.get('amount')
            
            try:
                allocation = NewBudgetAllocation.objects.select_related('approved_budget').get(id=allocation_id)
                new_amount = Decimal(new_amount)
                
                # Calculate amount already used
                amount_used = (allocation.pre_amount_used + 
                             allocation.pr_amount_used + 
                             allocation.ad_amount_used)
                
                # Validation: new amount must be >= amount used
                if new_amount < amount_used:
                    messages.error(request, 
                        f"New allocation amount (₱{new_amount:,.2f}) cannot be less than "
                        f"amount already used (₱{amount_used:,.2f}).")
                    return redirect('budget_allocation')
                
                # Calculate difference
                old_amount = allocation.allocated_amount
                difference = new_amount - old_amount
                
                # If increasing allocation, check if approved budget has enough balance
                if difference > 0:
                    if difference > allocation.approved_budget.remaining_budget:
                        messages.error(request, 
                            f"Cannot increase allocation. Approved budget only has "
                            f"₱{allocation.approved_budget.remaining_budget:,.2f} remaining.")
                        return redirect('budget_allocation')
                    
                    # Decrease approved budget's remaining balance
                    allocation.approved_budget.remaining_budget -= difference
                    allocation.approved_budget.save()
                
                # If decreasing allocation, return amount to approved budget
                elif difference < 0:
                    # Return the difference to approved budget
                    allocation.approved_budget.remaining_budget += abs(difference)
                    allocation.approved_budget.save()
                
                # Update allocation
                allocation.allocated_amount = new_amount
                allocation.remaining_balance = new_amount - amount_used
                allocation.save()
                
                messages.success(request, 
                    f"Successfully updated allocation for {allocation.end_user.get_full_name()}. "
                    f"New amount: ₱{new_amount:,.2f} "
                    f"({'increased' if difference > 0 else 'decreased' if difference < 0 else 'unchanged'} by ₱{abs(difference):,.2f})")
                
                log_audit_trail(
                    request=request,
                    action='UPDATE',
                    model_name='BudgetAllocation',
                    record_id=allocation.id,
                    detail=f"Updated allocation for {allocation.end_user.get_full_name()} "
                           f"from ₱{old_amount:,.2f} to ₱{new_amount:,.2f}"
                )
                
                return redirect('budget_allocation')
                
            except NewBudgetAllocation.DoesNotExist:
                messages.error(request, "Allocation not found.")
                return redirect('budget_allocation')
            except Exception as e:
                messages.error(request, f"Error updating allocation: {str(e)}")
                return redirect('budget_allocation')
        
        else:
            # Get form data
            approved_budget_id = request.POST.get('approved_budget')
            mfo = request.POST.get('mfo')
            end_user_id = request.POST.get('end_user')
            amount = request.POST.get('amount')
            
            # Validation
            if not all([approved_budget_id, mfo, end_user_id, amount]):
                messages.error(request, "All fields are required.")
                return redirect('budget_allocation')
            
            try:
                # Get approved budget
                approved_budget = NewApprovedBudget.objects.get(id=approved_budget_id)
                
                # Get end user
                end_user = User.objects.get(id=end_user_id)
                
                # Validate amount
                amount = Decimal(amount)
                if amount <= 0:
                    messages.error(request, "Amount must be greater than zero.")
                    return redirect('budget_allocation')
                
                # Check if approved budget has enough remaining balance
                if amount > approved_budget.remaining_budget:
                    messages.error(request, 
                        f"Insufficient budget. Available: ₱{approved_budget.remaining_budget:,.2f}, "
                        f"Requested: ₱{amount:,.2f}")
                    return redirect('budget_allocation')
                
                # Check if allocation already exists for this user and budget
                existing_allocation = NewBudgetAllocation.objects.filter(
                    approved_budget=approved_budget,
                    end_user=end_user
                ).first()
                
                if existing_allocation:
                    messages.warning(request, 
                        f"Allocation already exists for {end_user.get_full_name()} "
                        f"({end_user.department}) under {approved_budget.title}. "
                        f"Please edit the existing allocation instead.")
                    return redirect('budget_allocation')
                
                # Create allocation - use end_user's department field
                allocation = NewBudgetAllocation.objects.create(
                    approved_budget=approved_budget,
                    department=end_user.department,  # Use end_user's department
                    end_user=end_user,
                    allocated_amount=amount,
                    remaining_balance=amount
                )
                
                # Update approved budget remaining balance
                approved_budget.remaining_budget -= amount
                approved_budget.save()
                
                messages.success(request, 
                    f"Successfully allocated ₱{amount:,.2f} to {end_user.get_full_name()} "
                    f"({end_user.department}) from {approved_budget.title}.")
                
                log_audit_trail(
                    request=request,
                    action='CREATE',
                    model_name='BudgetAllocation',
                    record_id=allocation.id,
                    detail=f"Allocated ₱{amount:,.2f} to {end_user.get_full_name()} "
                        f"- MFO: {mfo}, Department: {end_user.department} "
                        f"from {approved_budget.title} (FY {approved_budget.fiscal_year})"
                )
                
                return redirect('budget_allocation')
                
            except NewApprovedBudget.DoesNotExist:
                messages.error(request, "Selected approved budget not found.")
                return redirect('budget_allocation')
            except User.DoesNotExist:
                messages.error(request, "Selected end user not found.")
                return redirect('budget_allocation')
            except Exception as e:
                messages.error(request, f"Error creating allocation: {str(e)}")
                return redirect('budget_allocation')
    
    # GET request - Display allocations
    current_year = str(datetime.now().year)
    
    # Get year filter for summary cards
    summary_year = request.GET.get('summary_year', current_year)
    
    # Get allocations for summary statistics based on selected year
    if summary_year == 'all':
        summary_allocations = NewBudgetAllocation.objects.all()
    else:
        summary_allocations = NewBudgetAllocation.objects.filter(
            approved_budget__fiscal_year=summary_year
        )
    
    # Calculate summary statistics based on selected year
    allocation_stats = summary_allocations.aggregate(
        total_allocated=Sum('allocated_amount'),
        total_remaining=Sum('remaining_balance'),
        total_departments=Count('department', distinct=True)
    )
    
    total_allocated = allocation_stats['total_allocated'] or Decimal('0')
    total_remaining = allocation_stats['total_remaining'] or Decimal('0')
    total_departments = allocation_stats['total_departments'] or 0
    
    # Calculate utilization rate
    utilization_rate = 0
    if total_allocated > 0:
        total_used = total_allocated - total_remaining
        utilization_rate = round((total_used / total_allocated) * 100, 1)
    
    # Get all allocations for the table (not filtered by year)
    allocations_list = NewBudgetAllocation.objects.select_related(
        'approved_budget', 'end_user'
    ).order_by('-allocated_at')
    
    # Apply table filters
    fiscal_year_filter = request.GET.get('fiscal_year')
    mfo_filter = request.GET.get('mfo')
    department_filter = request.GET.get('department')
    amount_min = request.GET.get('amount_min')
    amount_max = request.GET.get('amount_max')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search')
    
    if fiscal_year_filter:
        allocations_list = allocations_list.filter(approved_budget__fiscal_year=fiscal_year_filter)
    
    if mfo_filter:
        allocations_list = allocations_list.filter(end_user__mfo=mfo_filter)
    
    if department_filter:
        allocations_list = allocations_list.filter(department__icontains=department_filter)
    
    if amount_min:
        try:
            allocations_list = allocations_list.filter(allocated_amount__gte=Decimal(amount_min))
        except:
            pass
    
    if amount_max:
        try:
            allocations_list = allocations_list.filter(allocated_amount__lte=Decimal(amount_max))
        except:
            pass
    
    if date_from:
        allocations_list = allocations_list.filter(allocated_at__date__gte=date_from)
    
    if date_to:
        allocations_list = allocations_list.filter(allocated_at__date__lte=date_to)
    
    if search_query:
        allocations_list = allocations_list.filter(
            Q(end_user__fullname__icontains=search_query) |
            Q(end_user__username__icontains=search_query) |
            Q(end_user__email__icontains=search_query)
        )
    
    # Get data for dropdowns
    approved_budgets = NewApprovedBudget.objects.filter(
        is_active=True,
        remaining_budget__gt=0
    ).order_by('-fiscal_year')
    
    # Get unique MFOs from users
    mfos = User.objects.filter(
        mfo__isnull=False
    ).exclude(mfo='').values_list('mfo', flat=True).distinct().order_by('mfo')
    
    # Get end users grouped by MFO
    end_users_by_mfo = {}
    for mfo in mfos:
        users = User.objects.filter(
            is_superuser=False,
            is_admin=False,
            mfo=mfo
        ).order_by('department', 'fullname')
        
        if users.exists():
            end_users_by_mfo[mfo] = list(users)
    
    # Get available years for dropdown (from approved budgets linked to allocations)
    available_years = NewBudgetAllocation.objects.select_related('approved_budget').values_list(
        'approved_budget__fiscal_year', flat=True
    ).distinct().order_by('-approved_budget__fiscal_year')
    
    # Pagination
    paginator = Paginator(allocations_list, 10)
    page_number = request.GET.get('page')
    
    try:
        allocations = paginator.page(page_number)
    except PageNotAnInteger:
        allocations = paginator.page(1)
    except EmptyPage:
        allocations = paginator.page(paginator.num_pages)
    
    context = {
        'allocations': allocations,
        'total_allocated': total_allocated,
        'total_remaining': total_remaining,
        'total_departments': total_departments,
        'utilization_rate': utilization_rate,
        'approved_budgets': approved_budgets,
        'mfos': mfos,
        'end_users_by_mfo': end_users_by_mfo,
        'available_years': available_years,
        'selected_year': summary_year,
        'current_year': current_year,
    }
    
    return render(request, 'admin_panel/budget_allocation.html', context)

@role_required('admin', login_url='/admin/')
def export_allocation_excel(request, allocation_id):
    """Export single budget allocation to Excel"""
    allocation = get_object_or_404(
        NewBudgetAllocation.objects.select_related('approved_budget', 'end_user'), 
        id=allocation_id
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Allocation {allocation_id}"
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = "BUDGET ALLOCATION REPORT"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    current_row = 3
    
    # Allocation ID
    ws[f'A{current_row}'] = "Allocation ID"
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'A{current_row}'].border = thin_border
    ws[f'B{current_row}'] = allocation.id
    ws[f'B{current_row}'].border = thin_border
    current_row += 1
    
    # Approved Budget Section
    current_row += 1
    ws[f'A{current_row}'] = "APPROVED BUDGET INFORMATION"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    budget_data = [
        ("Budget Title:", allocation.approved_budget.title),
        ("Fiscal Year:", allocation.approved_budget.fiscal_year),
        ("Total Budget Amount:", f"₱{allocation.approved_budget.amount:,.2f}"),
        ("Budget Remaining:", f"₱{allocation.approved_budget.remaining_budget:,.2f}"),
    ]
    
    for label, value in budget_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].border = thin_border
        ws[f'B{current_row}'] = value
        ws[f'B{current_row}'].border = thin_border
        current_row += 1
    
    # User Information Section
    current_row += 1
    ws[f'A{current_row}'] = "END USER INFORMATION"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    user_data = [
        ("Full Name:", allocation.end_user.fullname),
        ("Username:", allocation.end_user.username),
        ("Email:", allocation.end_user.email),
        ("MFO:", allocation.end_user.mfo or "N/A"),
        ("Department:", allocation.department),
        ("Position:", allocation.end_user.position or "N/A"),
    ]
    
    for label, value in user_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].border = thin_border
        ws[f'B{current_row}'] = value
        ws[f'B{current_row}'].border = thin_border
        current_row += 1
    
    # Financial Information Section
    current_row += 1
    ws[f'A{current_row}'] = "FINANCIAL SUMMARY"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    financial_data = [
        ("Allocated Amount:", f"₱{allocation.allocated_amount:,.2f}"),
        ("Remaining Balance:", f"₱{allocation.remaining_balance:,.2f}"),
        ("PRE Amount Used:", f"₱{allocation.pre_amount_used:,.2f}"),
        ("PR Amount Used:", f"₱{allocation.pr_amount_used:,.2f}"),
        ("AD Amount Used:", f"₱{allocation.ad_amount_used:,.2f}"),
    ]
    
    total_used = allocation.pre_amount_used + allocation.pr_amount_used + allocation.ad_amount_used
    financial_data.append(("Total Amount Used:", f"₱{total_used:,.2f}"))
    
    # Calculate utilization rate
    utilization_rate = (total_used / allocation.allocated_amount * 100) if allocation.allocated_amount > 0 else 0
    financial_data.append(("Utilization Rate:", f"{utilization_rate:.2f}%"))
    
    for label, value in financial_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].border = thin_border
        ws[f'B{current_row}'] = value
        ws[f'B{current_row}'].border = thin_border
        
        # Highlight total row
        if "Total Amount Used" in label:
            ws[f'A{current_row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'B{current_row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'B{current_row}'].font = Font(bold=True)
        
        current_row += 1
    
    # Timeline Section
    current_row += 1
    ws[f'A{current_row}'] = "TIMELINE"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    ws[f'A{current_row}'] = "Allocated At:"
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'A{current_row}'].border = thin_border
    ws[f'B{current_row}'] = allocation.allocated_at.strftime("%B %d, %Y %I:%M %p")
    ws[f'B{current_row}'].border = thin_border
    current_row += 1
    
    # Footer
    current_row += 2
    ws[f'A{current_row}'] = f"Generated on: {allocation.allocated_at.strftime('%B %d, %Y %I:%M %p')}"
    ws[f'A{current_row}'].font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(f'A{current_row}:B{current_row}')
    
    # Prepare response
    filename = f"Budget_Allocation_{allocation.id}_{allocation.end_user.username}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@role_required('admin', login_url='/admin/')
def bulk_export_allocations(request):
    """Export all budget allocations to Excel"""
    
    # Get all allocations
    allocations = NewBudgetAllocation.objects.select_related(
        'approved_budget', 'end_user'
    ).order_by('-allocated_at')
    
    # Apply filters if any (same as the table view)
    summary_year = request.GET.get('summary_year')
    if summary_year and summary_year != 'all':
        allocations = allocations.filter(approved_budget__fiscal_year=summary_year)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Allocations"
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 22
    
    # Title
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = "BUDGET ALLOCATIONS REPORT"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    # Export info
    from django.utils import timezone
    ws.merge_cells('A2:M2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {timezone.now().strftime('%B %d, %Y %I:%M %p')} | Total Records: {allocations.count()}"
    info_cell.font = Font(italic=True, size=10, color="666666")
    info_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = [
        'ID',
        'Approved Budget',
        'FY',
        'MFO',
        'Department',
        'End User',
        'Allocated (₱)',
        'Remaining (₱)',
        'PRE Used (₱)',
        'PR Used (₱)',
        'AD Used (₱)',
        'Utilization %',
        'Allocated At'
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[4].height = 25
    
    # Data rows
    row_num = 5
    for allocation in allocations:
        total_used = (allocation.pre_amount_used + 
                     allocation.pr_amount_used + 
                     allocation.ad_amount_used)
        
        utilization = (total_used / allocation.allocated_amount * 100) if allocation.allocated_amount > 0 else 0
        
        row_data = [
            allocation.id,
            allocation.approved_budget.title,
            allocation.approved_budget.fiscal_year,
            allocation.end_user.mfo or 'N/A',
            allocation.department,
            f"{allocation.end_user.fullname} ({allocation.end_user.username})",
            float(allocation.allocated_amount),
            float(allocation.remaining_balance),
            float(allocation.pre_amount_used),
            float(allocation.pr_amount_used),
            float(allocation.ad_amount_used),
            round(utilization, 2),
            allocation.allocated_at.strftime('%b %d, %Y %I:%M %p')
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            
            # Format currency columns
            if col_num in [7, 8, 9, 10, 11]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Format percentage column
            if col_num == 12:
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color code utilization
                if utilization < 50:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif utilization < 80:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C6500")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
            
            # Alternate row colors
            if row_num % 2 == 0:
                if col_num != 12:  # Don't override utilization coloring
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        row_num += 1
    
    # Summary row
    row_num += 1
    summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    
    # Calculate totals
    total_allocated = sum(float(a.allocated_amount) for a in allocations)
    total_remaining = sum(float(a.remaining_balance) for a in allocations)
    total_pre_used = sum(float(a.pre_amount_used) for a in allocations)
    total_pr_used = sum(float(a.pr_amount_used) for a in allocations)
    total_ad_used = sum(float(a.ad_amount_used) for a in allocations)
    total_used = total_pre_used + total_pr_used + total_ad_used
    avg_utilization = (total_used / total_allocated * 100) if total_allocated > 0 else 0
    
    ws.merge_cells(f'A{row_num}:F{row_num}')
    summary_label = ws.cell(row=row_num, column=1)
    summary_label.value = "TOTAL / AVERAGE"
    summary_label.font = summary_font
    summary_label.fill = summary_fill
    summary_label.alignment = Alignment(horizontal='right', vertical='center')
    summary_label.border = thin_border
    
    # Totals
    totals = [
        total_allocated,
        total_remaining,
        total_pre_used,
        total_pr_used,
        total_ad_used,
        round(avg_utilization, 2)
    ]
    
    for col_offset, total in enumerate(totals, 7):
        cell = ws.cell(row=row_num, column=col_offset)
        cell.value = total
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right' if col_offset != 12 else 'center', vertical='center')
        
        if col_offset in [7, 8, 9, 10, 11]:
            cell.number_format = '#,##0.00'
        elif col_offset == 12:
            cell.number_format = '0.00"%"'
    
    # Empty cell for last column
    cell = ws.cell(row=row_num, column=13)
    cell.fill = summary_fill
    cell.border = thin_border
    
    # Prepare response
    from datetime import datetime
    filename = f"Budget_Allocations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@role_required('admin', login_url='/admin/')
def institutional_funds(request):
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        
        if action == 'edit':
            # Handle Edit
            budget_id = request.POST.get('budget_id')
            title = request.POST.get('title')
            fiscal_year = request.POST.get('fiscal_year')  # Now editable
            amount = request.POST.get('amount')
            description = request.POST.get('description', '')
            new_documents = request.FILES.getlist('supporting_documents')
            remove_doc_ids = request.POST.getlist('remove_documents[]')
            
            try:
                budget = NewApprovedBudget.objects.get(id=budget_id)
                
                # Validate fiscal year
                if not fiscal_year:
                    messages.error(request, "Fiscal year is required.")
                    return redirect("institutional_funds")
                
                # Check if fiscal year is being changed and if new fiscal year already exists
                if budget.fiscal_year != fiscal_year:
                    if NewApprovedBudget.objects.filter(fiscal_year=fiscal_year).exclude(id=budget_id).exists():
                        messages.error(request, f"A budget for fiscal year {fiscal_year} already exists.")
                        return redirect("institutional_funds")
                
                # Validate amount
                amount = Decimal(amount)
                if amount <= 0:
                    messages.error(request, "Amount must be greater than zero.")
                    return redirect("institutional_funds")
                
                # Update budget fields
                budget.title = title
                budget.fiscal_year = fiscal_year  # Now updating fiscal year
                budget.amount = amount
                budget.description = description
                
                # Recalculate remaining budget if amount changed
                old_allocated = budget.amount - budget.remaining_budget
                budget.remaining_budget = amount - old_allocated
                
                budget.save()
                
                # Remove documents marked for deletion
                if remove_doc_ids:
                    SupportingDocument.objects.filter(
                        id__in=remove_doc_ids,
                        approved_budget=budget
                    ).delete()
                
                # Add new documents
                if new_documents:
                    allowed_formats = ['pdf', 'doc', 'docx', 'xls', 'xlsx']
                    max_file_size = 10 * 1024 * 1024
                    
                    for doc in new_documents:
                        if doc.size > max_file_size:
                            messages.warning(request, f"File '{doc.name}' exceeds 10MB limit and was skipped.")
                            continue
                        
                        ext = doc.name.split('.')[-1].lower()
                        if ext not in allowed_formats:
                            messages.warning(request, f"File '{doc.name}' has invalid format and was skipped.")
                            continue
                        
                        SupportingDocument.objects.create(
                            approved_budget=budget,
                            document=doc,
                            uploaded_by=request.user
                        )
                
                messages.success(request, f"Budget '{title}' for FY {fiscal_year} updated successfully!")
                
                log_audit_trail(
                    request=request,
                    action='UPDATE',
                    model_name='ApprovedBudget',
                    record_id=budget.id,
                    detail=f"Updated approved budget: {title} for fiscal year {fiscal_year}"
                )
                
                return redirect("institutional_funds")
                
            except NewApprovedBudget.DoesNotExist:
                messages.error(request, "Budget not found.")
                return redirect("institutional_funds")
            except Exception as e:
                messages.error(request, f"Error updating budget: {str(e)}")
                return redirect("institutional_funds")
    
        else:
            # Handle Create (existing code)
            title = request.POST.get('title')
            fiscal_year = request.POST.get('fiscal_year')
            amount = request.POST.get('amount')
            description = request.POST.get('description', '')
            supporting_documents = request.FILES.getlist('supporting_documents')  # Multiple files
            
            # Validation
            if not all([title, fiscal_year, amount]):
                messages.error(request, "Title, Fiscal Year, and Amount are required.")
                return redirect("institutional_funds")
            
            if not supporting_documents:
                messages.error(request, "At least one supporting document is required.")
                return redirect("institutional_funds")
            
            try:
                amount = Decimal(amount)
                if amount <= 0:
                    messages.error(request, "Amount must be greater than zero.")
                    return redirect("institutional_funds")
            except (ValueError, TypeError, decimal.InvalidOperation):
                messages.error(request, "Invalid amount entered.")
                return redirect("institutional_funds")
            
            # Check if budget for this fiscal year already exists
            if NewApprovedBudget.objects.filter(fiscal_year=fiscal_year).exists():
                messages.error(request, f"Approved budget for fiscal year {fiscal_year} already exists.")
                return redirect("institutional_funds")
            
            # Validate files
            allowed_formats = ['pdf', 'doc', 'docx', 'xls', 'xlsx']
            max_file_size = 10 * 1024 * 1024  # 10MB
            
            for doc in supporting_documents:
                # Check file size
                if doc.size > max_file_size:
                    messages.error(request, f"File '{doc.name}' exceeds 10MB limit.")
                    return redirect("institutional_funds")
                
                # Check file format
                ext = doc.name.split('.')[-1].lower()
                if ext not in allowed_formats:
                    messages.error(request, f"File '{doc.name}' has invalid format. Allowed: {', '.join(allowed_formats).upper()}")
                    return redirect("institutional_funds")
            
            # Create the approved budget
            try:
                budget = NewApprovedBudget.objects.create(
                    title=title,
                    fiscal_year=fiscal_year,
                    amount=amount,
                    remaining_budget=amount,
                    description=description,
                    created_by=request.user
                )
                
                # Create supporting document records for each uploaded file
                document_count = 0
                for doc in supporting_documents:
                    SupportingDocument.objects.create(
                        approved_budget=budget,
                        document=doc,
                        uploaded_by=request.user
                    )
                    document_count += 1
                
                messages.success(request, 
                    f"Approved budget '{title}' for FY {fiscal_year} successfully added with {document_count} supporting document(s).")
                
                log_audit_trail(
                    request=request,
                    action='CREATE',
                    model_name='ApprovedBudget',
                    record_id=budget.id,
                    detail=f"Created approved budget: {title} for fiscal year {fiscal_year} with amount ₱{intcomma(amount)}. Uploaded {document_count} supporting documents."
                )
                
                return redirect("institutional_funds")
                
            except Exception as e:
                messages.error(request, f"Error creating approved budget: {str(e)}")
                return redirect("institutional_funds")
    
    # GET request - Display budgets
    current_year = str(datetime.now().year)
    
    # Get year filter for summary cards (separate from table filters)
    summary_year = request.GET.get('summary_year', current_year)
    
    # Get budgets for summary statistics
    if summary_year == 'all':
        summary_budgets = NewApprovedBudget.objects.all()
    else:
        summary_budgets = NewApprovedBudget.objects.filter(fiscal_year=summary_year)
    
    # Calculate summary statistics based on selected year
    summary_stats = summary_budgets.aggregate(
        total_approved=Sum('amount'),
        total_remaining=Sum('remaining_budget'),
        total_count=Count('id')
    )
    
    total_approved_budget = summary_stats['total_approved'] or Decimal('0')
    total_remaining_budget = summary_stats['total_remaining'] or Decimal('0')
    total_budget_count = summary_stats['total_count'] or 0
    
    # Calculate utilization rate
    budget_utilization_rate = 0
    if total_approved_budget > 0:
        total_allocated = total_approved_budget - total_remaining_budget
        budget_utilization_rate = round((total_allocated / total_approved_budget) * 100, 1)
    
    # Get table budgets (with all table filters applied)
    approved_budgets_list = NewApprovedBudget.objects.order_by('-created_at')
    
    # Apply table filters (keep your existing filter logic)
    fiscal_year_filter = request.GET.get('fiscal_year')
    amount_min = request.GET.get('amount_min')
    amount_max = request.GET.get('amount_max')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search')
    
    if fiscal_year_filter:
        approved_budgets_list = approved_budgets_list.filter(fiscal_year=fiscal_year_filter)
    
    if amount_min:
        try:
            approved_budgets_list = approved_budgets_list.filter(amount__gte=Decimal(amount_min))
        except:
            pass
    
    if amount_max:
        try:
            approved_budgets_list = approved_budgets_list.filter(amount__lte=Decimal(amount_max))
        except:
            pass
    
    if date_from:
        approved_budgets_list = approved_budgets_list.filter(created_at__date__gte=date_from)
    
    if date_to:
        approved_budgets_list = approved_budgets_list.filter(created_at__date__lte=date_to)
    
    if search_query:
        from django.db.models import Q
        approved_budgets_list = approved_budgets_list.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(fiscal_year__icontains=search_query)
        )
    
    # Get available years for dropdown
    available_years = NewApprovedBudget.objects.values_list('fiscal_year', flat=True).distinct().order_by('-fiscal_year')
    
    # Pagination
    paginator = Paginator(approved_budgets_list, 10)
    page_number = request.GET.get('page')
    
    try:
        approved_budgets = paginator.page(page_number)
    except PageNotAnInteger:
        approved_budgets = paginator.page(1)
    except EmptyPage:
        approved_budgets = paginator.page(paginator.num_pages)
    
    context = {
        'approved_budgets': approved_budgets,
        'total_approved_budget': total_approved_budget,
        'total_remaining_budget': total_remaining_budget,
        'total_budget_count': total_budget_count,
        'budget_utilization_rate': budget_utilization_rate,
        'available_years': available_years,
        'selected_year': summary_year,
        'current_year': current_year,
    }
    
    return render(request, 'admin_panel/institutional_funds.html', context)

@role_required('admin', login_url='/admin/')
def download_document(request, document_id):
    """Download supporting document with 'Save As' dialog"""
    document = get_object_or_404(SupportingDocument, id=document_id)
    
    # Check if file exists
    if not os.path.exists(document.document.path):
        raise Http404("Document file not found")
    
    # Open the file
    file_handle = open(document.document.path, 'rb')
    
    # Create response with attachment header to force download dialog
    response = FileResponse(file_handle)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = f'attachment; filename="{document.file_name}"'
    
    return response

@role_required('admin', login_url='/admin/')
def export_budget_excel(request, budget_id):
    """Export single approved budget to Excel"""
    budget = get_object_or_404(NewApprovedBudget, id=budget_id)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Budget {budget.fiscal_year}"
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = "APPROVED BUDGET REPORT"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    
    # Add some spacing
    current_row = 3
    
    # Budget Information Section
    ws[f'A{current_row}'] = "Budget Information"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    # Budget details
    budget_data = [
        ("Title:", budget.title),
        ("Fiscal Year:", budget.fiscal_year),
        ("Approved Amount:", f"₱{budget.amount:,.2f}"),
        ("Remaining Balance:", f"₱{budget.remaining_budget:,.2f}"),
        ("Allocated Amount:", f"₱{budget.amount - budget.remaining_budget:,.2f}"),
        ("Created By:", budget.created_by.get_full_name() if budget.created_by else "N/A"),
        ("Created At:", budget.created_at.strftime("%B %d, %Y %I:%M %p")),
        ("Description:", budget.description or "N/A"),
    ]
    
    for label, value in budget_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'B{current_row}'] = value
        ws[f'B{current_row}'].border = thin_border
        current_row += 1
    
    # Supporting Documents Section
    current_row += 2
    ws[f'A{current_row}'] = "Supporting Documents"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    # Document headers
    ws[f'A{current_row}'] = "File Name"
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'A{current_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws[f'A{current_row}'].border = thin_border
    
    ws[f'B{current_row}'] = "File Type"
    ws[f'B{current_row}'].font = Font(bold=True)
    ws[f'B{current_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws[f'B{current_row}'].border = thin_border
    current_row += 1
    
    # List documents
    documents = budget.supporting_documents.all()
    if documents.exists():
        for doc in documents:
            ws[f'A{current_row}'] = doc.file_name
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'] = doc.file_format.upper()
            ws[f'B{current_row}'].border = thin_border
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No documents attached"
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'A{current_row}'].font = Font(italic=True, color="999999")
        ws[f'A{current_row}'].border = thin_border
    
    # Footer
    current_row += 2
    ws[f'A{current_row}'] = f"Generated on: {budget.created_at.strftime('%B %d, %Y %I:%M %p')}"
    ws[f'A{current_row}'].font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(f'A{current_row}:B{current_row}')
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Budget_{budget.fiscal_year}_{budget.title.replace(" ", "_")}.xlsx"'
    
    wb.save(response)
    return response

@role_required('admin', login_url='/admin/')
def bulk_export_budgets(request):
    """Export all approved budgets to Excel with filters applied"""
    
    # Get all budgets with same filters as the table view
    budgets = NewApprovedBudget.objects.order_by('-created_at')
    
    # Apply filters from GET parameters
    fiscal_year_filter = request.GET.get('fiscal_year')
    amount_min = request.GET.get('amount_min')
    amount_max = request.GET.get('amount_max')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search')
    
    if fiscal_year_filter:
        budgets = budgets.filter(fiscal_year=fiscal_year_filter)
    
    if amount_min:
        try:
            budgets = budgets.filter(amount__gte=Decimal(amount_min))
        except:
            pass
    
    if amount_max:
        try:
            budgets = budgets.filter(amount__lte=Decimal(amount_max))
        except:
            pass
    
    if date_from:
        budgets = budgets.filter(created_at__date__gte=date_from)
    
    if date_to:
        budgets = budgets.filter(created_at__date__lte=date_to)
    
    if search_query:
        budgets = budgets.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(fiscal_year__icontains=search_query)
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Budgets"
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 22
    
    # Title
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = "APPROVED BUDGETS REPORT"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    # Export info
    from django.utils import timezone
    ws.merge_cells('A2:I2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {timezone.now().strftime('%B %d, %Y %I:%M %p')} | Total Records: {budgets.count()}"
    info_cell.font = Font(italic=True, size=10, color="666666")
    info_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = [
        'ID', 
        'Title', 
        'Fiscal Year', 
        'Amount (₱)', 
        'Remaining Balance (₱)', 
        'Allocated (₱)',
        'Utilization %',
        'Documents', 
        'Created At'
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[4].height = 25
    
    # Data rows
    row_num = 5
    for budget in budgets:
        allocated = budget.amount - budget.remaining_budget
        utilization = (allocated / budget.amount * 100) if budget.amount > 0 else 0
        
        # Get document count and formats
        docs = budget.supporting_documents.all()
        doc_info = f"{docs.count()} file(s)"
        if docs.exists():
            formats = ", ".join(set([doc.file_format.upper() for doc in docs]))
            doc_info = f"{docs.count()} file(s) ({formats})"
        
        row_data = [
            budget.id,
            budget.title,
            budget.fiscal_year,
            float(budget.amount),
            float(budget.remaining_budget),
            float(allocated),
            round(utilization, 2),
            doc_info,
            budget.created_at.strftime('%b %d, %Y %I:%M %p')
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            
            # Format currency columns
            if col_num in [4, 5, 6]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Format percentage column
            if col_num == 7:
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        row_num += 1
    
    # Summary row
    row_num += 1
    summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    
    # Calculate totals
    total_amount = sum(float(b.amount) for b in budgets)
    total_remaining = sum(float(b.remaining_budget) for b in budgets)
    total_allocated = total_amount - total_remaining
    avg_utilization = (total_allocated / total_amount * 100) if total_amount > 0 else 0
    
    ws.merge_cells(f'A{row_num}:C{row_num}')
    summary_label = ws.cell(row=row_num, column=1)
    summary_label.value = "TOTAL / AVERAGE"
    summary_label.font = summary_font
    summary_label.fill = summary_fill
    summary_label.alignment = Alignment(horizontal='right', vertical='center')
    summary_label.border = thin_border
    
    # Total amount
    cell = ws.cell(row=row_num, column=4)
    cell.value = total_amount
    cell.font = summary_font
    cell.fill = summary_fill
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = thin_border
    
    # Total remaining
    cell = ws.cell(row=row_num, column=5)
    cell.value = total_remaining
    cell.font = summary_font
    cell.fill = summary_fill
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = thin_border
    
    # Total allocated
    cell = ws.cell(row=row_num, column=6)
    cell.value = total_allocated
    cell.font = summary_font
    cell.fill = summary_fill
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = thin_border
    
    # Average utilization
    cell = ws.cell(row=row_num, column=7)
    cell.value = round(avg_utilization, 2)
    cell.font = summary_font
    cell.fill = summary_fill
    cell.number_format = '0.00"%"'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    
    # Empty cells for remaining columns
    for col in [8, 9]:
        cell = ws.cell(row=row_num, column=col)
        cell.fill = summary_fill
        cell.border = thin_border
    
    # Prepare response
    from datetime import datetime
    filename = f"Approved_Budgets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@role_required('admin', login_url='/admin/')
def admin_logout(request):
    log_audit_trail(
        request=request,
        action='LOGOUT',
        model_name='User',
        record_id=request.user.id,
        detail=f"User {request.user.username} logged out."
    )
    logout(request)
    return redirect('admin_login')

@role_required('admin', login_url='/admin/')
def audit_trail(request):
    
    # Get all audit records and users
    audit_records = AuditTrail.objects.select_related('user').all()
    departments = User.objects.all().values_list('department', flat=True).distinct()
    
    # Filter by department
    department = request.GET.get('department')
    if department:
        audit_records = audit_records.filter(user__department=department)
    
    # Filter by action if specified
    action_filter = request.GET.get('action')
    if action_filter:
        audit_records = audit_records.filter(action=action_filter)
        
    # Filter by date range if specified
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        audit_records = audit_records.filter(
            timestamp__date__range=[start_date, end_date]
        )
        
    # Pagination
    paginator = Paginator(audit_records, 15)  # 15 records per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'action_choices': AuditTrail.ACTION_CHOICES,
        'departments': departments,
    }
    return render(request, 'admin_panel/audit_trail.html', context)

@role_required('admin', login_url='/admin/')
def budget_re_alignment(request):
    try:
        re_alignment_request = Budget_Realignment.objects.all()
    except Budget_Realignment.DoesNotExist:
        re_alignment_request = None
    return render(request, 'admin_panel/budget_re-alignment.html', {'re_alignment_request': re_alignment_request})

@role_required('admin', login_url='/admin/')
def handle_re_alignment_request_action(request, pk):
    req = get_object_or_404(Budget_Realignment, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            try:
                with transaction.atomic():
                    source_alloc = BudgetAllocation.objects.select_for_update().get(papp=req.source_papp)
                    target_alloc = BudgetAllocation.objects.select_for_update().get(papp=req.target_papp)
                    
                    if source_alloc.remaining_budget < req.amount:
                        messages.error(request, "Insufficient budget in source PAPP to realign.")
                        return redirect('budget_re_alignment')
                    
                    # Deduct from source and add to target
                    source_alloc.remaining_budget -= req.amount
                    target_alloc.remaining_budget += req.amount
                    
                    # Save both allocations
                    source_alloc.save()
                    target_alloc.save()

                    # Update request status
                    req.status = 'Approved'
                    req.approved_by = request.user
                    req.save()
                    
                    messages.success(request, f'Budget Realignment Request of {req.requested_by.department} has been approved.')
            except BudgetAllocation.DoesNotExist:
                messages.error(request, "Source or Target PAPP not found.")
            except Exception as e:
                messages.error(request, f"Something went wrong: {e}")
        elif action == 'reject':
            req.status = 'rejected'
            req.approved_by = request.user
            req.save()
            messages.error(request, f'Budget Realignment Request of {req.requested_by.department} has been rejected.')

        
        
        # action = request.POST.get('action')
        # if action == 'approve':
        #     req.status = 'Approved'
        #     messages.success(request, f'Budget Realignment Request of {req.requested_by.department} has been approved.')
        # elif action == 'reject':
        #     req.status = 'Rejected'
        #     messages.error(request, f'Budget Realignment Request of {req.requested_by.department} has been rejected.')
        # req.approved_by = request.user
        # req.save()
        
    return redirect('budget_realignment')

@role_required('admin', login_url='/admin/')
def pre_request_page(request):
    # Filter only PREs that were approved by the approving officer
    """
    View for the Admin to view all approved PREs from different departments.

    This view shows a table of all approved PREs, with the ability to filter by department.
    The table shows the department, submitted by, submitted on, budget allocation, and status of each PRE.
    """
    
    STATUS_CHOICES = (
        ('Pending', 'Pending'),
        ('Partially Approved', 'Partially Approved'),
        ('Rejected', 'Rejected'),
        ('Approved', 'Approved')
    )
    
    pres = DepartmentPRE.objects.all().select_related('submitted_by', 'budget_allocation__approved_budget').order_by('-created_at')

    # Department filter
    dept = request.GET.get('department')
    if dept:
        pres = pres.filter(department=dept)
        
    # Status filter
    status = request.GET.get('status')
    if status:
        pres = pres.filter(status=status)

    # Distinct department list for filter select
    departments = DepartmentPRE.objects.all().values_list('department', flat=True).distinct()
    
    context = {
        'pres': pres,
        'departments': departments,
        'status_choices': STATUS_CHOICES,
    }

    return render(request, "admin_panel/pre_request.html", context)


@role_required('admin', login_url='/admin/')
def admin_preview_pre(request, pk: int):
    pre = get_object_or_404(DepartmentPRE.objects.select_related('submitted_by', 'budget_allocation__approved_budget'), pk=pk)

    # Reuse the same simple aggregation scheme as AO preview
    from decimal import Decimal

    def to_decimal(value):
        try:
            return Decimal(str(value)) if value not in (None, "") else Decimal('0')
        except Exception:
            return Decimal('0')

    payload = pre.data or {}

    sections_spec = [
        {
            'title': 'Personnel Services',
            'color_class': 'bg-yellow-100',
            'items': [
                {'label': 'Basic Salary', 'name': 'basic_salary'},
                {'label': 'Honoraria', 'name': 'honoraria'},
                {'label': 'Overtime Pay', 'name': 'overtime_pay'},
            ]
        },
        {
            'title': 'Maintenance and Other Operating Expenses',
            'color_class': 'bg-blue-100',
            'items': [
                {'label': 'Travelling Expenses', 'is_group': True},
                {'label': 'Travelling expenses-local', 'name': 'travel_local', 'indent': True},
                {'label': 'Travelling Expenses-foreign', 'name': 'travel_foreign', 'indent': True},
                {'label': 'Training and Scholarship expenses', 'is_group': True},
                {'label': 'Training Expenses', 'name': 'training_expenses', 'indent': True},
                {'label': 'Supplies and materials expenses', 'is_group': True},
                {'label': 'Office Supplies Expenses', 'name': 'office_supplies_expenses', 'indent': True},
                {'label': 'Accountable Form Expenses', 'name': 'accountable_form_expenses', 'indent': True},
                {'label': 'Agricultural and Marine Supplies Expenses', 'name': 'agri_marine_supplies_expenses', 'indent': True},
                {'label': 'Drugs and Medicines', 'name': 'drugs_medicines', 'indent': True},
                {'label': 'Medical, Dental & Laboratory Supplies Expenses', 'name': 'med_dental_lab_supplies_expenses', 'indent': True},
                {'label': 'Food Supplies Expenses', 'name': 'food_supplies_expenses', 'indent': True},
                {'label': 'Fuel, Oil and Lubricants Expenses', 'name': 'fuel_oil_lubricants_expenses', 'indent': True},
                {'label': 'Textbooks and Instructional Materials Expenses', 'name': 'textbooks_instructional_materials_expenses', 'indent': True},
                {'label': 'Construction Materials Expenses', 'name': 'construction_material_expenses', 'indent': True},
                {'label': 'Other Supplies & Materials Expenses', 'name': 'other_supplies_materials_expenses', 'indent': True},
                {'label': 'Semi-expendable Machinery Equipment', 'is_group': True},
                {'label': 'Machinery', 'name': 'semee_machinery', 'indent': True},
                {'label': 'Office Equipment', 'name': 'semee_office_equipment', 'indent': True},
                {'label': 'Information and Communications Technology Equipment', 'name': 'semee_information_communication', 'indent': True},
                {'label': 'Communications Equipment', 'name': 'semee_communications_equipment', 'indent': True},
                {'label': 'Disaster Response and Rescue Equipment', 'name': 'semee_drr_equipment', 'indent': True},
                {'label': 'Medical Equipment', 'name': 'semee_medical_equipment', 'indent': True},
                {'label': 'Printing Equipment', 'name': 'semee_printing_equipment', 'indent': True},
                {'label': 'Sports Equipment', 'name': 'semee_sports_equipment', 'indent': True},
                {'label': 'Technical and Scientific Equipment', 'name': 'semee_technical_scientific_equipment', 'indent': True},
                {'label': 'ICT Equipment', 'name': 'semee_ict_equipment', 'indent': True},
                {'label': 'Other Machinery and Equipment', 'name': 'semee_other_machinery_equipment', 'indent': True},
                {'label': 'Semi-expendable Furnitures and Fixtures', 'is_group': True},
                {'label': 'Furniture and Fixtures', 'name': 'furniture_fixtures', 'indent': True},
                {'label': 'Books', 'name': 'books', 'indent': True},
                {'label': 'Utility Expenses', 'is_group': True},
                {'label': 'Water Expenses', 'name': 'water_expenses', 'indent': True},
                {'label': 'Electricity Expenses', 'name': 'electricity_expenses', 'indent': True},
                {'label': 'Communication Expenses', 'is_group': True},
                {'label': 'Postage and Courier Services', 'name': 'postage_courier_services', 'indent': True},
                {'label': 'Telephone Expenses', 'name': 'telephone_expenses', 'indent': True},
                {'label': 'Telephone Expenses (Landline)', 'name': 'telephone_expenses_landline', 'indent': True},
                {'label': 'Internet Subscription Expenses', 'name': 'internet_subscription_expenses', 'indent': True},
                {'label': 'Cable, Satellite, Telegraph & Radio Expenses', 'name': 'cable_satellite_telegraph_radio_expenses', 'indent': True},
                {'label': 'Awards/Rewards and Prizes', 'is_group': True},
                {'label': 'Awards/Rewards Expenses', 'name': 'awards_rewards_expenses', 'indent': True},
                {'label': 'Prizes', 'name': 'prizes', 'indent': True},
                {'label': 'Survey, Research, Exploration, and Development Expenses', 'is_group': True},
                {'label': 'Survey Expenses', 'name': 'survey_expenses', 'indent': True},
                {'label': 'Survey, Research, Exploration, and Development expenses', 'name': 'survey_research_exploration_development_expenses', 'indent': True},
                {'label': 'Professional Services', 'is_group': True},
                {'label': 'Legal Services', 'name': 'legal_services', 'indent': True},
                {'label': 'Auditing Services', 'name': 'auditing_services', 'indent': True},
                {'label': 'Consultancy Services', 'name': 'consultancy_services', 'indent': True},
                {'label': 'Other Professional Services', 'name': 'other_professional_servies', 'indent': True},
                {'label': 'General Services', 'is_group': True},
                {'label': 'Security Services', 'name': 'security_services', 'indent': True},
                {'label': 'Janitorial Services', 'name': 'janitorial_services', 'indent': True},
                {'label': 'Other General Services', 'name': 'other_general_services', 'indent': True},
                {'label': 'Environment/Sanitary Services', 'name': 'environment/sanitary_services', 'indent': True},
                {'label': 'Repair and Maintenance', 'is_group': True},
                {'label': 'Repair & Maintenance - Land Improvements', 'name': 'repair_maintenance_land_improvements', 'indent': True},
                {'label': 'Repair & Maintenance - Buildings and Structures', 'is_group': True},
                {'label': 'Buildings', 'name': 'buildings', 'indent': True},
                {'label': 'School Buildings', 'name': 'school_buildings', 'indent': True},
                {'label': 'Hostels and Dormitories', 'name': 'hostel_dormitories', 'indent': True},
                {'label': 'Other Structures', 'name': 'other_structures', 'indent': True},
                {'label': 'Repairs and Maintenance - Machinery and Equipment', 'is_group': True},
                {'label': 'Machinery', 'name': 'repair_maintenance_machinery', 'indent': True},
                {'label': 'Office Equipment', 'name': 'repair_maintenance_office_equipment', 'indent': True},
                {'label': 'ICT Equipment', 'name': 'repair_maintenance_ict_equipment', 'indent': True},
                {'label': 'Agricultural and Forestry Equipment', 'name': 'repair_maintenance_agri_forestry_equipment', 'indent': True},
                {'label': 'Marine and Fishery Equipment', 'name': 'repair_maintenance_marine_fishery_equipment', 'indent': True},
                {'label': 'Airport Equipment', 'name': 'repair_maintenance_airport_equipment', 'indent': True},
                {'label': 'Communication Equipment', 'name': 'repair_maintenance_communication_equipment', 'indent': True},
                {'label': 'Disaster, Response and Rescue Equipment', 'name': 'repair_maintenance_drre_equipment', 'indent': True},
                {'label': 'Medical Equipment', 'name': 'repair_maintenance_medical_equipment', 'indent': True},
                {'label': 'Printing Equipment', 'name': 'repair_maintenance_printing_equipment', 'indent': True},
                {'label': 'Sports Equipment', 'name': 'repair_maintenance_sports_equipment', 'indent': True},
                {'label': 'Technical and Scientific Equipment', 'name': 'repair_maintenance_technical_scientific_equipment', 'indent': True},
                {'label': 'Other Machinery and Equipment', 'name': 'repair_maintenance_other_machinery_equipment', 'indent': True},
                {'label': 'Repairs and Maintenance - Transportation Equipment', 'is_group': True},
                {'label': 'Motor Vehicles', 'name': 'repair_maintenance_motor', 'indent': True},
                {'label': 'Other Transportation Equipment', 'name': 'repair_maintenance_other_transportation_equipment', 'indent': True},
                {'label': 'Repairs and Maintenance - Furniture & Fixtures', 'name': 'repair_maintenance_furniture_fixtures'},
                {'label': 'Repairs and Maintenance - Semi-Expendable Machinery and Equipment', 'name': 'repair_maintenance_semi_expendable_machinery_equipment'},
                {'label': 'Repairs and Maintenance - Other Property, Plant and Equipment', 'name': 'repair_maintenance_other_property_plant_equipment'},
                {'label': 'Taxes, Insurance Premiums and Other Fees', 'is_group': True},
                {'label': 'Taxes, Duties and Licenses', 'name': 'taxes_duties_licenses', 'indent': True},
                {'label': 'Fidelity Bond Premiums', 'name': 'fidelity_bond_premiums', 'indent': True},
                {'label': 'Insurance Expenses', 'name': 'insurance_expenses', 'indent': True},
                {'label': 'Labor and Wages', 'is_group': True},
                {'label': 'Labor and Wages', 'name': 'labor_wages', 'indent': True},
                {'label': 'Other Maintenance and Operating Expenses', 'is_group': True},
                {'label': 'Advertising Expenses', 'name': 'advertising_expenses', 'indent': True},
                {'label': 'Printing and Publication Expenses', 'name': 'printing_publication_expenses', 'indent': True},
                {'label': 'Representation Expenses', 'name': 'representation_expenses', 'indent': True},
                {'label': 'Transportation and Delivery Expenses', 'name': 'transportation_delivery_expenses', 'indent': True},
                {'label': 'Rent/Lease Expenses', 'name': 'rent/lease_expenses', 'indent': True},
                {'label': 'Membership Dues and contributions to organizations', 'name': 'membership_dues_contribute_to_org', 'indent': True},
                {'label': 'Subscription Expenses', 'name': 'subscription_expenses', 'indent': True},
                {'label': 'Website Maintenance', 'name': 'website_maintenance', 'indent': True},
                {'label': 'Other Maintenance and Operating Expenses', 'name': 'other_maintenance_operating_expenses', 'indent': True},
            ]
        },
        {
            'title': 'Capital Outlays',
            'color_class': 'bg-green-100',
            'items': [
                # Placeholder; extend as needed
                {'label': 'Land', 'is_group': True},
                {'label': 'Land', 'name': 'land', 'indent': True},
                {'label': 'Land Improvements', 'is_group': True},
                {'label': 'Land Improvements, Aquaculture Structure', 'name': 'land_improvements_aqua_structure', 'indent': True},
                {'label': 'Infrastructure Assets', 'is_group': True},
                {'label': 'Water Supply Systems', 'name': 'water_supply_systems', 'indent': True},
                {'label': 'Power Supply Systems', 'name': 'power_supply_systems', 'indent': True},
                {'label': 'Other Infrastructure Assets', 'name': 'other_infra_assets', 'indent': True},
                {'label': 'Building and Other Structures', 'is_group': True},
                {'label': 'Building', 'name': 'bos_building', 'indent': True},
                {'label': 'School Buildings', 'name': 'bos_school_buildings', 'indent': True},
                {'label': 'Hostels and Dormitories', 'name': 'bos_hostels_dorm', 'indent': True},
                {'label': 'Other Structures', 'name': 'other_structures', 'indent': True},
                {'label': 'Machinery and Equipment', 'is_group': True},
                {'label': 'Machinery', 'name': 'me_machinery', 'indent': True},
                {'label': 'Office Equipment', 'name': 'me_office_equipment', 'indent': True},
                {'label': 'Information and Communication Technology Equipment', 'name': 'me_ict_equipment', 'indent': True},
                {'label': 'Communication Equipment', 'name': 'me_communication_equipment', 'indent': True},
                {'label': 'Disaster Response and Rescue Equipment', 'name': 'me_drre', 'indent': True},
                {'label': 'Medical Equipment', 'name': 'me_medical_equipment', 'indent': True},
                {'label': 'Printing Equipment', 'name': 'me_printing_equipment', 'indent': True},
                {'label': 'Sports Equipment', 'name': 'me_sports_equipment', 'indent': True},
                {'label': 'Technical and Scientific Equipment', 'name': 'me_technical_scientific_equipment', 'indent': True},
                {'label': 'Other Machinery and Equipment', 'name': 'me_other_machinery_equipment', 'indent': True},
                {'label': 'Transportation Equipment', 'is_group': True},
                {'label': 'Motor Vehicles', 'name': 'te_motor', 'indent': True},
                {'label': 'Other Transportation Equipment', 'name': 'te_other_transpo_equipment', 'indent': True},
                {'label': 'Furniture, Fixtures and Books', 'is_group': True},
                {'label': 'Furniture and Fixtures', 'name': 'ffb_furniture_fixtures', 'indent': True},
                {'label': 'Books', 'name': 'ffb_books', 'indent': True},
                {'label': 'Construction in Progress', 'is_group': True},
                {'label': 'Construction in Progress - Land Improvements', 'name': 'cp_land_improvements', 'indent': True},
                {'label': 'Construction in Progress - Infrastructure Assets', 'name': 'cp_infra_assets', 'indent': True},
                {'label': 'Construction in Progress - Buildings and Other Structures', 'name': 'cp_building_other_structures', 'indent': True},
                {'label': 'Construction in Progress - Leased Assets', 'name': 'cp_leased_assets', 'indent': True},
                {'label': 'Construction in Progress - Leased Assets Improvements', 'name': 'cp_leased_assets_improvements', 'indent': True},
                {'label': 'Intangible Assets', 'is_group': True},
                {'label': 'Computer Software', 'name': 'ia_computer_software', 'indent': True},
                {'label': 'Websites', 'name': 'ia_websites', 'indent': True},
                {'label': 'Other Tangible Assets', 'name': 'ia_other_tangible_assets', 'indent': True},
                
            ]
        },
    ]

    sections = []
    for spec in sections_spec:
        items = []
        for row in spec['items']:
            if row.get('is_group'):
                items.append({'is_group': True, 'label': row['label']})
            else:
                name = row['name']
                q1 = to_decimal(payload.get(f"{name}_q1"))
                q2 = to_decimal(payload.get(f"{name}_q2"))
                q3 = to_decimal(payload.get(f"{name}_q3"))
                q4 = to_decimal(payload.get(f"{name}_q4"))
                total = q1 + q2 + q3 + q4
                items.append({'label': row['label'], 'indent': row.get('indent', False), 'q1': q1, 'q2': q2, 'q3': q3, 'q4': q4, 'total': total})
        from decimal import Decimal as _D
        total_q1 = sum((it['q1'] for it in items if not it.get('is_group')), _D('0'))
        total_q2 = sum((it['q2'] for it in items if not it.get('is_group')), _D('0'))
        total_q3 = sum((it['q3'] for it in items if not it.get('is_group')), _D('0'))
        total_q4 = sum((it['q4'] for it in items if not it.get('is_group')), _D('0'))
        total_overall = sum((it['total'] for it in items if not it.get('is_group')), _D('0'))
        sections.append({'title': spec['title'], 'color_class': spec['color_class'], 'items': items, 'total_q1': total_q1, 'total_q2': total_q2, 'total_q3': total_q3, 'total_q4': total_q4, 'total_overall': total_overall})

    from decimal import Decimal as _DD
    grand_total_q1 = sum((sec['total_q1'] for sec in sections), _DD('0'))
    grand_total_q2 = sum((sec['total_q2'] for sec in sections), _DD('0'))
    grand_total_q3 = sum((sec['total_q3'] for sec in sections), _DD('0'))
    grand_total_q4 = sum((sec['total_q4'] for sec in sections), _DD('0'))
    grand_total_overall = sum((sec['total_overall'] for sec in sections), _DD('0'))

    return render(request, 'admin_panel/preview_pre.html', {
        'pre': pre,
        'sections': sections,
        'grand_total_q1': grand_total_q1,
        'grand_total_q2': grand_total_q2,
        'grand_total_q3': grand_total_q3,
        'grand_total_q4': grand_total_q4,
        'grand_total_overall': grand_total_overall,
        'prepared_by': pre.submitted_by.fullname,
        'certified_by': pre.certified_by_name,
        'approved_by': pre.approved_by_name,
    })


@role_required('admin', login_url='/admin/')
def admin_handle_pre_action(request, pk: int):
    pre = get_object_or_404(DepartmentPRE, pk=pk)
    budget_allocation = get_object_or_404(BudgetAllocation, id=pre.budget_allocation_id)
    # budget_allocation = BudgetAllocation.objects.filter(department=pre.department, id=pre.budget_allocation_id).first
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            pre.status = 'Partially Approved'
            pre.approved_by_admin = True
            audit_trail_action = 'Partially APPROVE'
        elif action == 'reject':
            pre.status = 'Rejected'
            pre.approved_by_admin = False
            budget_allocation.is_compiled = False
            budget_allocation.save(update_fields=['is_compiled'])
            audit_trail_action = 'REJECT'
        pre.save()
        log_audit_trail(
            request=request,
            action=audit_trail_action,
            model_name='DepartmentPRE',
            record_id=pre.id,
            detail=f"Admin {action} a PRE with ID {pre.id} from department {pre.department}."
        )
    return redirect('pre_request_page')

@role_required('admin', login_url='/admin/')
def preview_purchase_request(request, pk:int):
    pr = get_object_or_404(PurchaseRequest.objects.select_related('requested_by', 'budget_allocation__approved_budget', 'source_pre'),
        pk=pk,
    )
    
    return render(request, 'admin_panel/preview_purchase_request.html', {
        'pr': pr,
    })
    
@role_required('admin', login_url='/admin/')
def departments_ad_request(request):
    """
    Enhanced admin view for Activity Design requests with status counts and filtering
    """
    from apps.budgets.models import ActivityDesign

    STATUS_CHOICES = (
        ('Pending', 'Pending'),
        ('Partially Approved', 'Partially Approved'),
        ('Approved', 'Approved'),
        ('Rejected', 'Rejected')
    )

    # Get all Activity Designs with related data
    ads = ActivityDesign.objects.select_related(
        'submitted_by',
        'budget_allocation',
        'budget_allocation__approved_budget'
    ).prefetch_related(
        'pre_allocations',
        'pre_allocations__pre_line_item'
    ).order_by('-submitted_at')

    # Calculate status counts before filtering
    status_counts = {
        'total': ads.count(),
        'pending': ads.filter(status='Pending').count(),
        'partially_approved': ads.filter(status='Partially Approved').count(),
        'approved': ads.filter(status='Approved').count(),
        'rejected': ads.filter(status='Rejected').count(),
    }

    # Get distinct departments
    departments = (
        User.objects.filter(is_staff=False, is_approving_officer=False)
        .exclude(department__isnull=True)
        .values_list('department', flat=True)
        .distinct()
        .order_by('department')
    )

    # Apply filters
    department_filter = request.GET.get('department')
    if department_filter:
        ads = ads.filter(department=department_filter)

    status_filter = request.GET.get('status')
    if status_filter:
        ads = ads.filter(status=status_filter)

    context = {
        'ads': ads,
        'status_counts': status_counts,
        'departments': departments,
        'selected_department': department_filter,
        'status_choices': STATUS_CHOICES,
    }
    return render(request, 'admin_panel/departments_ad_request.html', context)

@role_required('admin', login_url='/admin/')
def admin_preview_activity_design(request, pk: int):
    from apps.budgets.models import ActivityDesign
    activity_design = get_object_or_404(ActivityDesign, id=pk)
    return render(request, 'admin_panel/preview_activity_design.html', {'activity': activity_design})

@role_required('admin', login_url='/admin/')
def handle_activity_design_request(request, pk):
    """
    View for Admin to approve or reject department activity designs.
    Updates the status of the activity design and associated budget allocation.
    Accepts UUID as pk parameter.
    """
    from apps.budgets.models import ActivityDesign

    activity_design = get_object_or_404(ActivityDesign, id=pk)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'approve':
            if activity_design.status != 'Pending':
                messages.warning(request,
                    f"AD {activity_design.ad_number} cannot be approved. "
                    f"Current status: {activity_design.status}")
                return redirect('departments_ad_request')

            try:
                # Get allocations info for detailed feedback
                allocations = activity_design.pre_allocations.all()
                line_items_info = ", ".join([
                    f"{alloc.pre_line_item.item_name} ({alloc.quarter})"
                    for alloc in allocations[:3]
                ])
                if allocations.count() > 3:
                    line_items_info += f" and {allocations.count() - 3} more"

                # 1. Update status to Partially Approved
                activity_design.status = 'Partially Approved'
                activity_design.partially_approved_at = timezone.now()
                activity_design.save(update_fields=['status', 'partially_approved_at', 'updated_at'])

                # 2. 🔥 AUTO-CONVERT DOCUMENTS TO PDF
                from .document_converter import convert_ad_documents_to_pdf
                print(f"🔄 Starting document conversion for AD {activity_design.ad_number}")
                conversion_results = convert_ad_documents_to_pdf(activity_design)

                # 3. Check conversion results and show appropriate messages
                if conversion_results['main_document']:
                    if conversion_results['main_document_format'] == 'PDF':
                        messages.success(request,
                            f'✅ AD {activity_design.ad_number} partially approved!')
                        messages.info(request,
                            f'📄 Main document converted to PDF successfully.')
                    else:
                        messages.success(request,
                            f'✅ AD {activity_design.ad_number} partially approved!')
                        messages.info(request,
                            f'📄 Main document already in PDF format.')
                else:
                    # Conversion completely failed
                    messages.success(request,
                        f'✅ AD {activity_design.ad_number} partially approved!')
                    messages.warning(request,
                        f'⚠️ Document conversion failed. Original files are available for download. Manual PDF conversion may be needed.')

                # Show specific warnings (only if there are issues)
                if conversion_results['warnings']:
                    for warning in conversion_results['warnings']:
                        messages.warning(request, warning)

                # Show errors (if any critical errors)
                if conversion_results['errors']:
                    for error in conversion_results['errors']:
                        messages.error(request, f"❌ {error}")

                # 4. Log audit trail
                log_audit_trail(
                    request=request,
                    action='APPROVE',
                    model_name='ActivityDesign',
                    record_id=activity_design.id,
                    detail=f'Activity Design {activity_design.ad_number} partially approved by Admin. '
                           f'Amount: ₱{activity_design.total_amount:,.2f} from {allocations.count()} line items: {line_items_info}. '
                           f'Documents converted: Main={conversion_results["main_document"]}, '
                           f'Supporting={len(conversion_results["supporting_docs"])}'
                )

                # 5. Create notification for end user
                from apps.budgets.models import SystemNotification
                SystemNotification.objects.create(
                    recipient=activity_design.submitted_by,
                    title='AD Partially Approved - Ready to Print',
                    message=f'Your AD {activity_design.ad_number} has been partially approved. '
                            f'Download the PDF documents, print them, and get them signed by the Approving Officer.',
                    content_type='ad',
                    object_id=activity_design.id
                )

                messages.info(request,
                    f'💰 Total Amount: ₱{activity_design.total_amount:,.2f} from {allocations.count()} line item(s)')

            except Exception as e:
                print(f"❌ Error in approval process: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error processing approval: {str(e)}")
                return redirect('departments_ad_request')
        elif action == 'reject':
            try:
                from apps.budgets.models import ActivityDesignAllocation

                # Get allocations
                allocations = ActivityDesignAllocation.objects.filter(
                    activity_design=activity_design
                ).select_related('pre_line_item')

                if not allocations.exists():
                    messages.warning(request, "No allocations found to reverse.")
                    return redirect('departments_ad_request')

                total_released = Decimal('0')
                released_details = []

                # Reverse allocations
                for ad_allocation in allocations:
                    line_item = ad_allocation.pre_line_item
                    allocated_amount = ad_allocation.allocated_amount
                    quarter = ad_allocation.quarter

                    # ✅ NO NEED to update line_item - consumption is calculated dynamically!
                    # Just track for logging
                    total_released += allocated_amount
                    released_details.append(
                        f"{line_item.item_name} {quarter}: ₱{allocated_amount:,.2f}"
                    )

                    print(f"✅ Will release ₱{allocated_amount} from {line_item.item_name} {quarter}")

                # ✅ Delete allocations - this automatically "releases" the budget
                allocations.delete()

                # Update budget allocation tracking
                if activity_design.budget_allocation:
                    activity_design.budget_allocation.ad_amount_used = max(
                        Decimal('0'),
                        activity_design.budget_allocation.ad_amount_used - total_released
                    )
                    activity_design.budget_allocation.update_remaining_balance()
                    activity_design.budget_allocation.save()

                print(f"✅ Deleted {len(released_details)} allocation(s), budget automatically released")

            except Exception as e:
                print(f"❌ ERROR: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error processing rejection: {str(e)}")
                return redirect('departments_ad_request')

            # Update AD status
            activity_design.status = 'Rejected'

            # Log audit trail
            log_audit_trail(
                request=request,
                action='REJECT',
                model_name='ActivityDesign',
                record_id=activity_design.id,
                detail=f'Activity Design {activity_design.ad_number} rejected by Admin. '
                       f'Released ₱{total_released:,.2f} back to budget. '
                       f'Details: {", ".join(released_details)}'
            )

            # Create notification
            from apps.budgets.models import SystemNotification
            SystemNotification.objects.create(
                recipient=activity_design.submitted_by,
                title='AD Rejected',
                message=f'Your AD {activity_design.ad_number} has been rejected. '
                        f'₱{total_released:,.2f} has been returned to your budget.',
                content_type='ad',
                object_id=activity_design.id
            )

            messages.success(request,
                f'❌ Activity Design {activity_design.ad_number} has been rejected.')
            messages.info(request,
                f'💰 Released ₱{total_released:,.2f} back to budget')

            # Show detailed breakdown if multiple allocations
            if len(released_details) > 1:
                for detail in released_details:
                    messages.info(request, f"  • {detail}")

        activity_design.save(update_fields=['status', 'updated_at'])

    return redirect('departments_ad_request')


@role_required('admin', login_url='/admin/')
def pre_budget_realignment_admin(request):
    """Admin view for PRE budget realignment requests"""
    
    # Get all realignment requests
    realignment_requests = PREBudgetRealignment.objects.select_related(
        'requested_by', 'approved_by', 'source_pre', 'target_pre'
    ).order_by('-created_at')
    
    status_counts = {
        'total': realignment_requests.count(),
        'pending': realignment_requests.filter(status='Pending').count(),
        'approved': realignment_requests.filter(status='Approved').count(),
        'partially_approved': realignment_requests.filter(status='Partially Approved').count(),
        'rejected': realignment_requests.filter(status='Rejected').count(),
    }
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        realignment_requests = realignment_requests.filter(status=status_filter)
    
    context = {
        'realignment_requests': realignment_requests,
        'status_filter': status_filter,
        'status_choices': PREBudgetRealignment.STATUS_CHOICES,
        'status_counts': status_counts,
    }
    
    return render(request, 'admin_panel/pre_budget_realignment.html', context)

@role_required('admin', login_url='/admin/')
def handle_pre_realignment_admin_action(request, pk):
    """Handle approve/reject actions for PRE realignment requests"""
    
    realignment = get_object_or_404(PREBudgetRealignment, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'partial_approve':
            realignment.status='Partially Approved'
            realignment.approved_by_admin = True
            realignment.partial_approved_by = request.user
            realignment.save()
            
            # Log audit trail
            log_audit_trail(
                request=request,
                action='PARTIAL_APPROVE',
                model_name='PREBudgetRealignment',
                record_id=realignment.id,
                detail=f'Partially approved budget realignment: {realignment.source_item_display} → {realignment.target_item_display} (₱{realignment.amount:,.2f})',
            )
            
            messages.success(request, f'Budget realignment partially approved.')
        
        elif action == 'reject':
            realignment.status = 'Rejected'
            realignment.approved_by = request.user
            realignment.save()
            
            # Log audit trail
            log_audit_trail(
                request=request,
                action='REJECT',
                model_name='PREBudgetRealignment',
                record_id=realignment.id,
                detail=f'Rejected budget realignment: {realignment.source_item_display} → {realignment.target_item_display} (₱{realignment.amount:,.2f})',
            )
            
            messages.success(request, 'Budget realignment rejected.')
    
    return redirect('pre_budget_realignment_admin')

@role_required('admin', login_url='/admin/')
def admin_pre_list(request):
    """
    Admin view to list all PRE submissions with filters, search, and pagination
    """
    # Check if user is admin/staff
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')
    
    # Base queryset - exclude drafts
    pres = NewDepartmentPRE.objects.exclude(status='Draft').select_related(
        'submitted_by', 
        'budget_allocation__approved_budget'
    ).order_by('-created_at')
    
    # === FILTERS ===
    
    # 1. Search Filter (PRE ID or Employee Name)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        pres = pres.filter(
            Q(id__icontains=search_query) |
            Q(submitted_by__fullname__icontains=search_query) |
            Q(submitted_by__username__icontains=search_query) |
            Q(submitted_by__email__icontains=search_query) |
            Q(department__icontains=search_query)
        )
    
    # 2. Department Filter
    department_filter = request.GET.get('department', '').strip()
    if department_filter:
        pres = pres.filter(department=department_filter)
    
    # 3. Status Filter
    status_filter = request.GET.get('status', '').strip()
    if status_filter:
        pres = pres.filter(status=status_filter)
    
    # 4. Date Range Filter
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            pres = pres.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            messages.warning(request, "Invalid 'From Date' format.")
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            pres = pres.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            messages.warning(request, "Invalid 'To Date' format.")
    
    # === STATISTICS ===
    # Calculate stats for all PREs (excluding drafts)
    all_pres = NewDepartmentPRE.objects.exclude(status='Draft')
    stats = {
        'total': all_pres.count(),
        'pending': all_pres.filter(status='Pending').count(),
        'approved': all_pres.filter(status='Approved').count(),
        'rejected': all_pres.filter(status='Rejected').count(),
    }
    
    # === PAGINATION ===
    paginator = Paginator(pres, 10)  # Show 10 PREs per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # === GET DEPARTMENTS FOR FILTER ===
    departments = NewDepartmentPRE.objects.exclude(
        status='Draft'
    ).values_list('department', flat=True).distinct().order_by('department')
    
    # === STATUS CHOICES ===
    status_choices = NewDepartmentPRE.STATUS_CHOICES
    
    context = {
        'pres': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'stats': stats,
        'departments': departments,
        'status_choices': status_choices,
    }
    
    return render(request, 'admin_panel/pre_list.html', context)

@role_required('admin', login_url='/admin/')
def admin_pre_detail(request, pre_id):
    """
    Admin view to see detailed PRE information with budget tracking breakdown
    """
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')

    pre = get_object_or_404(
        NewDepartmentPRE.objects.select_related(
            'submitted_by',
            'budget_allocation__approved_budget'
        ).prefetch_related(
            'line_items__category',
            'line_items__subcategory',
            'receipts'
        ),
        id=pre_id
    )

    # Get approval history
    approval_history = RequestApproval.objects.filter(
        content_type='pre',
        object_id=pre.id
    ).select_related('approved_by').order_by('-approved_at')

    # Calculate totals by category
    from django.db.models import Sum
    category_totals = pre.line_items.values(
        'category__name'
    ).annotate(
        total=Sum('q1_amount') + Sum('q2_amount') + Sum('q3_amount') + Sum('q4_amount')
    ).order_by('category__sort_order')

    # Calculate budget tracking breakdown for each line item
    line_items_with_breakdown = []
    for item in pre.line_items.all():
        item_data = {
            'item': item,
            'quarters': []
        }

        for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
            breakdown = item.get_quarter_breakdown(quarter)
            item_data['quarters'].append(breakdown)

        line_items_with_breakdown.append(item_data)

    context = {
        'pre': pre,
        'approval_history': approval_history,
        'category_totals': category_totals,
        'line_items_with_breakdown': line_items_with_breakdown,
    }

    return render(request, 'admin_panel/pre_detail.html', context)

@role_required('admin', login_url='/admin/')
def admin_handle_pre_action(request, pre_id):
    """
    Handle Approve/Reject actions for PRE submissions
    This is a basic implementation - will be enhanced with modals in Component 4
    """
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('dashboard')
    
    if request.method != 'POST':
        return redirect('admin_pre_list')
    
    pre = get_object_or_404(NewDepartmentPRE, id=pre_id)
    action = request.POST.get('action')
    
    if action == 'approve':
        if pre.status == 'Pending':
            pre.status = 'Partially Approved'
            pre.partially_approved_at = timezone.now()
            pre.save()
            
            # Create approval record
            RequestApproval.objects.create(
                content_type='pre',
                object_id=pre.id,
                approved_by=request.user,
                approval_level='partial',
                comments=request.POST.get('comment', '')
            )
            
            # Create notification
            SystemNotification.objects.create(
                recipient=pre.submitted_by,
                title='PRE Partially Approved',
                message=f'Your PRE for {pre.department} has been partially approved.',
                content_type='pre',
                object_id=pre.id
            )
            
            messages.success(request, f'PRE {str(pre.id)[:8]} has been partially approved. Please generate PDF for printing.')
            
            # TODO: Trigger PDF generation
            # generate_pre_pdf(pre)
        else:
            messages.warning(request, 'This PRE cannot be approved.')
    
    elif action == 'reject':
        if pre.status == 'Pending':
            pre.status = 'Rejected'
            pre.rejection_reason = request.POST.get('reason', 'No reason provided')
            pre.save()
            
            # Create approval record
            RequestApproval.objects.create(
                content_type='pre',
                object_id=pre.id,
                approved_by=request.user,
                approval_level='rejected',
                comments=pre.rejection_reason
            )
            
            # Create notification
            SystemNotification.objects.create(
                recipient=pre.submitted_by,
                title='PRE Rejected',
                message=f'Your PRE for {pre.department} has been rejected. Reason: {pre.rejection_reason}',
                content_type='pre',
                object_id=pre.id
            )
            
            messages.success(request, f'PRE {str(pre.id)[:8]} has been rejected.')
        else:
            messages.warning(request, 'This PRE cannot be rejected.')
    
    else:
        messages.error(request, 'Invalid action.')
    
    return redirect('admin_pre_list')

@role_required('admin', login_url='/admin/')
def admin_approve_pre_with_comment(request, pre_id):
    """
    Advanced approve with comment (AJAX endpoint for modal)
    Auto-generates PDF when approving
    """
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    pre = get_object_or_404(NewDepartmentPRE, id=pre_id)  # Use your model name
    
    if pre.status != 'Pending':
        return JsonResponse({
            'success': False, 
            'error': f'This PRE cannot be approved. Current status: {pre.status}'
        }, status=400)
    
    # Get comment from request
    comment = request.POST.get('comment', '').strip()
    
    # Update PRE to Partially Approved
    pre.status = 'Partially Approved'
    pre.partially_approved_at = timezone.now()
    pre.admin_notes = comment
    pre.save()
    
    # 🔥 THIS IS THE KEY PART - Auto-generate PDF
    # try:
    #     from .pdf_generator import save_pre_pdf
    #     pdf_url = save_pre_pdf(pre)
    #     print(f"✅ PDF generated successfully: {pdf_url}")
    # except Exception as e:
    #     # If PDF generation fails, still approve but log error
    #     print(f"❌ PDF generation failed: {str(e)}")
    #     pdf_url = None
    
    from .excel_to_pdf_converter import generate_pre_pdf_from_excel

    # Auto-generate PDF from Excel
    try:
        pdf_url = generate_pre_pdf_from_excel(pre)
        if pdf_url:
            print(f"✅ PDF generated successfully: {pdf_url}")
        else:
            print("⚠️ Auto-conversion failed - manual upload needed")
    except Exception as e:
        print(f"❌ PDF generation failed: {str(e)}")
        pdf_url = None
    
    # Create approval record
    RequestApproval.objects.create(
        content_type='pre',
        object_id=pre.id,
        approved_by=request.user,
        approval_level='partial',
        comments=comment
    )
    
    # Create notification for submitter
    SystemNotification.objects.create(
        recipient=pre.submitted_by,
        title='PRE Partially Approved - Ready to Print',
        message=f'Your PRE for {pre.department} has been partially approved. You can now download and print the PDF document.',
        content_type='pre',
        object_id=pre.id
    )
    
    return JsonResponse({
        'success': True,
        'message': f'PRE approved and PDF generated successfully',
        'pre_id': str(pre.id),
        'new_status': pre.status,
        'pdf_generated': pdf_url is not None  # Include this info
    })

@role_required('admin', login_url='/admin/')
def admin_upload_manual_pdf(request, pre_id):
    """
    Admin manually uploads converted PDF
    """
    if not request.user.is_staff:
        messages.error(request, "Permission denied")
        return redirect('admin_pre_list')
    
    if request.method != 'POST':
        return redirect('admin_pre_detail', pre_id=pre_id)
    
    pre = get_object_or_404(NewDepartmentPRE, id=pre_id)
    
    if 'manual_pdf' not in request.FILES:
        messages.error(request, "No PDF file uploaded")
        return redirect('admin_pre_detail', pre_id=pre_id)
    
    pdf_file = request.FILES['manual_pdf']
    
    # Validate file
    if not pdf_file.name.endswith('.pdf'):
        messages.error(request, "Only PDF files are allowed")
        return redirect('admin_pre_detail', pre_id=pre_id)
    
    try:
        from .excel_to_pdf_converter import enable_manual_pdf_upload
        
        pdf_url = enable_manual_pdf_upload(pre, pdf_file)
        
        if pdf_url:
            messages.success(request, "PDF uploaded successfully! Users can now download it.")
        else:
            messages.error(request, "Failed to upload PDF")
            
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('admin_pre_detail', pre_id=pre_id)

@role_required('admin', login_url='/admin/')
def admin_reject_pre_with_reason(request, pre_id):
    """
    Advanced reject with reason (AJAX endpoint for modal)
    """
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    pre = get_object_or_404(NewDepartmentPRE, id=pre_id)
    
    if pre.status != 'Pending':
        return JsonResponse({
            'success': False, 
            'error': f'This PRE cannot be rejected. Current status: {pre.status}'
        }, status=400)
    
    # Get reason from request
    reason = request.POST.get('reason', '').strip()
    
    if not reason:
        return JsonResponse({
            'success': False, 
            'error': 'Rejection reason is required'
        }, status=400)
    
    # Update PRE
    pre.status = 'Rejected'
    pre.rejection_reason = reason
    pre.save()
    
    # Create approval record
    RequestApproval.objects.create(
        content_type='pre',
        object_id=pre.id,
        approved_by=request.user,
        approval_level='rejected',
        comments=reason
    )
    
    # Create notification
    SystemNotification.objects.create(
        recipient=pre.submitted_by,
        title='PRE Rejected',
        message=f'Your PRE for {pre.department} has been rejected. Reason: {reason}',
        content_type='pre',
        object_id=pre.id
    )
    
    return JsonResponse({
        'success': True,
        'message': f'PRE {str(pre.id)[:8]} rejected',
        'pre_id': str(pre.id),
        'new_status': pre.status
    })


@role_required('admin', login_url='/admin/')
def admin_update_pre_status(request, pre_id):
    """
    General status update endpoint
    """
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    pre = get_object_or_404(NewDepartmentPRE, id=pre_id)
    new_status = request.POST.get('status')
    comment = request.POST.get('comment', '')
    
    # Valid status transitions
    valid_statuses = dict(NewDepartmentPRE.STATUS_CHOICES).keys()
    
    if new_status not in valid_statuses:
        return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)
    
    old_status = pre.status
    pre.status = new_status
    
    # Update timestamps based on status
    if new_status == 'Partially Approved':
        pre.partially_approved_at = timezone.now()
    elif new_status == 'Approved':
        pre.final_approved_at = timezone.now()
    
    pre.save()
    
    # Create status change record
    RequestApproval.objects.create(
        content_type='pre',
        object_id=pre.id,
        approved_by=request.user,
        approval_level='status_change',
        comments=f'Status changed from {old_status} to {new_status}. {comment}'
    )
    
    return JsonResponse({
        'success': True,
        'message': f'Status updated from {old_status} to {new_status}',
        'new_status': new_status
    })
    
@role_required('admin', login_url='/admin/')
def admin_preview_pre(request, pre_id):
    """
    Backward compatibility - redirects to detail view
    """
    return redirect('admin_pre_detail', pre_id=pre_id)


# === PLACEHOLDER FUNCTIONS (Will implement in later components) ===

@role_required('admin', login_url='/admin/')
def admin_generate_pre_pdf(request, pre_id):
    """
    Generate PDF for partially approved PRE
    """
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('dashboard')
    
    pre = get_object_or_404(NewDepartmentPRE, id=pre_id)
    
    # Check if PRE is in correct status
    if pre.status not in ['Partially Approved', 'Approved']:
        messages.error(request, "PDF can only be generated for partially approved or approved PREs.")
        return redirect('admin_pre_detail', pre_id=pre_id)
    
    try:
        # Import PDF generator
        from .pdf_generator import save_pre_pdf
        
        # Generate and save PDF
        pdf_url = save_pre_pdf(pre)
        
        messages.success(request, f'PDF generated successfully! Click below to download.')
        
        # Redirect to detail page with success message
        return redirect('admin_pre_detail', pre_id=pre_id)
        
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('admin_pre_detail', pre_id=pre_id)


@role_required('admin', login_url='/admin/')
def admin_upload_signed_copy(request, pre_id):
    """
    Upload scanned signed copy of PRE
    """
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('dashboard')
    
    pre = get_object_or_404(NewDepartmentPRE, id=pre_id)
    
    if request.method != 'POST':
        return redirect('admin_pre_detail', pre_id=pre_id)
    
    # Check if file was uploaded
    if 'signed_copy' not in request.FILES:
        messages.error(request, "No file was uploaded.")
        return redirect('admin_pre_detail', pre_id=pre_id)
    
    signed_file = request.FILES['signed_copy']
    
    # Validate file extension
    allowed_extensions = ['pdf', 'jpg', 'jpeg', 'png']
    file_ext = signed_file.name.split('.')[-1].lower()
    
    if file_ext not in allowed_extensions:
        messages.error(request, f"Invalid file type. Allowed: {', '.join(allowed_extensions)}")
        return redirect('admin_pre_detail', pre_id=pre_id)
    
    # Validate file size (max 10MB)
    if signed_file.size > 10 * 1024 * 1024:
        messages.error(request, "File size must be less than 10MB.")
        return redirect('admin_pre_detail', pre_id=pre_id)
    
    try:
        # Save the signed copy
        pre.final_approved_scan = signed_file
        pre.status = 'Approved'
        pre.final_approved_at = timezone.now()
        pre.save()
        
        # Create approval record
        RequestApproval.objects.create(
            content_type='pre',
            object_id=pre.id,
            approved_by=request.user,
            approval_level='final',
            comments='Signed copy uploaded and marked as fully approved'
        )
        
        # Create notification
        SystemNotification.objects.create(
            recipient=pre.submitted_by,
            title='PRE Fully Approved',
            message=f'Your PRE for {pre.department} has been fully approved. The signed copy has been uploaded.',
            content_type='pre',
            object_id=pre.id
        )
        
        # Update budget allocation
        if pre.budget_allocation:
            pre.budget_allocation.pre_amount_used += pre.total_amount
            pre.budget_allocation.update_remaining_balance()
        
        messages.success(request, 'Signed copy uploaded successfully! PRE is now fully approved.')
        
    except Exception as e:
        messages.error(request, f'Error uploading file: {str(e)}')
    
    return redirect('admin_pre_detail', pre_id=pre_id)


@role_required('admin', login_url='/admin/')
def admin_download_pre(request, pre_id):
    """
    Download PRE document
    """
    pre = get_object_or_404(NewDepartmentPRE, id=pre_id)
    
    if pre.uploaded_excel_file:
        from django.http import FileResponse
        return FileResponse(pre.uploaded_excel_file.open('rb'), 
                          as_attachment=True,
                          filename=f'PRE_{pre.department}_{pre.id}.xlsx')
    else:
        messages.error(request, "No file available for download")
        return redirect('admin_pre_detail', pre_id=pre_id)


@role_required('admin', login_url='/admin/')
def admin_upload_approved_document(request, pre_id):
    """
    Admin uploads scanned approved document and marks PRE as fully approved
    """
    pre = get_object_or_404(
        NewDepartmentPRE.objects.select_related('submitted_by', 'budget_allocation'),
        id=pre_id
    )
    
    # Check if PRE is in correct status
    if pre.status != 'Partially Approved':
        messages.error(request, f'Cannot upload documents. PRE status is "{pre.status}". Only "Partially Approved" PREs can receive documents.')
        return redirect('admin_pre_detail', pre_id=pre_id)
    
    if request.method == 'POST':
        form = ApprovedDocumentUploadForm(request.POST, request.FILES, instance=pre)
        
        if form.is_valid():
            try:
                # Save uploaded file
                pre = form.save(commit=False)
                
                # Mark as approved with admin details
                pre.approve_with_documents(request.user)
                
                # Create system notification
                from apps.budgets.models import SystemNotification
                SystemNotification.objects.create(
                    recipient=pre.submitted_by,
                    title='PRE Fully Approved',
                    message=f'Your PRE for {pre.department} has been fully approved with signed documents.',
                    content_type='pre',
                    object_id=pre.id
                )
                
                # Log audit trail
                log_audit_trail(
                    request=request,
                    action='APPROVE',
                    model_name='DepartmentPRE',
                    record_id=str(pre.id),
                    detail=f'PRE {str(pre.id)[:8]} fully approved with uploaded documents by {request.user.username}'
                )
                
                messages.success(
                    request,
                    f'✅ PRE {str(pre.id)[:8]} has been fully approved! '
                    f'Document uploaded successfully and end user can now use line items for PR/AD.'
                )
                
                return redirect('admin_pre_detail', pre_id=pre_id)
                
            except Exception as e:
                messages.error(request, f'Error uploading document: {str(e)}')
                return redirect('admin_pre_detail', pre_id=pre_id)
        else:
            # Form validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    
    else:
        form = ApprovedDocumentUploadForm(instance=pre)
    
    context = {
        'pre': pre,
        'form': form,
    }
    
    return render(request, 'admin_panel/upload_approved_document.html', context)


@role_required('admin', login_url='/admin/')
@require_http_methods(["POST"])
def admin_quick_approve_with_upload(request, pre_id):
    """
    AJAX endpoint for quick approval with document upload
    """
    pre = get_object_or_404(NewDepartmentPRE, id=pre_id)
    
    if pre.status != 'Partially Approved':
        return JsonResponse({
            'success': False,
            'error': f'PRE must be Partially Approved first. Current status: {pre.status}'
        }, status=400)
    
    uploaded_file = request.FILES.get('approved_document')
    
    if not uploaded_file:
        return JsonResponse({
            'success': False,
            'error': 'No file uploaded'
        }, status=400)
    
    # Validate file
    max_size = 10 * 1024 * 1024  # 10MB
    if uploaded_file.size > max_size:
        return JsonResponse({
            'success': False,
            'error': 'File size exceeds 10MB limit'
        }, status=400)
    
    try:
        # Save file and approve
        pre.approved_documents = uploaded_file
        pre.approve_with_documents(request.user)
        
        # Notification
        from apps.budgets.models import SystemNotification
        SystemNotification.objects.create(
            recipient=pre.submitted_by,
            title='PRE Fully Approved',
            message=f'Your PRE for {pre.department} has been fully approved.',
            content_type='pre',
            object_id=pre.id
        )
        
        # Audit log
        log_audit_trail(
            request=request,
            action='APPROVE',
            model_name='DepartmentPRE',
            record_id=str(pre.id),
            detail=f'PRE fully approved with uploaded document'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'PRE {str(pre.id)[:8]} fully approved!',
            'new_status': pre.status,
            'redirect_url': f'/admin/pre/{pre.id}/'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)